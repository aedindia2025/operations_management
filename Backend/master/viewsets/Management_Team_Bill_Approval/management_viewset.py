import io

import openpyxl
from django.db import connection
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from master.apps.Management_Team_Bill_Approval.management_model import ManagementApproval
from master.serializers.Management_Team_Bill_Approval.management_serializer import ManagementApprovalSerializer


_COLUMN_CACHE = {}


def db_fetch(sql, params=None):
    with connection.cursor() as cursor:
        cursor.execute(sql, params or [])
        cols = [c[0] for c in cursor.description]
        return [dict(zip(cols, row)) for row in cursor.fetchall()]


def db_fetchone(sql, params=None):
    rows = db_fetch(sql, params)
    return rows[0] if rows else None


def table_columns(table_name):
    cached = _COLUMN_CACHE.get(table_name)
    if cached is not None:
        return cached
    try:
        with connection.cursor() as cursor:
            cursor.execute(f"SHOW COLUMNS FROM `{table_name}`")
            cols = {row[0] for row in cursor.fetchall()}
    except Exception:
        cols = set()
    _COLUMN_CACHE[table_name] = cols
    return cols


def pick_col(columns, *candidates):
    for col in candidates:
        if col and col in columns:
            return col
    return None


def safe_str(value):
    return "" if value is None else str(value)


def safe_float(value):
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def safe_int(value):
    try:
        return int(value or 0)
    except Exception:
        return 0


def fmt_date_expr(col, alias, fmt="%d-%m-%Y"):
    if not col:
        return f"'' AS {alias}"
    return f"DATE_FORMAT({col}, '{fmt.replace('%', '%%')}') AS {alias}"


def fmt_max_date_expr(col, alias, fmt="%d-%m-%Y"):
    if not col:
        return f"'' AS {alias}"
    return f"DATE_FORMAT(MAX({col}), '{fmt.replace('%', '%%')}') AS {alias}"


def name_expr(col, alias):
    if not col:
        return f"'' AS {alias}"
    return (
        f"CASE WHEN MAX({col}) IS NULL OR MAX({col}) = '' THEN '' "
        f"ELSE COALESCE(get_staff_name(MAX({col})), CAST(MAX({col}) AS CHAR), '') END AS {alias}"
    )


def status_text(sts):
    sts = str(sts or "0")
    if sts == "1":
        return "Approved"
    if sts == "2":
        return "Rejected"
    return "Pending"


def payment_status_text(sts):
    return "Paid" if str(sts or "0") == "1" else "Pending"


def get_user_id(request):
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        return auth_header.split(" ", 1)[1].strip()
    return safe_str(request.data.get("user_id"))


class ManagementApprovalView(APIView):
    def post(self, request):
        action = request.data.get("action")
        if action == "datatable":
            data = ManagementApproval.objects.filter(approval_status="pending").order_by("-id")
            return Response({"data": ManagementApprovalSerializer(data, many=True).data})
        if action == "approved_list":
            data = ManagementApproval.objects.filter(approval_status="approved")
            return Response({"data": ManagementApprovalSerializer(data, many=True).data})
        if action == "rejected_list":
            data = ManagementApproval.objects.filter(approval_status="rejected")
            return Response({"data": ManagementApprovalSerializer(data, many=True).data})
        if action == "approve":
            obj = ManagementApproval.objects.get(id=request.data.get("unique_id"))
            obj.approval_status = "approved"
            obj.approved_by = request.data.get("approved_by")
            obj.approved_at = request.data.get("approved_at")
            obj.save()
            return Response({"status": True, "msg": "Approved Successfully"})
        if action == "reject":
            obj = ManagementApproval.objects.get(id=request.data.get("unique_id"))
            obj.approval_status = "rejected"
            obj.reject_reason = request.data.get("reject_reason")
            obj.save()
            return Response({"status": True, "msg": "Rejected Successfully"})
        if action == "create":
            serializer = ManagementApprovalSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response({"status": True, "msg": "Created"})
            return Response({"status": False, "error": serializer.errors})
        return Response({"status": False, "msg": "Invalid Action"})


def list_rows(tab, search="", from_date="", to_date="", page=1, length=10):
    vpd_cols = table_columns("vendor_payment_details")
    if not vpd_cols:
        return [], 0

    bill_date = pick_col(vpd_cols, "bill_date")
    invoice_date = pick_col(vpd_cols, "vendor_inv_attach_approval_date", "invoice_date")
    invoice_no = pick_col(vpd_cols, "veninvverifyid", "invoice_no")
    vendor_name = pick_col(vpd_cols, "vendor_name")
    vendor_id = pick_col(vpd_cols, "vendor_id")
    dc_num = pick_col(vpd_cols, "dc_num")
    amount = pick_col(vpd_cols, "amount")
    additional_charges = pick_col(vpd_cols, "additionalcharges", "additional_charges")
    tds = pick_col(vpd_cols, "acctdsvalue")
    other = pick_col(vpd_cols, "accotherdeduction")
    advance = pick_col(vpd_cols, "advancepayment")
    payable = pick_col(vpd_cols, "acctotalpaybleamount")
    vendor_app = pick_col(vpd_cols, "vendor_bill_approval")
    vendor_app_date = pick_col(vpd_cols, "vendor_bill_approval_date")
    finance_app = pick_col(vpd_cols, "finance_approved_by")
    finance_app_date = pick_col(vpd_cols, "finance_approved_date")
    mgmt_app = pick_col(vpd_cols, "managment_team_approvedby")
    mgmt_app_date = pick_col(vpd_cols, "managment_team_approvaldate")
    mgmt_status = pick_col(vpd_cols, "managment_team_approval_sts")
    remark_col = pick_col(vpd_cols, "managementremark")

    where = ["COALESCE(is_delete,0)=0"]
    params = []
    if "finance_approval" in vpd_cols:
        where.append("COALESCE(finance_approval,0)=1")
    if "accstatus" in vpd_cols:
        where.append("COALESCE(accstatus,0)=1")
    if mgmt_status:
        where.append(f"COALESCE({mgmt_status},0) = 0" if tab == "pending" else f"COALESCE({mgmt_status},0) IN (1,2)")
    if from_date and bill_date:
        where.append(f"DATE({bill_date}) >= %s")
        params.append(from_date)
    if to_date and bill_date:
        where.append(f"DATE({bill_date}) <= %s")
        params.append(to_date)

    having = ""
    if search:
        like = f"%{search}%"
        search_parts = ["bill_no LIKE %s"]
        params.append(like)
        if invoice_no:
            search_parts.append(f"MAX(CAST({invoice_no} AS CHAR)) LIKE %s")
            params.append(like)
        if vendor_name:
            search_parts.append(f"MAX(CAST({vendor_name} AS CHAR)) LIKE %s")
            params.append(like)
        having = f"HAVING ({' OR '.join(search_parts)})"

    amount_expr = "0"
    if amount and additional_charges:
        amount_expr = f"ROUND(IFNULL(SUM({amount}),0) + MAX(IFNULL({additional_charges},0)),2)"
    elif amount:
        amount_expr = f"ROUND(IFNULL(SUM({amount}),0),2)"
    elif additional_charges:
        amount_expr = f"ROUND(MAX(IFNULL({additional_charges},0)),2)"

    sql = f"""
        SELECT
            bill_no,
            {fmt_max_date_expr(bill_date, 'bill_date')},
            {fmt_max_date_expr(invoice_date, 'invoice_date')},
            {f"MAX(CAST({invoice_no} AS CHAR)) AS invoice_no" if invoice_no else "'' AS invoice_no"},
            {f"MAX(CAST({vendor_name} AS CHAR)) AS vendor_name" if vendor_name else "'' AS vendor_name"},
            {f"MAX(CAST(get_vendor_address({vendor_id}) AS CHAR)) AS vendor_address" if vendor_id else "'' AS vendor_address"},
            {f"MAX(CAST(get_vendor_company_name({vendor_id}) AS CHAR)) AS vendor_reg_id" if vendor_id else "'' AS vendor_reg_id"},
            {f"MAX(CAST(get_vendor_contact({vendor_id}) AS CHAR)) AS vendor_phone" if vendor_id else "'' AS vendor_phone"},
            {f"COUNT(DISTINCT {dc_num}) AS dc_count" if dc_num else "0 AS dc_count"},
            {amount_expr} AS amount,
            {f"ROUND(IFNULL(SUM({tds}),0),2) AS tds_deduction" if tds else "0 AS tds_deduction"},
            {f"ROUND(IFNULL(SUM({other}),0),2) AS other_deduction" if other else "0 AS other_deduction"},
            {f"ROUND(IFNULL(MAX({advance}),0),2) AS advance_amount" if advance else "0 AS advance_amount"},
            {f"ROUND(IFNULL(MAX({payable}),0),2) AS total_payable" if payable else "0 AS total_payable"},
            {name_expr(vendor_app, 'vendor_approved_by')},
            {fmt_max_date_expr(vendor_app_date, 'vendor_approved_date')},
            {name_expr(finance_app, 'accounts_approved_by')},
            {fmt_max_date_expr(finance_app_date, 'accounts_approved_date')},
            {name_expr(mgmt_app, 'mgmt_approved_by')},
            {fmt_max_date_expr(mgmt_app_date, 'mgmt_approved_date')},
            {f"MAX(COALESCE({mgmt_status},0)) AS mgmt_status" if mgmt_status else "0 AS mgmt_status"},
            {f"MAX(CAST({remark_col} AS CHAR)) AS remarks" if remark_col else "'' AS remarks"}
        FROM vendor_payment_details
        WHERE {' AND '.join(where)}
        GROUP BY bill_no, vendor_id
        {having}
        ORDER BY MAX(unique_id) DESC
    """
    count_sql = f"SELECT COUNT(*) AS total FROM ({sql}) t"
    total = safe_int((db_fetchone(count_sql, params) or {}).get("total"))
    rows = db_fetch(f"{sql} LIMIT %s OFFSET %s", params + [length, max(page - 1, 0) * length])
    return rows, total


class ManagementBillListView(APIView):
    def get(self, request):
        tab = safe_str(request.query_params.get("tab", "pending")).lower()
        search = safe_str(request.query_params.get("search"))
        from_date = safe_str(request.query_params.get("from_date"))
        to_date = safe_str(request.query_params.get("to_date"))
        page = safe_int(request.query_params.get("page")) or 1
        length = safe_int(request.query_params.get("length")) or 10
        rows, total = list_rows(tab, search, from_date, to_date, page, length)
        data = []
        for idx, row in enumerate(rows, start=((page - 1) * length) + 1):
            vendor = safe_str(row.get("vendor_name"))
            reg = safe_str(row.get("vendor_reg_id"))
            addr = safe_str(row.get("vendor_address"))
            phone = safe_str(row.get("vendor_phone"))
            vendor_line = vendor
            if reg:
                vendor_line = f"{vendor}-{reg}"
            if addr or phone:
                vendor_line = f"{vendor_line} {addr}{' - Ph.No. ' + phone if phone else ''}".strip()
            data.append({
                "id": f"{safe_str(row.get('bill_no'))}-{idx}",
                "bill_no": safe_str(row.get("bill_no")),
                "bill_date": safe_str(row.get("bill_date")),
                "invoice_no": safe_str(row.get("invoice_no")),
                "invoice_date": safe_str(row.get("invoice_date")),
                "vendor_name": vendor_line,
                "dc_count": safe_int(row.get("dc_count")),
                "amount": safe_float(row.get("amount")),
                "tds_deduction": safe_float(row.get("tds_deduction")),
                "other_deduction": safe_float(row.get("other_deduction")),
                "advance_amount": safe_float(row.get("advance_amount")),
                "total_payable": safe_float(row.get("total_payable")),
                "vendor_approved_by": safe_str(row.get("vendor_approved_by")),
                "vendor_approved_date": safe_str(row.get("vendor_approved_date")),
                "accounts_approved_by": safe_str(row.get("accounts_approved_by")),
                "accounts_approved_date": safe_str(row.get("accounts_approved_date")),
                "mgmt_approved_by": safe_str(row.get("mgmt_approved_by")),
                "mgmt_approved_date": safe_str(row.get("mgmt_approved_date")),
                "remarks": safe_str(row.get("remarks")),
                "status": status_text(row.get("mgmt_status")),
            })
        return Response({"data": data, "total": total, "page": page, "pages": max(1, -(-total // length))})


class ManagementBillDetailView(APIView):
    def get(self, request):
        bill_no = safe_str(request.query_params.get("bill_no")).strip()
        if not bill_no:
            return Response({"error": "bill_no required"}, status=status.HTTP_400_BAD_REQUEST)

        vpd_cols = table_columns("vendor_payment_details")
        vc_cols = table_columns("vendor_creation")
        dc_date_col = pick_col(vpd_cols, "dc_date")
        invoice_date_col = pick_col(vpd_cols, "vendor_inv_attach_approval_date", "invoice_date")
        po_date_col = pick_col(vpd_cols, "po_date")
        po_num_col = pick_col(vpd_cols, "po_num")
        invoice_no_col = pick_col(vpd_cols, "invoice_no", "veninvverifyid")
        invoice_qty_col = pick_col(vpd_cols, "invoice_qty")
        rate_col = pick_col(vpd_cols, "rate")
        gst_col = pick_col(vpd_cols, "gst")
        amount_col = pick_col(vpd_cols, "amount")
        additional_charges_col = pick_col(vpd_cols, "additionalcharges", "additional_charges")
        join_vendor = "LEFT JOIN vendor_creation v ON v.unique_id = vb.vendor_id" if "vendor_id" in vpd_cols and "unique_id" in vc_cols else ""

        detail_sql = f"""
            SELECT
                vb.bill_no,
                {fmt_date_expr('vb.bill_date' if 'bill_date' in vpd_cols else '', 'bill_date')},
                {f"CAST(vb.veninvverifyid AS CHAR) AS invoice_no" if 'veninvverifyid' in vpd_cols else "'' AS invoice_no"},
                {fmt_date_expr('vb.vendor_inv_attach_approval_date' if 'vendor_inv_attach_approval_date' in vpd_cols else '', 'invoice_date')},
                {f"CAST(v.inv_verfiy_attach AS CHAR) AS invoice_attach_url" if False else f"CAST(vb.inv_verfiy_attach AS CHAR) AS invoice_attach_url" if 'inv_verfiy_attach' in vpd_cols else "'#' AS invoice_attach_url"},
                {f"CAST(vb.po_ven_filename AS CHAR) AS po_attach_url" if 'po_ven_filename' in vpd_cols else "'#' AS po_attach_url"},
                {f"IFNULL(vb.acctdsvalue,0) AS tds_deduction" if 'acctdsvalue' in vpd_cols else "0 AS tds_deduction"},
                {f"IFNULL(vb.accotherdeduction,0) AS others_deduction" if 'accotherdeduction' in vpd_cols else "0 AS others_deduction"},
                {f"IFNULL(vb.advancepayment,0) AS advance_amount" if 'advancepayment' in vpd_cols else "0 AS advance_amount"},
                {f"IFNULL(vb.acctotalpaybleamount,0) AS total_payable" if 'acctotalpaybleamount' in vpd_cols else "0 AS total_payable"},
                {f"IFNULL(vb.{additional_charges_col},0) AS additional_charges" if additional_charges_col else "0 AS additional_charges"},
                {f"CAST(vb.managementremark AS CHAR) AS remarks" if 'managementremark' in vpd_cols else "'' AS remarks"},
                {f"CAST(vb.managment_team_reject_reason AS CHAR) AS reject_reason" if 'managment_team_reject_reason' in vpd_cols else "'' AS reject_reason"},
                {f"CAST(v.company_name AS CHAR) AS vendor_name" if 'company_name' in vc_cols else "'' AS vendor_name"},
                {f"CAST(v.address AS CHAR) AS vendor_address" if 'address' in vc_cols else "'' AS vendor_address"},
                {f"CAST(v.gst_no AS CHAR) AS vendor_gst" if 'gst_no' in vc_cols else "'' AS vendor_gst"},
                {f"CAST(v.pan_no AS CHAR) AS vendor_pan" if 'pan_no' in vc_cols else "'' AS vendor_pan"},
                {f"CAST(v.mail_id AS CHAR) AS vendor_email" if 'mail_id' in vc_cols else "'' AS vendor_email"},
                {f"CAST(v.contact_no AS CHAR) AS vendor_phone" if 'contact_no' in vc_cols else "'' AS vendor_phone"},
                {f"CAST(v.bank_name AS CHAR) AS bank_name" if 'bank_name' in vc_cols else "'' AS bank_name"},
                {f"CAST(v.branch_name AS CHAR) AS branch" if 'branch_name' in vc_cols else "'' AS branch"},
                {f"CAST(v.account_no AS CHAR) AS account_no" if 'account_no' in vc_cols else "'' AS account_no"},
                {f"CAST(v.ifsc_code AS CHAR) AS ifsc_code" if 'ifsc_code' in vc_cols else "'' AS ifsc_code"},
                {f"CAST(v.acc_holder_name AS CHAR) AS account_holder" if 'acc_holder_name' in vc_cols else "'' AS account_holder"},
                {f"CAST(v.pan_attach_file_name AS CHAR) AS pan_copy_url" if 'pan_attach_file_name' in vc_cols else "'#' AS pan_copy_url"},
                {f"CAST(v.bank_proof AS CHAR) AS bank_proof_url" if 'bank_proof' in vc_cols else "'#' AS bank_proof_url"}
            FROM vendor_payment_details vb
            {join_vendor}
            WHERE vb.bill_no = %s AND COALESCE(vb.is_delete,0)=0
            LIMIT 1
        """
        bill = db_fetchone(detail_sql, [bill_no])
        if not bill:
            return Response({"error": "Bill not found"}, status=status.HTTP_404_NOT_FOUND)

        dc_sql = """
            SELECT
                dc_num AS dc_no,
                {dc_date_expr},
                {invoice_no_expr},
                {invoice_date_expr},
                {po_no_expr},
                {po_date_expr},
                {invoice_qty_expr},
                COALESCE(get_address((SELECT consignee_unique_id FROM view_outsource_vendor_verified_invoice WHERE vendor_payment_details.dc_num = view_outsource_vendor_verified_invoice.dc_number LIMIT 1)), '') AS consignee_address,
                {unit_price_expr},
                {basic_amount_expr},
                {gst_expr},
                {gst_amount_expr},
                {total_amount_expr}
            FROM vendor_payment_details
            WHERE bill_no = %s AND COALESCE(is_delete,0)=0
            ORDER BY unique_id
        """.format(
            dc_date_expr=fmt_date_expr(f"vendor_payment_details.{dc_date_col}" if dc_date_col else "", "dc_date"),
            invoice_no_expr=f"CAST(vendor_payment_details.{invoice_no_col} AS CHAR) AS invoice_no" if invoice_no_col else "'' AS invoice_no",
            invoice_date_expr=fmt_date_expr(f"vendor_payment_details.{invoice_date_col}" if invoice_date_col else "", "invoice_date"),
            po_no_expr=f"CAST(vendor_payment_details.{po_num_col} AS CHAR) AS po_no" if po_num_col else "'' AS po_no",
            po_date_expr=fmt_date_expr(f"vendor_payment_details.{po_date_col}" if po_date_col else "", "po_date"),
            invoice_qty_expr=f"IFNULL(vendor_payment_details.{invoice_qty_col},0) AS invoice_qty" if invoice_qty_col else "0 AS invoice_qty",
            unit_price_expr=f"IFNULL(vendor_payment_details.{rate_col},0) AS unit_price" if rate_col else "0 AS unit_price",
            basic_amount_expr=(
                f"(IFNULL(vendor_payment_details.{invoice_qty_col},0) * IFNULL(vendor_payment_details.{rate_col},0)) AS basic_amount"
                if invoice_qty_col and rate_col else "0 AS basic_amount"
            ),
            gst_expr=f"IFNULL(vendor_payment_details.{gst_col},0) AS gst" if gst_col else "0 AS gst",
            gst_amount_expr=(
                f"((IFNULL(vendor_payment_details.{invoice_qty_col},0) * IFNULL(vendor_payment_details.{rate_col},0)) * (IFNULL(vendor_payment_details.{gst_col},0) / 100)) AS gst_amount"
                if invoice_qty_col and rate_col and gst_col else "0 AS gst_amount"
            ),
            total_amount_expr=f"IFNULL(vendor_payment_details.{amount_col},0) AS total_amount" if amount_col else "0 AS total_amount",
        )
        dc_rows = db_fetch(dc_sql, [bill_no])
        dc_items = []
        total_amount = 0.0
        for idx, row in enumerate(dc_rows, start=1):
            total_amount += safe_float(row.get("total_amount"))
            dc_items.append({
                "s_no": idx,
                "dc_no": safe_str(row.get("dc_no")),
                "dc_date": safe_str(row.get("dc_date")),
                "invoice_no": safe_str(row.get("invoice_no")),
                "invoice_date": safe_str(row.get("invoice_date")),
                "po_no": safe_str(row.get("po_no")),
                "po_date": safe_str(row.get("po_date")),
                "consignee_address": safe_str(row.get("consignee_address")),
                "invoice_qty": safe_float(row.get("invoice_qty")),
                "unit_price": safe_float(row.get("unit_price")),
                "basic_amount": safe_float(row.get("basic_amount")),
                "gst": f"{safe_float(row.get('gst')):g} %",
                "gst_amount": safe_float(row.get("gst_amount")),
                "total_amount": safe_float(row.get("total_amount")),
            })

        additional_charges = safe_float(bill.get("additional_charges"))
        grand_total_amount = total_amount + additional_charges

        approval_sql = """
            SELECT
                COALESCE(get_staff_name(MAX(vendor_bill_created_by)), '') AS bill_created_by,
                COALESCE(DATE_FORMAT(MAX(vendor_bill_created_date), '%%d-%%m-%%Y %%H:%%i:%%s'), '') AS bill_created_at,
                COALESCE(get_staff_name(MAX(vendor_bill_approval)), '') AS operation_by,
                COALESCE(DATE_FORMAT(MAX(vendor_bill_approval_date), '%%d-%%m-%%Y %%H:%%i:%%s'), '') AS operation_at,
                COALESCE(MAX(vendor_bill_app_status), 0) AS operation_status,
                COALESCE(get_staff_name(MAX(vendor_account_approved_by)), '') AS account_entry_by,
                COALESCE(DATE_FORMAT(MAX(vendor_account_approval_date), '%%d-%%m-%%Y %%H:%%i:%%s'), '') AS account_entry_at,
                COALESCE(MAX(acc_ent_sts), 0) AS account_entry_status,
                COALESCE(get_staff_name(MAX(finance_approved_by)), '') AS accounts_approval_by,
                COALESCE(DATE_FORMAT(MAX(finance_approved_date), '%%d-%%m-%%Y %%H:%%i:%%s'), '') AS accounts_approval_at,
                COALESCE(MAX(finance_approval), 0) AS accounts_approval_status,
                COALESCE(get_staff_name(MAX(managment_team_approvedby)), '') AS management_by,
                COALESCE(DATE_FORMAT(MAX(managment_team_approvaldate), '%%d-%%m-%%Y %%H:%%i:%%s'), '') AS management_at,
                COALESCE(MAX(managment_team_approval_sts), 0) AS management_status,
                COALESCE(MAX(transaction_id), '') AS payment_ref,
                COALESCE(DATE_FORMAT(MAX(transaction_date), '%%d-%%m-%%Y'), '') AS payment_date,
                COALESCE(MAX(acctotalpaybleamount), 0) AS payment_amount,
                COALESCE(MAX(accounts_approval), 0) AS payment_status
            FROM vendor_payment_details
            WHERE bill_no = %s AND COALESCE(is_delete,0)=0
        """
        app = db_fetchone(approval_sql, [bill_no]) or {}

        return Response({
            "id": safe_str(bill.get("bill_no")),
            "vendor_name": safe_str(bill.get("vendor_name")),
            "vendor_address": safe_str(bill.get("vendor_address")),
            "vendor_gst": safe_str(bill.get("vendor_gst")),
            "vendor_pan": safe_str(bill.get("vendor_pan")),
            "vendor_email": safe_str(bill.get("vendor_email")),
            "vendor_phone": safe_str(bill.get("vendor_phone")),
            "vendor_bill_no": safe_str(bill.get("bill_no")),
            "vendor_bill_date": safe_str(bill.get("bill_date")),
            "vendor_invoice_no": safe_str(bill.get("invoice_no")),
            "vendor_invoice_date": safe_str(bill.get("invoice_date")),
            "invoice_attach_url": safe_str(bill.get("invoice_attach_url")) or "#",
            "po_attach_url": safe_str(bill.get("po_attach_url")) or "#",
            "bank_name": safe_str(bill.get("bank_name")),
            "branch": safe_str(bill.get("branch")),
            "account_no": safe_str(bill.get("account_no")),
            "ifsc_code": safe_str(bill.get("ifsc_code")),
            "account_holder": safe_str(bill.get("account_holder")),
            "pan_copy_url": safe_str(bill.get("pan_copy_url")) or "#",
            "bank_proof_url": safe_str(bill.get("bank_proof_url")) or "#",
            "remarks": safe_str(bill.get("remarks")),
            "reject_reason": safe_str(bill.get("reject_reason")),
            "dc_items": dc_items,
            "total_amount": total_amount,
            "additional_charges": additional_charges,
            "grand_total_amount": grand_total_amount,
            "tds_deduction": safe_float(bill.get("tds_deduction")),
            "others_deduction": safe_float(bill.get("others_deduction")),
            "advance_amount": safe_float(bill.get("advance_amount")),
            "total_payable": safe_float(bill.get("total_payable")) or total_amount,
            "approvals": [{
                "s_no": 1,
                "bill_created_by": safe_str(app.get("bill_created_by")),
                "bill_created_at": safe_str(app.get("bill_created_at")),
                "bill_created_status": "Bill Created",
                "operation_by": safe_str(app.get("operation_by")),
                "operation_at": safe_str(app.get("operation_at")),
                "operation_status": status_text(app.get("operation_status")),
                "account_entry_by": safe_str(app.get("account_entry_by")),
                "account_entry_at": safe_str(app.get("account_entry_at")),
                "account_entry_status": status_text(app.get("account_entry_status")),
                "accounts_approval_by": safe_str(app.get("accounts_approval_by")),
                "accounts_approval_at": safe_str(app.get("accounts_approval_at")),
                "accounts_approval_status": status_text(app.get("accounts_approval_status")),
                "management_by": safe_str(app.get("management_by")),
                "management_at": safe_str(app.get("management_at")),
                "management_status": status_text(app.get("management_status")),
                "payment_ref": safe_str(app.get("payment_ref")),
                "payment_date": safe_str(app.get("payment_date")),
                "payment_amount": f"{safe_float(app.get('payment_amount')):,.2f}",
                "payment_status": payment_status_text(app.get("payment_status")),
            }],
        })


class ManagementBillUpdateRemarkView(APIView):
    def post(self, request):
        bill_no = safe_str(request.data.get("bill_no")).strip()
        remark = safe_str(request.data.get("remark"))
        if not bill_no:
            return Response({"error": "bill_no required"}, status=status.HTTP_400_BAD_REQUEST)
        with connection.cursor() as cursor:
            cursor.execute("UPDATE vendor_payment_details SET managementremark=%s WHERE bill_no=%s AND COALESCE(is_delete,0)=0", [remark, bill_no])
            try:
                cursor.execute("UPDATE vendor_payment_details_main SET managementremark=%s WHERE bill_no=%s AND COALESCE(is_delete,0)=0", [remark, bill_no])
            except Exception:
                pass
        return Response({"status": 1, "msg": "update"})


class ManagementBillApproveView(APIView):
    def post(self, request):
        bill_no = safe_str(request.data.get("bill_no")).strip()
        action_status = safe_str(request.data.get("status", "Approved"))
        reject_reason = safe_str(request.data.get("remark"))
        user_id = get_user_id(request)
        if not bill_no:
            return Response({"error": "bill_no required"}, status=status.HTTP_400_BAD_REQUEST)

        status_map = {"Approved": 1, "Rejected": 2, "Pending": 0}
        status_value = status_map.get(action_status, 0)

        with connection.cursor() as cursor:
            if status_value == 1:
                cursor.execute(
                    """
                    UPDATE vendor_payment_details
                    SET managment_team_allocated=1,
                        managment_team_approval_sts=1,
                        managment_team_approvedby=%s,
                        managment_team_approvaldate=NOW()
                    WHERE bill_no=%s AND COALESCE(is_delete,0)=0
                    """,
                    [user_id, bill_no],
                )
                try:
                    cursor.execute(
                        """
                        UPDATE vendor_payment_details_main
                        SET managment_team_allocated=1,
                            managment_team_approval_sts=1,
                            managment_team_approvedby=%s,
                            managment_team_approvaldate=NOW()
                        WHERE bill_no=%s AND COALESCE(is_delete,0)=0
                        """,
                        [user_id, bill_no],
                    )
                except Exception:
                    pass
                try:
                    cursor.execute(
                        """
                        UPDATE invoice_verfication_table iv
                        JOIN vendor_payment_details vpd ON vpd.dc_num = iv.dc_number
                        SET iv.managment_team_allocated=1,
                            iv.managment_team_approval_sts=1
                        WHERE vpd.bill_no=%s AND COALESCE(vpd.is_delete,0)=0
                        """,
                        [bill_no],
                    )
                except Exception:
                    pass
            elif status_value == 2:
                cursor.execute(
                    """
                    UPDATE vendor_payment_details
                    SET managment_team_allocated=0,
                        managment_team_approval_sts=%s,
                        managment_team_reject_reason=%s,
                        managment_team_approvedby=%s,
                        managment_team_approvaldate=NOW()
                    WHERE bill_no=%s AND COALESCE(is_delete,0)=0
                    """,
                    [status_value, reject_reason, user_id, bill_no],
                )
                try:
                    cursor.execute(
                        """
                        UPDATE vendor_payment_details_main
                        SET managment_team_allocated=0,
                            managment_team_approval_sts=%s,
                            managment_team_reject_reason=%s,
                            managment_team_approvedby=%s,
                            managment_team_approvaldate=NOW()
                        WHERE bill_no=%s AND COALESCE(is_delete,0)=0
                        """,
                        [status_value, reject_reason, user_id, bill_no],
                    )
                except Exception:
                    pass
                try:
                    cursor.execute(
                        """
                        UPDATE invoice_verfication_table iv
                        JOIN vendor_payment_details vpd ON vpd.dc_num = iv.dc_number
                        SET iv.managment_team_allocated=0,
                            iv.managment_team_approval_sts=%s,
                            iv.vendor_finance_approval=0,
                            iv.vendor_payment_allocated=0,
                            iv.vendor_bill_no=''
                        WHERE vpd.bill_no=%s AND COALESCE(vpd.is_delete,0)=0
                        """,
                        [status_value, bill_no],
                    )
                except Exception:
                    pass
            else:
                cursor.execute(
                    """
                    UPDATE vendor_payment_details
                    SET managment_team_allocated=0,
                        managment_team_approval_sts=0,
                        managment_team_reject_reason='',
                        managment_team_approvedby='',
                        managment_team_approvaldate=NULL
                    WHERE bill_no=%s AND COALESCE(is_delete,0)=0
                    """,
                    [bill_no],
                )
                try:
                    cursor.execute(
                        """
                        UPDATE vendor_payment_details_main
                        SET managment_team_allocated=0,
                            managment_team_approval_sts=0,
                            managment_team_reject_reason='',
                            managment_team_approvedby='',
                            managment_team_approvaldate=NULL
                        WHERE bill_no=%s AND COALESCE(is_delete,0)=0
                        """,
                        [bill_no],
                    )
                except Exception:
                    pass

        return Response({"status": 1, "msg": action_status.lower()})


class ManagementBillExportView(APIView):
    def get(self, request):
        rows, _ = list_rows("complete", "", "", "", 1, 100000)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Management Bill Approval"
        ws.merge_cells("A1:M1")
        ws["A1"] = "Management Team Bill Approval"
        ws["A1"].font = Font(bold=True, size=13)
        ws["A1"].alignment = Alignment(horizontal="center")
        headers = ["S.No", "Bill No", "Bill Date", "Invoice No", "Invoice Date", "Vendor Name", "DC Count", "Amount", "TDS Deduction", "Other Deduction", "Advance Amount", "Total Payable", "Status"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=header).font = Font(bold=True)
        for idx, row in enumerate(rows, 1):
            ws.append([
                idx, safe_str(row.get("bill_no")), safe_str(row.get("bill_date")), safe_str(row.get("invoice_no")),
                safe_str(row.get("invoice_date")), safe_str(row.get("vendor_name")), safe_int(row.get("dc_count")),
                safe_float(row.get("amount")), safe_float(row.get("tds_deduction")), safe_float(row.get("other_deduction")),
                safe_float(row.get("advance_amount")), safe_float(row.get("total_payable")), status_text(row.get("mgmt_status")),
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="ManagementBillApproval.xlsx"'
        return response
