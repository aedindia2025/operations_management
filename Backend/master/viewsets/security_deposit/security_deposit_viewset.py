"""
master/viewsets/security_deposit_viewset.py

Full Python/Django-REST-Framework equivalent of:
  crud.php   – all switch-case actions
  form.php   – edit-form data endpoint
  view.php   – bill detail view endpoint
  excel.php  – Excel export
  partial_payment.php – partial payment page data

Tables involved (mirrors PHP):
  bill_submission_sub          → BillSubmissionSub model
  bill_submission_main_table   → BillSubmissionMain model
  bill_submission_form         → BillSubmissionForm model
  view_payment_entry_list      → read-only DB view (PaymentEntryView)
  view_partial_elcot_entry_final → PartialElcotEntryFinal view
  view_bill_submission_main_table → BillSubmissionMainView
  bg_creation_main             → BgCreationMain
  sign_doc_verification_detail → SignDocVerification
  installation_details         → InstallationDetails
  po_form                      → PoForm
"""

from __future__ import annotations

import datetime
import re
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment
from django.db import transaction, connection
from django.db.models import Q, Count, Max
from django.http import HttpResponse
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.viewsets import ViewSet

from master.apps.signed_document_verification.signed_doc_verification_model import (
    InstallationDetails,
    SignDocVerificationDetail,
)
from master.apps.vendor_bill_creation.vendorbillcreationmodel import (
    BillSubmissionForm,
    BillSubmissionMainTable as BillSubmissionMain,
    BillSubmissionSub,
)
from master.models import (
    InvoiceCreation,
    PartialElcotEntryFinal,
    PaymentEntryView,
    ViewBillSubmissionMain,
    ViewPartialBill1,
    ViewPaymentEntryList,
)

from master.serializers.security_deposit.security_deposit_serializer import (
    SecurityDepositCreateUpdateSerializer,
    PartialBillMainSerializer,
    PartialBillCreationAddUpdateSerializer,
)


# ─────────────────────────────────────────────────────────────────────────────
BgCreationMain = None
PoForm = None

# Helper utilities
# ─────────────────────────────────────────────────────────────────────────────

def _indian_money(value) -> str:
    """Python equivalent of moneyFormatIndia1() / indian_format_value()."""
    try:
        val = round(float(value or 0))
        s = str(val)
        last3 = s[-3:]
        rest = s[:-3]
        if rest:
            rest = re.sub(r"\B(?=(\d{2})+(?!\d))", ",", rest)
            return f"{rest},{last3}"
        return last3
    except Exception:
        return str(value or "0")


def _claim_value_cal(claim_percent, invoice_value) -> float:
    """Port of PHP claim_value_cal(): (invoice_value * claim_percent) / 100."""
    try:
        return round((float(invoice_value) * float(claim_percent)) / 100)
    except Exception:
        return 0.0


def _remaining_percent(claim_pct) -> int:
    """100 minus the stored claim percent (e.g. 75 → 25)."""
    try:
        return 100 - int(claim_pct)
    except Exception:
        return 0


def _fmt_date(d) -> str:
    """Format a date/string to dd-mm-yyyy."""
    if not d:
        return ""
    if isinstance(d, (datetime.date, datetime.datetime)):
        return d.strftime("%d-%m-%Y")
    try:
        return datetime.datetime.strptime(str(d), "%Y-%m-%d").strftime("%d-%m-%Y")
    except Exception:
        return str(d)


def _generate_bill_no() -> str:
    """
    Python port of bill_no_get() from crud.php.
    Pattern: BN-YYMM-NNNN   (e.g. BN-2504-0003)
    """
    now = datetime.date.today()
    yy = now.strftime("%y")
    mm = now.strftime("%m")
    prefix = f"BN-{yy}{mm}-"

    last = (
        BillSubmissionMain.objects
        .filter(bill_no__startswith="BN-", is_delete=False)
        .order_by("-id")
        .values_list("bill_no", flat=True)
        .first()
    )

    if last:
        parts = last.split("-")
        try:
            # Check if still the same year
            last_yy = parts[1][:2]
            if last_yy == yy:
                seq = int(parts[2]) + 1
                return f"{prefix}{str(seq).zfill(4)}"
        except (IndexError, ValueError):
            pass

    return f"{prefix}0001"


def _generate_unique_id(prefix: str = "") -> str:
    """Generate a simple unique id (mirrors PHP unique_id() helper)."""
    import uuid
    return prefix + uuid.uuid4().hex[:18]


def _paginate(qs, start: int, length: int):
    total = qs.count()
    if length and length != -1:
        page = qs[start: start + length]
    else:
        page = qs
    return page, total


def _apply_date_filter(qs, from_date: str, to_date: str, opt: str,
                       po_field="po_date", bill_field="bill_created_date"):
    if not from_date or not to_date:
        return qs
    if opt == "40":
        return qs.filter(**{f"{po_field}__gte": from_date, f"{po_field}__lte": to_date})
    if opt == "50":
        return qs.filter(**{f"{bill_field}__gte": from_date, f"{bill_field}__lte": to_date})
    return qs


def _get_customer_detail(po_unique_id: str) -> dict:
    """
    Python equivalent of the MySQL stored procedure
    GetCustomerDetailByponum(po_unique_id).
    Adapt to your actual model relationships.
    """
    try:
        with connection.cursor() as cur:
            cur.callproc("GetCustomerDetailByponum", [po_unique_id])
            columns = [col[0] for col in cur.description]
            row = cur.fetchone()
            if row:
                return dict(zip(columns, row))
    except Exception:
        pass
    return {}


def _dc_ir_filter(qs):
    """
    Apply the complex DC-required / IR-status EXISTS filter from crud.php.
    Mirrors the big WHERE clause added in both 'datatable' and
    'partial_payment_datatable' cases.
    """
    from django.db.models import Exists, OuterRef
    

    dc_required_exists = InstallationDetails.objects.filter(
        invoice_no=OuterRef("invoice_no"),
        documents_type="DC",
        dc_required="1",
    )
    dc_received_exists = SignDocVerificationDetail.objects.filter(
        invoice_no=OuterRef("invoice_no"),
        dc_received_status="1",
    )
    ir_no_dc_exists = SignDocVerificationDetail.objects.filter(
        invoice_no=OuterRef("invoice_no"),
        dc_received_status="2",
        ir_status="1",
    )
    not_dc_required_exists = InstallationDetails.objects.filter(
        invoice_no=OuterRef("invoice_no"),
        documents_type="DC",
        dc_required="1",
    )

    return qs.filter(
        Q(
            Exists(dc_required_exists) & Exists(dc_received_exists)
        ) | Q(
            ~Exists(not_dc_required_exists) & Exists(ir_no_dc_exists)
        )
    )


# ─────────────────────────────────────────────────────────────────────────────
# ViewSet
# ─────────────────────────────────────────────────────────────────────────────

class SecurityDepositViewSet(ViewSet):
    """
    Router registration (in master/apps/urls.py):

        router.register(
            r"security-deposit",
            SecurityDepositViewSet,
            basename="security-deposit",
        )

    Endpoints produced:
      GET  /security-deposit/list/                  → main datatable (tab 1)
      GET  /security-deposit/partial-bill-list/     → partial bill datatable (tab 2)
      GET  /security-deposit/partial-payment-list/  → partial payment modal datatable
      GET  /security-deposit/bill-generate-list/    → bill-generate datatable
      GET  /security-deposit/customer-details/      → customer details lookup
      GET  /security-deposit/form-data/             → payment edit-form pre-fill
      GET  /security-deposit/view/                  → bill detail view page data
      POST /security-deposit/create-update/         → create or update bill_submission_form
      POST /security-deposit/partial-bill-main/     → bulk partial bill insert
      POST /security-deposit/partial-bill-add/      → add row to bill_submission_sub
      GET  /security-deposit/export-excel/          → Excel download
      DELETE /security-deposit/{pk}/delete/         → soft-delete
    """

    permission_classes = [IsAuthenticated]

    # ──────────────────────────────────────────────────────────────────────────
    # 1. MAIN DATATABLE  (PHP case 'datatable')
    # ──────────────────────────────────────────────────────────────────────────

    @action(detail=False, methods=["get"], url_path="list")
    def datatable(self, request):
        search    = request.query_params.get("search", "")
        length    = int(request.query_params.get("length", 10))
        start     = int(request.query_params.get("start", 0))
        from_date = request.query_params.get("from_date", "")
        to_date   = request.query_params.get("to_date", "")
        opt       = request.query_params.get("opt", "")
        draw      = int(request.query_params.get("draw", 1))

        qs = (
            BillSubmissionSub.objects
            .filter(bill_status="2", sd_status="0")
            .exclude(partial_bill_status=3)
        )
        qs = _dc_ir_filter(qs)
        qs = _apply_date_filter(qs, from_date, to_date, opt,
                                po_field="po_date", bill_field="bill_created_date")

        if search:
            qs = qs.filter(
                Q(po_num__icontains=search)
                | Q(invoice_no__icontains=search)
                | Q(bill_no__icontains=search)
            )

        page_qs, total = _paginate(qs, start, length)
        data = []
        for idx, row in enumerate(page_qs, start=start + 1):
            claim_pct = _remaining_percent(row.claim_percentage)
            inv_val   = float(row.invoice_value or 0)
            claim_val = _claim_value_cal(claim_pct, inv_val)

            # Customer details via stored procedure
            cust = _get_customer_detail(row.bill_form_unique_id)
            dept     = cust.get("department", "")
            district = cust.get("district", "")
            state    = cust.get("state_name", "")

            data.append({
                "s_no"           : idx,
                "bill_details"   : f"{row.bill_no} / {_fmt_date(row.bill_created_date)}",
                "po_details"     : f"{row.po_num} / {_fmt_date(row.po_date)}",
                "invoice_details": f"{row.invoice_no} / {_fmt_date(row.invoice_date)}",
                "customer_details": f"{dept} / {district} / {state}",
                "invoice_value"  : _indian_money(inv_val),
                "claim_percentage": f"{claim_pct}%",
                "claim_value"    : _indian_money(claim_val),
                "unique_id"      : row.unique_id,
                "bill_form_unique_id": row.bill_form_unique_id,
                "po_date"        : _fmt_date(row.po_date),
                "invoice_date"   : _fmt_date(row.invoice_date),
                "po_num"         : row.po_num,
                "invoice_no"     : row.invoice_no,
                "bill_no"        : row.bill_no,
                "bill_created_date": _fmt_date(row.bill_created_date),
                "partial_bill_status": row.partial_bill_status,
            })

        return Response({
            "draw"            : draw,
            "recordsTotal"    : total,
            "recordsFiltered" : total,
            "data"            : data,
        })

    # ──────────────────────────────────────────────────────────────────────────
    # 2. PARTIAL BILL DATATABLE  (PHP case 'partial_bill_datatable')
    # ──────────────────────────────────────────────────────────────────────────

    @action(detail=False, methods=["get"], url_path="partial-bill-list")
    def partial_bill_datatable(self, request):
        search = request.query_params.get("search", "")
        length = int(request.query_params.get("length", 10))
        start  = int(request.query_params.get("start", 0))
        draw   = int(request.query_params.get("draw", 1))

        # Mirrors PHP: WHERE bill_status='1' on view_partial_elcot_entry_final
        # Adapt the model/view name to your ORM setup.
        qs = PartialElcotEntryFinal.objects.filter(bill_status="1")

        if search:
            qs = qs.filter(
                Q(po_num__icontains=search) | Q(invoice_no__icontains=search)
            )

        page_qs, total = _paginate(qs, start, length)
        data = []
        for idx, row in enumerate(page_qs, start=start + 1):
            cust = _get_customer_detail(row.unique_id)
            dept     = cust.get("department", "")
            district = cust.get("district", "")
            state    = cust.get("state_name", "")

            bg_num = row.bg_num or "NA"
            bg_date = _fmt_date(row.bg_date) if row.bg_date else "NA"

            data.append({
                "file_org_name"  : "",        # checkbox HTML – built by frontend
                "s_no"           : idx,
                "bg_details"     : f"{bg_num} / {bg_date}",
                "po_details"     : f"{row.po_num} / {_fmt_date(row.po_date)}",
                "invoice_details": f"{row.invoice_no} / {_fmt_date(row.invoice_date)}",
                "customer_details": f"{dept} / {district} / {state}",
                "invoice_qty"    : row.invoice_qty,
                "invoice_value"  : row.invoice_value,
                "claim_amount"   : row.claim_amount,
                "unique_id"      : row.unique_id,
                "ledger_name"    : getattr(row, "ledger_name", ""),
                "ledger_no"      : getattr(row, "ledger_no", ""),
                "bg_date"        : bg_date,
                "bg_num"         : bg_num,
                "po_date"        : _fmt_date(row.po_date),
                "invoice_date"   : _fmt_date(row.invoice_date),
                "invoice_no"     : row.invoice_no or "NA",
                "po_num"         : row.po_num,
            })

        return Response({
            "draw": draw, "recordsTotal": total,
            "recordsFiltered": total, "data": data,
        })

    # ──────────────────────────────────────────────────────────────────────────
    # 3. PARTIAL PAYMENT MODAL DATATABLE  (PHP case 'partial_payment_datatable')
    # ──────────────────────────────────────────────────────────────────────────

    @action(detail=False, methods=["get"], url_path="partial-payment-list")
    def partial_payment_datatable(self, request):
        search = request.query_params.get("search", "")
        length = int(request.query_params.get("length", 10))
        start  = int(request.query_params.get("start", 0))
        draw   = int(request.query_params.get("draw", 1))
        po_id  = request.query_params.get("po_id", "")

        qs = (
            BillSubmissionSub.objects
            .filter(bill_status="2", sd_status="0")
            .exclude(partial_bill_status=3)
        )
        qs = _dc_ir_filter(qs)

        if po_id:
            qs = qs.filter(bill_form_unique_id=po_id)

        if search:
            qs = qs.filter(
                Q(po_num__icontains=search) | Q(invoice_no__icontains=search)
            )

        page_qs, total = _paginate(qs, start, length)
        data = []
        for idx, row in enumerate(page_qs, start=start + 1):
            claim_pct = _remaining_percent(row.claim_percentage)
            inv_val   = float(row.invoice_value or 0)
            claim_val = _claim_value_cal(claim_pct, inv_val)

            cust = _get_customer_detail(row.bill_form_unique_id)
            dept     = cust.get("department", "")
            district = cust.get("district", "")
            state    = cust.get("state_name", "")

            data.append({
                "chek_box"       : "",      # checkbox HTML – built by frontend
                "s_no"           : idx,
                "bg_no"          : "",
                "bill_details"   : f"{row.bill_no} / {_fmt_date(row.bill_created_date)}",
                "po_details"     : f"{row.po_num} / {_fmt_date(row.po_date)}",
                "invoice_details": f"{row.invoice_no} / {_fmt_date(row.invoice_date)}",
                "customer_details": f"{dept} / {district} / {state}",
                "invoice_value"  : _indian_money(inv_val),
                "claim_percentage": f"{claim_pct}%",
                "claim_value"    : _indian_money(claim_val),
                "bill_form_unique_id": row.bill_form_unique_id,
                "po_date"        : _fmt_date(row.po_date),
                "invoice_date"   : _fmt_date(row.invoice_date),
                "po_num"         : row.po_num,
                "invoice_no"     : row.invoice_no,
                "bill_no"        : row.bill_no,
                "bill_created_date": _fmt_date(row.bill_created_date),
                "partial_bill_status": row.partial_bill_status,
                "bg_value"       : "",
                "invoice_qty"    : row.invoice_qty,
            })

        return Response({
            "draw": draw, "recordsTotal": total,
            "recordsFiltered": total, "data": data,
        })

    # ──────────────────────────────────────────────────────────────────────────
    # 4. BILL GENERATE DATATABLE  (PHP case 'bill_generate_datatable')
    # ──────────────────────────────────────────────────────────────────────────

    @action(detail=False, methods=["get"], url_path="bill-generate-list")
    def bill_generate_datatable(self, request):
        search    = request.query_params.get("search", "")
        length    = int(request.query_params.get("length", 10))
        start     = int(request.query_params.get("start", 0))
        from_date = request.query_params.get("from_date", "")
        to_date   = request.query_params.get("to_date", "")
        opt       = request.query_params.get("opt1", "")
        draw      = int(request.query_params.get("draw", 1))

        qs = (
            BillSubmissionMain.objects
            .filter(partial_bill_status="3", is_delete=False)
        )
        qs = _apply_date_filter(qs, from_date, to_date, opt,
                                po_field="po_date", bill_field="bill_created_date")

        page_qs, total = _paginate(qs, start, length)
        data = []
        for idx, row in enumerate(page_qs, start=start + 1):
            inv_val      = float(row.invoice_value or 0)
            claim_pct    = float(row.claim_amount or 0)
            claim_amt    = (inv_val * claim_pct) / 100

            # Invoice count for this bill_no
            inv_count = (
                BillSubmissionForm.objects
                .filter(bill_form_main_unique_id=row.bill_form_main_unique_id,
                        is_delete=False)
                .count()
            )

            cust = _get_customer_detail(row.bill_form_main_unique_id)
            dept     = cust.get("department", "")
            district = cust.get("district", "")
            state    = cust.get("state_name", "")

            data.append({
                "s_no"           : idx,
                "bill_details"   : f"{row.bill_no} / {_fmt_date(row.bill_created_date)}",
                "po_num"         : f"{row.po_num} / {_fmt_date(row.po_date)}",
                "customer_details": f"{dept} / {district} / {state}",
                "invcount"       : inv_count,
                "invoice_value"  : _indian_money(inv_val),
                "claim_amount"   : str(claim_pct),
                "claimamnt"      : _indian_money(claim_amt),
                "bg_num"         : row.bg_num or "",
                "view"           : f"/security-deposit/view/?unique_id={row.bill_form_main_unique_id}&bill_no={row.bill_no}",
                "invoice_no"     : row.invoice_no,
                "invoice_date"   : _fmt_date(row.invoice_date),
                "bill_status"    : row.bill_status,
                "po_date"        : _fmt_date(row.po_date),
                "bill_no"        : row.bill_no,
                "bill_form_main_unique_id": row.bill_form_main_unique_id,
                "bill_created_dates": _fmt_date(row.bill_created_date),
            })

        return Response({
            "draw": draw, "recordsTotal": total,
            "recordsFiltered": total, "data": data,
        })

    # ──────────────────────────────────────────────────────────────────────────
    # 5. CUSTOMER DETAILS LOOKUP  (PHP case 'customerDetails')
    # ──────────────────────────────────────────────────────────────────────────

    @action(detail=False, methods=["get"], url_path="customer-details")
    def customer_details(self, request):
        po_id = request.query_params.get("po_id", "")
        if not po_id:
            return Response({"error": "po_id required"}, status=status.HTTP_400_BAD_REQUEST)

        cust = _get_customer_detail(po_id)
        if not cust:
            return Response({"data": []})

        return Response({
            "data": [{
                "po_number" : cust.get("po_num", ""),
                "po_date"   : _fmt_date(cust.get("po_date", "")),
                "department": cust.get("department", ""),
                "address"   : cust.get("bill_address", ""),
                "district"  : cust.get("district", ""),
                "state"     : cust.get("state_name", ""),
                "contact"   : cust.get("contact_number", ""),
                "email"     : cust.get("email", ""),
            }]
        })

    # ──────────────────────────────────────────────────────────────────────────
    # 6. PAYMENT EDIT-FORM DATA  (PHP form.php)
    # ──────────────────────────────────────────────────────────────────────────

    @action(detail=False, methods=["get"], url_path="form-data")
    def form_data(self, request):
        unique_id  = request.query_params.get("unique_id", "")
        invoice_no = request.query_params.get("invoice_no", "")

        if not unique_id or not invoice_no:
            return Response({"error": "unique_id and invoice_no required"},
                            status=status.HTTP_400_BAD_REQUEST)

        # Fetch main record (mirrors PHP view_partial_bill_1 query in form.php)
        try:
            rec = ViewPartialBill1.objects.get(
                bill_form_unique_id=unique_id,
                invoice_no=invoice_no,
            )
        except Exception as exc:
            return Response({"error": str(exc)}, status=status.HTTP_404_NOT_FOUND)

        # Fetch sub-form entry
        sub = BillSubmissionForm.objects.filter(
            bill_form_unique_id=unique_id,
            invoice_no=invoice_no,
        ).first()

        cust = _get_customer_detail(unique_id)

        bg = rec.with_bg or ""
        if bg == "on":
            bg = "With Bg"

        # Claim-amount options based on current claim percentage
        CLAIM_OPTIONS_MAP = {
            95:  [{"id": "5",  "value": "5%"}, {"id": "25", "value": "25%"}],
            75:  [{"id": "25", "value": "25%"}, {"id": "20", "value": "20%"}],
            25:  [{"id": "75", "value": "75%"}, {"id": "20", "value": "20%"}],
            90:  [{"id": "10", "value": "10%"}],
        }
        try:
            ca_key = int(float(rec.claim_amount or 0))
        except Exception:
            ca_key = 0

        return Response({
            "po_num"          : rec.po_num,
            "po_date"         : _fmt_date(rec.po_date),
            "invoice_no"      : rec.invoice_no,
            "invoice_date"    : _fmt_date(rec.invoice_date),
            "invoice_value"   : _indian_money(rec.invoice_value),
            "invoice_qty"     : rec.invoice_qty,
            "bill_no"         : rec.bill_no,
            "e_no"            : rec.e_no or "",
            "bg"              : bg,
            "bill_submission_date": _fmt_date(rec.bill_submission_date) or "-",
            "department"      : cust.get("department", ""),
            "customer_address": cust.get("bill_address", ""),
            "customer_district": cust.get("district", ""),
            "customer_state"  : cust.get("state_name", ""),
            "customer_pincode": cust.get("pin", ""),
            "customer_contact": cust.get("contact_number", ""),
            "customer_email"  : cust.get("email", ""),
            "existing_unique_id"    : sub.unique_id if sub else "",
            "existing_claim_amount" : sub.claim_amount if sub else "",
            "claim_amount_options"  : CLAIM_OPTIONS_MAP.get(ca_key, []),
        })

    # ──────────────────────────────────────────────────────────────────────────
    # 7. BILL DETAIL VIEW  (PHP view.php)
    # ──────────────────────────────────────────────────────────────────────────

    @action(detail=False, methods=["get"], url_path="view")
    def view_bill(self, request):
        unique_id = request.query_params.get("unique_id", "")
        bill_no   = request.query_params.get("bill_no", "")

        if not unique_id or not bill_no:
            return Response({"error": "unique_id and bill_no required"},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            main = BillSubmissionMain.objects.get(
                bill_form_main_unique_id=unique_id,
                bill_no=bill_no,
            )
        except BillSubmissionMain.DoesNotExist:
            return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)

        cust = _get_customer_detail(unique_id)

        # Invoice sub-table rows
        sub_rows = ViewBillSubmissionMain.objects.filter(
            bill_form_main_unique_id=unique_id,
            bill_no=bill_no,
            is_delete=False,
        ).values("invoice_no", "invoice_date", "invoice_qty", "invoice_value", "claimamt")

        invoices = []
        for inv in sub_rows:
            # Consignee details
            consignee_id = (
                InvoiceCreation.objects
                .filter(invoice_no=inv["invoice_no"])
                .values_list("consignee_id", flat=True)
                .first()
            )
            cons_data = {}
            if consignee_id:
                try:
                    with connection.cursor() as cur:
                        cur.callproc("GetConsigneeDetailsById", [consignee_id])
                        cols = [c[0] for c in cur.description]
                        row = cur.fetchone()
                        if row:
                            cons_data = dict(zip(cols, row))
                except Exception:
                    pass

            invoices.append({
                "invoice_no"       : inv["invoice_no"],
                "invoice_date"     : _fmt_date(inv["invoice_date"]),
                "invoice_qty"      : inv["invoice_qty"],
                "invoice_value"    : _indian_money(inv["invoice_value"]),
                "claimamt"         : _indian_money(inv["claimamt"]),
                "consignee_name"   : cons_data.get("con_contact_name", ""),
                "consignee_address": cons_data.get("con_address", ""),
                "consignee_district": cons_data.get("con_district", ""),
                "consignee_state"  : cons_data.get("con_state_name", ""),
                "consignee_pincode": cons_data.get("con_pincode", ""),
            })

        return Response({
            "department"      : cust.get("department", ""),
            "customer_address": cust.get("bill_address", ""),
            "customer_district": cust.get("district", ""),
            "customer_state"  : cust.get("state_name", ""),
            "customer_pincode": cust.get("pin", ""),
            "customer_contact": cust.get("contact_number", ""),
            "customer_email"  : cust.get("email", ""),
            "po_num"          : main.po_num,
            "po_date"         : _fmt_date(main.po_date),
            "bill_no"         : main.bill_no,
            "show_bg"         : main.partial_bill_status != "3",
            "bg_num"          : main.bg_num or "",
            "bg_date"         : _fmt_date(main.bg_date),
            "bg_value"        : _indian_money(main.bg_value),
            "bg_percen"       : getattr(main, "bg_percen", ""),
            "invoices"        : invoices,
        })

    # ──────────────────────────────────────────────────────────────────────────
    # 8. CREATE / UPDATE  (PHP case 'createupdate')
    # ──────────────────────────────────────────────────────────────────────────

    @action(detail=False, methods=["post"], url_path="create-update")
    def create_update(self, request):
        ser = SecurityDepositCreateUpdateSerializer(data=request.data)
        if not ser.is_valid():
            return Response({"status": False, "error": ser.errors, "msg": "error"},
                            status=status.HTTP_400_BAD_REQUEST)

        d = ser.validated_data
        claim_pct  = float(d["claim_amount"])
        inv_val    = float(d["invoice_value"])
        claim_amt  = (inv_val * claim_pct) / 100
        unique_id  = d.get("unique_id", "")

        columns = {
            "bill_form_main_unique_id": d["bill_form_main_unique_id"],
            "po_num"                  : d["po_num"],
            "po_date"                 : d["po_date"],
            "invoice_no"              : d["invoice_no"],
            "invoice_date"            : d["invoice_date"],
            "invoice_value"           : inv_val,
            "claim_amount"            : claim_pct,
            "partial_bill_status"     : 2,
            "bill_status"             : d["bill_status"],
            "invoice_qty"             : d.get("invoice_qty", ""),
            "claim_amt"               : claim_amt,
        }

        try:
            with transaction.atomic():
                # Update bill_submission_form partial_bill_status
                BillSubmissionForm.objects.filter(
                    bill_form_main_unique_id=d["bill_form_main_unique_id"],
                    invoice_no=d["invoice_no"],
                    is_delete=False,
                ).update(partial_bill_status=d["bill_status"])

                if unique_id:
                    BillSubmissionForm.objects.filter(unique_id=unique_id).update(**columns)
                    msg = "update"
                else:
                    columns["unique_id"] = _generate_unique_id()
                    BillSubmissionForm.objects.create(**columns)
                    msg = "create"

        except Exception as exc:
            return Response({"status": False, "error": str(exc), "msg": "error"},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({"status": True, "msg": msg, "error": ""})

    # ──────────────────────────────────────────────────────────────────────────
    # 9. PARTIAL BILL MAIN INSERT  (PHP case 'partial_bill_main')
    # ──────────────────────────────────────────────────────────────────────────

    @action(detail=False, methods=["post"], url_path="partial-bill-main")
    def partial_bill_main(self, request):
        ser = PartialBillMainSerializer(data=request.data)
        if not ser.is_valid():
            return Response({"status": False, "error": ser.errors, "msg": "error"},
                            status=status.HTTP_400_BAD_REQUEST)

        d = ser.validated_data
        bill_form_unique_id = d["bill_form_unique_id"]
        bg_nums  = d.get("bg_num", [])
        bg_dates = d.get("bg_date", [])
        bg_values= d.get("bg_value", [])
        invoice_nos = d["invoice_no"]

        try:
            with transaction.atomic():
                bill_no = _generate_bill_no()

                # Insert into bill_submission_main_table
                BillSubmissionMain.objects.create(
                    bill_form_main_unique_id=bill_form_unique_id,
                    po_num          = d["po_num"],
                    po_date         = d["po_date"],
                    claim_amount    = d["claim_percent"],
                    invoice_no      = invoice_nos[0] if invoice_nos else "",
                    invoice_date    = d["invoice_date"][0] if d["invoice_date"] else "",
                    invoice_value   = d["invoice_value"][0] if d["invoice_value"] else 0,
                    invoice_qty     = d["invoice_qty"][0] if d["invoice_qty"] else "",
                    bg_num          = bg_nums[0] if bg_nums else "NA",
                    bg_date         = bg_dates[0] if bg_dates else None,
                    bg_value        = bg_values[0] if bg_values else 0,
                    claimamt        = d["claim_value"],
                    partial_bill_status = 3,
                    bill_status     = 2,
                    bill_no         = bill_no,
                    bill_created_date = datetime.date.today(),
                    unique_id       = _generate_unique_id(),
                )

                # Update bg_creation_main if bg_num is provided and not 'NA'
                for i, inv_no in enumerate(invoice_nos):
                    bg_n = bg_nums[i] if i < len(bg_nums) else "NA"
                    if bg_n and bg_n != "NA":
                        BgCreationMain.objects.filter(
                            form_main_unique_id=bill_form_unique_id,
                            bg_num=bg_n,
                            invoice_no=inv_no,
                        ).update(bill_status=2)

        except Exception as exc:
            return Response({"status": False, "error": str(exc), "msg": "error"},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({"status": True, "msg": "create", "error": ""})

    # ──────────────────────────────────────────────────────────────────────────
    # 10. PARTIAL BILL SUB-TABLE INSERT  (PHP case 'partial_bill_creation_add_update')
    # ──────────────────────────────────────────────────────────────────────────

    @action(detail=False, methods=["post"], url_path="partial-bill-add")
    def partial_bill_creation_add_update(self, request):
        ser = PartialBillCreationAddUpdateSerializer(data=request.data)
        if not ser.is_valid():
            return Response({"status": False, "error": ser.errors, "msg": "error"},
                            status=status.HTTP_400_BAD_REQUEST)

        d = ser.validated_data
        bill_form_unique_id = d["bill_form_unique_id"]

        # Get the last bill_no from bill_submission_main_table
        agg = BillSubmissionMain.objects.filter(
            is_delete=False
        ).aggregate(count=Count("unique_id"), last=Max("bill_no"))

        if not agg["count"]:
            return Response({"status": True, "msg": "already", "error": ""})

        last_bill_no = agg["last"]

        try:
            with transaction.atomic():
                # Update sign_doc_verification_detail bill_no
                SignDocVerificationDetail.objects.filter(
                    form_main_unique_id=bill_form_unique_id,
                    invoice_no=d["invoice_no"],
                ).update(bill_no=last_bill_no)

                # Update bill_submission_sub sd_status
                BillSubmissionSub.objects.filter(
                    bill_form_unique_id=bill_form_unique_id,
                    invoice_no=d["invoice_no"],
                ).update(sd_status=1)

                # Insert into bill_submission_sub
                BillSubmissionSub.objects.create(
                    bill_form_unique_id  = bill_form_unique_id,
                    po_num               = d["po_num"],
                    po_date              = d["po_date"],
                    invoice_no           = d["invoice_no"],
                    invoice_date         = d["invoice_date"],
                    invoice_value        = d["invoice_value"],
                    invoice_qty          = d.get("invoice_qty", ""),
                    claimamt             = d["claim_amount"],
                    claim_percentage     = d["claim_percent"],
                    partial_bill_status  = 3,
                    sd_status            = 2,
                    bill_status          = 2,
                    bill_no              = last_bill_no,
                    bill_created_date    = datetime.date.today(),
                    unique_id            = _generate_unique_id(),
                )

                # Insert into bill_submission_form (main)
                BillSubmissionForm.objects.create(
                    bill_form_main_unique_id = bill_form_unique_id,
                    po_num               = d["po_num"],
                    po_date              = d["po_date"],
                    invoice_no           = d["invoice_no"],
                    invoice_date         = d["invoice_date"],
                    invoice_value        = d["invoice_value"],
                    invoice_qty          = d.get("invoice_qty", ""),
                    claim_amount         = d["claim_percent"],
                    claim_amt            = d["claim_amount"],
                    partial_bill_status  = 3,
                    bill_status          = 2,
                    sd_status            = 2,
                    bill_no              = last_bill_no,
                    unique_id            = _generate_unique_id(),
                )

        except Exception as exc:
            return Response({"status": False, "error": str(exc), "msg": "error"},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({"status": True, "msg": "create", "error": ""})

    # ──────────────────────────────────────────────────────────────────────────
    # 11. SOFT DELETE  (PHP case 'delete')
    # ──────────────────────────────────────────────────────────────────────────

    @action(detail=True, methods=["delete"], url_path="delete")
    def soft_delete(self, request, pk=None):
        try:
            # Soft-delete on view_payment_entry_list means deleting from the
            # underlying base table.  Adapt to your actual writable model.
            BillSubmissionSub.objects.filter(unique_id=pk).update(is_delete=True)
        except Exception as exc:
            return Response({"status": False, "error": str(exc), "msg": "error"},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response({"status": True, "msg": "success_delete", "error": ""})

    # ──────────────────────────────────────────────────────────────────────────
    # 12. EXCEL EXPORT  (PHP excel.php)
    # ──────────────────────────────────────────────────────────────────────────

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        """
        Mirrors excel.php: queries view_payment_entry_list and produces
        a Payment.xlsx with columns:
          S.No | PO Number | PO Date | Consignee | Location |
          Invoice No | Invoice Date | Invoice Value | Bill Submission Date |
          ENo | LD Days | LD Value | Payment Receivable
        """
        qs = ViewPaymentEntryList.objects.all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payment"

        ws.merge_cells("A1:M1")
        ws["A1"] = "PAYMENT"
        ws["A1"].font      = Font(bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = [
            "S.No", "PO Number", "PO Date", "Consignee", "Location",
            "Invoice No", "Invoice Date", "Invoice Value",
            "Bill Submission Date", "ENo", "LD Days", "LD Value",
            "Payment Receivable",
        ]
        for col, h in enumerate(headers, start=1):
            cell      = ws.cell(row=3, column=col, value=h)
            cell.font = Font(bold=True)

        for i, row in enumerate(qs, start=1):
            po_date = _fmt_date(row.po_date)
            ws.append([
                i,
                row.po_num,
                po_date,
                row.con_contact_name,
                row.con_address,
                row.invoice_no,
                _fmt_date(row.invoice_date),
                row.invoice_value,
                _fmt_date(row.bill_submission_date),
                row.e_no,
                row.ld_days,
                row.ld_amount,
                row.payement_receive,
            ])

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = HttpResponse(
            buf,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="Payment.xlsx"'
        return resp
