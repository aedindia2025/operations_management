"""
views.py
Django REST Framework views that replace the PHP crud.php switch-case actions.

Endpoint map
────────────
PHP action              → Django URL / method
─────────────────────────────────────────────────────────────────────
datatable               → GET  /api/payment/list/
payment_completed_…     → GET  /api/payment/list/?status=completed
bill_view               → GET  /api/payment/<unique_id>/detail/
createupdate            → POST /api/payment/save/          (with file)
bill_cancel_add_update  → POST /api/payment/cancel/
delete                  → DELETE /api/payment/<unique_id>/delete/
user_permission         → GET  /api/payment/permissions/
excel export            → GET  /api/payment/export/excel/
"""

import os
import io
import uuid
from datetime import date

from django.conf import settings
from django.db import connection, transaction
from django.db.models import F
from django.http import HttpResponse

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser

from master.apps.customer_payment.customer_payment_model import (
    PaymentBillSubmissionMain,
    PaymentBillSubmissionSub,
    PaymentConsigneeDetail,
    PaymentUserScreenPermission,
)
from master.serializers.customer_payment.customer_payment_serializer import (
    PaymentListSerializer,
    PaymentDetailSerializer,
    PaymentCreateUpdateSerializer,
    BillCancelSerializer,
)
from master.serializers.customer_payment.utils import indian_format, get_financial_year, disdate, make_payment_pdf_name


# ─────────────────────────────────────────────────────────────────────────────
# Helper
# ─────────────────────────────────────────────────────────────────────────────

def _ok(data=None, msg="success", extra=None):
    payload = {"status": 1, "msg": msg, "data": data or []}
    if extra:
        payload.update(extra)
    return Response(payload)


def _err(msg="error", error="", data=None):
    return Response(
        {"status": 0, "msg": msg, "error": error, "data": data or []},
        status=status.HTTP_400_BAD_REQUEST,
    )


def _build_row(idx: int, sub: PaymentBillSubmissionSub, main: PaymentBillSubmissionMain, consignee: PaymentConsigneeDetail | None) -> dict:
    """Build a single list-row dict that matches the PHP datatable output."""
    invoice_val = float(sub.invoice_value or 0)
    claim_pct   = float(sub.claim_percentage or 0)
    claim_amt   = round(invoice_val * claim_pct / 100)

    return {
        "s_no":            idx,
        "unique_id":       main.bill_form_main_unique_id,
        "bill_no":         sub.bill_no,
        "po_num":          main.po_num,
        "customer":        consignee.con_contact_name if consignee else "",
        "invoice_no":      sub.invoice_no,
        "invoice_value":   indian_format(invoice_val),
        "claim_percentage": f"{claim_pct:.0f}%",
        "claimamt":        indian_format(claim_amt),
        "payment_status":  main.payment_status,
        # raw values for filtering
        "_invoice_value_raw": invoice_val,
        "_claimamt_raw":      claim_amt,
    }


EXTRA_PAYMENT_UPLOADS = {
    "courier_pod_file": ("courier_pod_file", "courier_pod_file_org"),
    "dc_file": ("dc_file", "dc_file_org"),
    "einvoice_file": ("einvoice_file", "einvoice_file_org"),
    "ir_file": ("ir_file", "ir_file_org"),
    "bg_copy_file": ("bg_copy_file", "bg_copy_file_org"),
    "fire_insurance_file": ("fire_insurance_file", "fire_insurance_file_org"),
    "marine_insurance_file": ("marine_insurance_file", "marine_insurance_file_org"),
    "burglary_file": ("burglary_file", "burglary_file_org"),
}


def _ensure_payment_document_columns():
    required = {column for pair in EXTRA_PAYMENT_UPLOADS.values() for column in pair}
    for table_name in ("bill_submission_main_table", "bill_submission_sub"):
        with connection.cursor() as cur:
            cur.execute(f"SHOW COLUMNS FROM `{table_name}`")
            existing = {row[0] for row in cur.fetchall()}
            for column in sorted(required - existing):
                cur.execute(f"ALTER TABLE `{table_name}` ADD COLUMN `{column}` varchar(255) DEFAULT NULL")


def _ensure_payment_document_upload_table():
    column_sql = ",\n".join(
        f"                  `{column}` varchar(255) DEFAULT NULL"
        for pair in EXTRA_PAYMENT_UPLOADS.values()
        for column in pair
    )
    with connection.cursor() as cur:
        cur.execute(
            f"""
            CREATE TABLE IF NOT EXISTS `customer_payment_document_uploads` (
              `id` int(11) NOT NULL AUTO_INCREMENT,
              `unique_id` varchar(50) NOT NULL,
{column_sql},
              `is_active` int(11) NOT NULL DEFAULT 1,
              `is_delete` int(11) NOT NULL DEFAULT 0,
              `created` timestamp NOT NULL DEFAULT current_timestamp(),
              `updated` timestamp NOT NULL DEFAULT current_timestamp() ON UPDATE current_timestamp(),
              PRIMARY KEY (`id`),
              UNIQUE KEY `customer_payment_document_uploads_unique_id` (`unique_id`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci
            """
        )


def _save_payment_document_uploads(request, inv_no_value):
    upload_dir = os.path.join(settings.MEDIA_ROOT, "uploads", "payment_documents")
    os.makedirs(upload_dir, exist_ok=True)
    saved_fields = {}

    for input_name, (stored_col, original_col) in EXTRA_PAYMENT_UPLOADS.items():
        uploaded_file = request.FILES.get(input_name)
        if not uploaded_file:
            continue

        original_name = os.path.basename(uploaded_file.name)
        _, ext = os.path.splitext(original_name)
        safe_ext = ext.lower() or ".bin"
        file_name = f"{inv_no_value}-{input_name}-{uuid.uuid4().hex[:8]}{safe_ext}"
        dest = os.path.join(upload_dir, file_name)
        with open(dest, "wb") as fh:
            for chunk in uploaded_file.chunks():
                fh.write(chunk)

        saved_fields[stored_col] = file_name
        saved_fields[original_col] = original_name

    return saved_fields


def _update_payment_document_columns(table_name, fields, where_sql, where_params):
    if not fields:
        return
    assignments = ", ".join(f"`{column}` = %s" for column in fields.keys())
    params = list(fields.values()) + list(where_params)
    with connection.cursor() as cur:
        cur.execute(f"UPDATE `{table_name}` SET {assignments} WHERE {where_sql}", params)


def _insert_payment_document_upload(fields):
    unique_id = f"CPDOC{uuid.uuid4().hex[:12]}"
    columns = ["unique_id", *fields.keys()]
    placeholders = ", ".join(["%s"] * len(columns))
    column_sql = ", ".join(f"`{column}`" for column in columns)
    params = [unique_id, *fields.values()]
    with connection.cursor() as cur:
        cur.execute(
            f"INSERT INTO `customer_payment_document_uploads` ({column_sql}) VALUES ({placeholders})",
            params,
        )
    return unique_id


# ─────────────────────────────────────────────────────────────────────────────
# 1. Payment List  (replaces datatable / payment_completed_datatable)
# ─────────────────────────────────────────────────────────────────────────────

class PaymentListView(APIView):
    """
    GET /api/payment/list/
    Query params:
        status      = pending | completed   (default: pending)
        from_date   = YYYY-MM-DD
        to_date     = YYYY-MM-DD
        opt         = 4 (PO Date) | 5 (Invoice Date) | 7 (Bill Submission Date)
        search      = string
        start       = int   (DataTables offset)
        length      = int   (DataTables page size)
        draw        = int   (DataTables draw counter)
    """

    def get(self, request):
        tab        = request.query_params.get("status", "pending")
        from_date  = request.query_params.get("from_date", "")
        to_date    = request.query_params.get("to_date", "")
        opt        = request.query_params.get("opt", "")
        search_val = request.query_params.get("search", "")
        draw       = int(request.query_params.get("draw", 1))
        start      = int(request.query_params.get("start", 0))
        length     = int(request.query_params.get("length", 10))

        # payment_status: 0 = pending, 1 = completed
        pmt_status = 1 if tab == "completed" else 0

        qs = PaymentBillSubmissionSub.objects.filter(
            is_delete=False,
            payment_status=pmt_status,
        )

        # Date filter
        if from_date and to_date and opt:
            direct_map = {
                "5": ("invoice_date__gte", "invoice_date__lte"),
                "7": ("bill_submission_date__gte", "bill_submission_date__lte"),
            }
            if opt in direct_map:
                qs = qs.filter(**{
                    direct_map[opt][0]: from_date,
                    direct_map[opt][1]: to_date,
                })
            elif opt == "4":
                # PO date lives on the main table - filter via bill_form_unique_id
                main_ids = PaymentBillSubmissionMain.objects.filter(
                    po_date__gte=from_date,
                    po_date__lte=to_date,
                    is_delete=False,
                ).values_list("bill_form_main_unique_id", flat=True)
                qs = qs.filter(bill_form_unique_id__in=main_ids)

        subs = list(qs.order_by("id"))

        main_ids = {sub.bill_form_unique_id for sub in subs}
        mains = {
            m.bill_form_main_unique_id: m
            for m in PaymentBillSubmissionMain.objects.filter(
                bill_form_main_unique_id__in=main_ids,
                is_delete=False,
            )
        }

        consignee_ids = {
            getattr(m, "consignee_unique_id", None)
            for m in mains.values()
            if getattr(m, "consignee_unique_id", None)
        }
        consignees = {
            c.unique_id: c
            for c in PaymentConsigneeDetail.objects.filter(unique_id__in=consignee_ids)
        }

        all_rows = []
        for sub in subs:
            main = mains.get(sub.bill_form_unique_id)
            if not main:
                continue
            consignee = consignees.get(getattr(main, "consignee_unique_id", None))
            all_rows.append(_build_row(0, sub, main, consignee))

        records_total = len(all_rows)

        if search_val:
            needle = search_val.lower()
            all_rows = [
                row for row in all_rows
                if needle in (f"{row['bill_no']} {row['invoice_no']} {row['customer']} {row['po_num']}".lower())
            ]

        filtered_count = len(all_rows)
        paged = all_rows[start: start + length]
        for idx, row in enumerate(paged, start=start + 1):
            row["s_no"] = idx

        return Response({
            "draw":            draw,
            "recordsTotal":    records_total,
            "recordsFiltered": filtered_count,
            "data":            paged,
        })


# ─────────────────────────────────────────────────────────────────────────────
# 2. Payment Detail  (replaces bill_view / view modal)
# ─────────────────────────────────────────────────────────────────────────────

class PaymentDetailView(APIView):
    """
    GET /api/payment/detail/
    Query params: unique_id, bill_no, invoice_no
    Returns a single dict that populates the modal (same keys as the PHP JS).
    """

    def get(self, request):
        unique_id  = request.query_params.get("unique_id", "")
        bill_no    = request.query_params.get("bill_no", "")
        invoice_no = request.query_params.get("invoice_no", "")

        if not (unique_id and bill_no and invoice_no):
            return _err("Missing parameters", "unique_id, bill_no and invoice_no are required")

        sub = PaymentBillSubmissionSub.objects.filter(
            bill_form_unique_id=unique_id,
            bill_no=bill_no,
            invoice_no=invoice_no,
            is_delete=False,
        ).first()

        if not sub:
            return _err("Record not found")

        main = PaymentBillSubmissionMain.objects.filter(
            bill_form_main_unique_id=unique_id,
            is_delete=False,
        ).first()

        consignee = PaymentConsigneeDetail.objects.filter(
            unique_id=getattr(main, "consignee_unique_id", None)
        ).first() if main else None

        # Compute derived values (mirrors PHP view logic)
        invoice_val   = float(sub.invoice_value or 0)
        claim_pct     = float(sub.claim_percentage or 0)
        claim_amt     = round(invoice_val * claim_pct / 100)
        gst_val       = float(sub.gst_value or 0)
        tds_val       = float(sub.tds_value or 0)
        ld_val        = float(sub.ld or 0)
        pmt_received  = float(sub.payement_receive or 0)
        tran_amt      = float(sub.tran_amt or 0)
        ttl_amount    = gst_val + tds_val + ld_val + tran_amt
        balance       = invoice_val - pmt_received

        row = {
            "bill_no":             sub.bill_no,
            "invoice_no":          sub.invoice_no,
            "customer":            consignee.con_contact_name if consignee else "",
            "address":             consignee.con_address if consignee else "",
            "pincode":             consignee.pincode if consignee else "",
            "contact_name":        consignee.con_contact_name if consignee else "",
            "contact_number":      consignee.contact_number if consignee else "",
            "email":               consignee.email if consignee else "",
            "bill_created_date":   disdate(main.created_at.date() if main else None),
            "bill_submission_date": disdate(sub.bill_submission_date),
            "bill_num":            sub.bill_no,
            "po_num":              main.po_num if main else "",
            "po_date":             disdate(main.po_date if main else None),
            "invoice_date":        disdate(sub.invoice_date),
            "invoice_value":       indian_format(invoice_val),
            "trans_date":          disdate(sub.payment_date),
            "trans_id":            "",
            "tran_amt":            indian_format(tran_amt),
            "payment_received":    indian_format(pmt_received),
            "ld_amount":           indian_format(float(sub.ld_amount or 0)),
            "gst_value":           indian_format(gst_val),
            "tds_value":           indian_format(tds_val),
            "payment_date":        disdate(sub.payment_date),
            "claimamt":            indian_format(claim_amt),
            "e_no":                sub.e_no,
            "district":            consignee.district if consignee else "",
            "state_name":          consignee.state_name if consignee else "",
            "ttl_amount":          indian_format(ttl_amount),
            "balance_amount":      indian_format(balance),
        }

        return _ok(data=[row])


# ─────────────────────────────────────────────────────────────────────────────
# 3. Create / Update  (replaces crud.php → case 'createupdate')
# ─────────────────────────────────────────────────────────────────────────────

class PaymentSaveView(APIView):
    """
    POST /api/payment/save/
    Accepts multipart/form-data (file upload) or JSON.
    Mirrors the PHP createupdate switch case exactly.
    """
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def post(self, request):
        data = request.data
        if str(data.get("upload_only", "")).lower() in {"1", "true", "yes"}:
            extra_upload_fields = _save_payment_document_uploads(request, "CustomerPayment")
            if not extra_upload_fields:
                return _err("Validation failed", "Please upload at least one document.")
            _ensure_payment_document_upload_table()
            upload_id = _insert_payment_document_upload(extra_upload_fields)
            return _ok(msg="create", extra={"upload_id": upload_id})

        ser  = PaymentCreateUpdateSerializer(data=data)
        if not ser.is_valid():
            return _err("Validation failed", str(ser.errors))

        v          = ser.validated_data
        unique_id  = v["unique_id"]
        bill_no    = v["bill_no"]
        my_inv_no  = v["my_inv_no"]
        acc_year   = get_financial_year()

        # Derive the PDF filename the same way PHP does
        # $parts = explode('/', $myinv_no);  $myinv_no_value = end($parts);
        inv_parts      = my_inv_no.split("/")
        inv_no_value   = inv_parts[-1] if inv_parts else my_inv_no
        file_name      = ""
        file_org_name  = ""

        uploaded_file = request.FILES.get("file")
        if uploaded_file:
            ext = os.path.splitext(uploaded_file.name)[1].lower()
            if uploaded_file.content_type == "application/pdf" and ext == ".pdf":
                file_name     = f"Payment-{inv_no_value}.pdf"
                file_org_name = uploaded_file.name
                upload_dir    = os.path.join(settings.MEDIA_ROOT, "uploads", "payment")
                os.makedirs(upload_dir, exist_ok=True)
                dest = os.path.join(upload_dir, file_name)
                with open(dest, "wb") as fh:
                    for chunk in uploaded_file.chunks():
                        fh.write(chunk)

        extra_upload_fields = _save_payment_document_uploads(request, inv_no_value)
        if extra_upload_fields:
            _ensure_payment_document_columns()

        with transaction.atomic():
            # ── Update bill_submission_main_table ─────────────────────────
            main_fields = {
                "payement_receive": v["payement_receive"],
                "payment_status":   v["payment_status"],
                "payment_date":     v["payment_date"],
                "ld_amount":        v["ld_amount"],
                "ld_days":          v["ld_days"],
                "acc_year":         acc_year,
            }
            if file_name:
                main_fields["file_name"]     = file_name
                main_fields["file_org_name"] = file_org_name

            updated_main = PaymentBillSubmissionMain.objects.filter(
                bill_form_main_unique_id=unique_id,
                bill_no=bill_no,
            ).update(**main_fields)

            # ── Update bill_submission_sub ─────────────────────────────────
            claim_pct = v["claim_amount"]
            sub_fields = {
                "payement_receive": v["payement_receive"],
                "payment_status":   v["payment_status"],
                "payment_date":     v["payment_date"],
                "ld_amount":        v["ld_amount"],
                "ld_days":          v["ld_days"],
                "claim_percentage": claim_pct,
                "gst":              v["gst"],
                "gst_value":        v["gst_value"],
                "tds":              v["tds"],
                "tds_value":        v["tds_value"],
                "ld":               v["ld"],
                "tran_amt":         v["tran_amt"],
                "rem_amt":          v["rem_amt"],
                "acc_year":         acc_year,
            }
            if file_name:
                sub_fields["file_name"]     = file_name
                sub_fields["file_org_name"] = file_org_name

            sub_qs = PaymentBillSubmissionSub.objects.filter(
                bill_form_unique_id=unique_id,
                bill_no=bill_no,
            )
            sub_qs.update(**sub_fields)
            _update_payment_document_columns(
                "bill_submission_main_table",
                extra_upload_fields,
                "`bill_form_main_unique_id` = %s AND `bill_no` = %s",
                [unique_id, bill_no],
            )
            _update_payment_document_columns(
                "bill_submission_sub",
                extra_upload_fields,
                "`bill_form_unique_id` = %s AND `bill_no` = %s",
                [unique_id, bill_no],
            )
            for sub_row in sub_qs:
                invoice_val = float(sub_row.invoice_value or 0)
                sub_row.claim_amount = round(invoice_val * float(claim_pct or 0) / 100, 2)
                sub_row.save(update_fields=["claim_amount", "updated_at"])

        if not updated_main:
            return _err("Record not found or not updated")

        return _ok(msg="update")


# ─────────────────────────────────────────────────────────────────────────────
# 4. Bill Cancel  (replaces case 'bill_cancel_add_update')
# ─────────────────────────────────────────────────────────────────────────────

class BillCancelView(APIView):
    """
    POST /api/payment/cancel/
    """

    def post(self, request):
        ser = BillCancelSerializer(data=request.data)
        if not ser.is_valid():
            return _err("Validation failed", str(ser.errors))

        v         = ser.validated_data
        unique_id = v["unique_id"]
        bill_no   = v["bill_no"]
        invoice_no = v["invoice_no"]

        updated = PaymentBillSubmissionSub.objects.filter(
            bill_form_unique_id=unique_id,
            bill_no=bill_no,
            invoice_no=invoice_no,
        ).update(is_delete=True)

        if not updated:
            return _err("Record not found", "already")

        return _ok(msg="success")


# ─────────────────────────────────────────────────────────────────────────────
# 5. Delete  (replaces case 'delete')
# ─────────────────────────────────────────────────────────────────────────────

class PaymentDeleteView(APIView):
    """
    DELETE /api/payment/<unique_id>/delete/
    Soft-deletes by setting is_delete=True on the main record.
    """

    def delete(self, request, unique_id):
        with transaction.atomic():
            updated_main = PaymentBillSubmissionMain.objects.filter(
                bill_form_main_unique_id=unique_id
            ).update(is_delete=True)
            PaymentBillSubmissionSub.objects.filter(
                bill_form_unique_id=unique_id
            ).update(is_delete=True)

        if not updated_main:
            return _err("Record not found", error="not found")

        return _ok(msg="success_delete")


# ─────────────────────────────────────────────────────────────────────────────
# 6. User Permission  (replaces case 'user_permission')
# ─────────────────────────────────────────────────────────────────────────────

class UserPermissionView(APIView):
    """
    GET /api/payment/permissions/?screen_id=<id>&user_type=<type>
    Returns a list of action_unique_ids the user is allowed to perform.
    """

    def get(self, request):
        screen_id = request.query_params.get("screen_id", "")
        user_type = request.query_params.get("user_type", "")

        perms = PaymentUserScreenPermission.objects.filter(
            screen_unique_id=screen_id,
            user_type=user_type,
            is_delete=0,
        ).values_list("action_unique_id", flat=True)

        return _ok(data=list(perms))


# ─────────────────────────────────────────────────────────────────────────────
# 7. Excel Export  (replaces excel.php)
# ─────────────────────────────────────────────────────────────────────────────

class PaymentExcelExportView(APIView):
    """
    GET /api/payment/export/excel/
    Streams an .xlsx file using openpyxl (replaces PHPExcel).
    Columns: S.No | PO Number | PO Date | Consignee | Location |
             Invoice No | Invoice Date | Invoice Value | Bill Submission Date |
             ENo | LD Days | LD Value | Payment Receivable
    """

    def get(self, request):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError:
            return Response(
                {"error": "openpyxl not installed. Run: pip install openpyxl"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payment"

        # Row 1 – Merged title (mirrors A1:M1 in PHP)
        ws.merge_cells("A1:M1")
        ws["A1"] = "PAYMENT"
        ws["A1"].font = Font(bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Row 3 – Headers
        headers = [
            "S.No", "PO Number", "PO Date", "Consignee", "Location",
            "Invoice No", "Invoice Date", "Invoice Value",
            "Bill Submission Date", "ENo", "LD Days", "LD Value",
            "Payment Receivable",
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)

        # Data rows starting at row 4
        subs = list(PaymentBillSubmissionSub.objects.filter(is_delete=False).order_by("id"))
        main_ids = {sub.bill_form_unique_id for sub in subs}
        mains = {
            m.bill_form_main_unique_id: m
            for m in PaymentBillSubmissionMain.objects.filter(
                bill_form_main_unique_id__in=main_ids,
                is_delete=False,
            )
        }
        consignee_ids = {
            getattr(m, "consignee_unique_id", None)
            for m in mains.values()
            if getattr(m, "consignee_unique_id", None)
        }
        consignees = {
            c.unique_id: c
            for c in PaymentConsigneeDetail.objects.filter(unique_id__in=consignee_ids)
        }

        row_num = 1
        for sub in subs:
            main = mains.get(sub.bill_form_unique_id)
            if not main:
                continue
            consignee = consignees.get(getattr(main, "consignee_unique_id", None))

            ws.append([
                row_num,
                main.po_num if main else "",
                disdate(main.po_date if main else None),
                consignee.con_contact_name if consignee else "",
                consignee.con_address if consignee else "",
                sub.invoice_no,
                disdate(sub.invoice_date),
                float(sub.invoice_value or 0),
                disdate(sub.bill_submission_date),
                sub.e_no,
                sub.ld_days,
                float(sub.ld_amount or 0),
                float(sub.payement_receive or 0),
            ])
            row_num += 1

        # Auto-width for all columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # Stream the file
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="Payment.xlsx"'
        return response
