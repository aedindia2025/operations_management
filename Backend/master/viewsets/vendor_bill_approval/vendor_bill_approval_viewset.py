from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.conf import settings
from django.db import connection
from django.http import HttpResponse, FileResponse, Http404
import openpyxl
from openpyxl.styles import Font, Alignment
import io
import os
from pathlib import Path
from urllib.parse import quote


LEGACY_UPLOAD_BASE = Path(
    os.environ.get("OTM_UPLOAD_BASE", r"Z:\xampp\htdocs\otm_beta\uploads")
)


_COLUMN_CACHE = {}


def db_fetch(sql, params=None):
    with connection.cursor() as cursor:
        cursor.execute(sql, params or [])
        cols = [c[0] for c in cursor.description]
        return [dict(zip(cols, row)) for row in cursor.fetchall()]


def db_fetchone(sql, params=None):
    with connection.cursor() as cursor:
        cursor.execute(sql, params or [])
        cols = [c[0] for c in cursor.description]
        row = cursor.fetchone()
        return dict(zip(cols, row)) if row else None


def table_columns(table_name):
    cached = _COLUMN_CACHE.get(table_name)
    if cached is not None:
        return cached

    try:
        with connection.cursor() as cursor:
            cursor.execute(f"SHOW COLUMNS FROM `{table_name}`")
            cols = {row[0] for row in cursor.fetchall()}
    except Exception:
        cols = set()

    _COLUMN_CACHE[table_name] = cols
    return cols


def pick_col(columns, *candidates):
    for col in candidates:
        if col and col in columns:
            return col
    return None


def safe_float(value):
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def safe_str(value):
    return "" if value is None else str(value)


def vendor_bill_file_url(request, file_kind, filename):
    value = safe_str(filename).strip()
    if not value or value == "#":
        return ""
    if value.startswith(("http://", "https://", "/api/", "/media/")):
        return value
    safe_name = Path(value).name
    return f"/api/master/vendor-bill-approval/files/{file_kind}/{quote(safe_name)}/"


def vendor_bill_file_candidates(file_kind, filename):
    safe_name = Path(safe_str(filename)).name
    if not safe_name:
        return []

    folder_map = {
        "invoice": [
            LEGACY_UPLOAD_BASE / "vendorpayment",
            Path(settings.MEDIA_ROOT) / "vendorpayment",
            Path(settings.BASE_DIR) / "uploads" / "vendorpayment",
        ],
        "po": [
            LEGACY_UPLOAD_BASE / "po_form" / "PO copy",
            LEGACY_UPLOAD_BASE / "po_form",
            LEGACY_UPLOAD_BASE / "purchase_order" / "PO copy",
            LEGACY_UPLOAD_BASE / "purchase_order",
            LEGACY_UPLOAD_BASE / "vendorpayment" / "PO copy",
            LEGACY_UPLOAD_BASE / "vendorpayment",
            Path(settings.MEDIA_ROOT) / "po_form" / "PO copy",
            Path(settings.MEDIA_ROOT) / "po_form",
            Path(settings.MEDIA_ROOT) / "purchase_order" / "PO copy",
            Path(settings.MEDIA_ROOT) / "purchase_order",
            Path(settings.MEDIA_ROOT) / "vendorpayment" / "PO copy",
            Path(settings.MEDIA_ROOT) / "vendorpayment",
            Path(settings.BASE_DIR) / "uploads" / "po_form" / "PO copy",
            Path(settings.BASE_DIR) / "uploads" / "po_form",
            Path(settings.BASE_DIR) / "uploads" / "purchase_order" / "PO copy",
            Path(settings.BASE_DIR) / "uploads" / "purchase_order",
            Path(settings.BASE_DIR) / "uploads" / "vendorpayment" / "PO copy",
            Path(settings.BASE_DIR) / "uploads" / "vendorpayment",
        ],
        "pan": [
            LEGACY_UPLOAD_BASE / "vendor_creation",
            Path(settings.MEDIA_ROOT) / "vendor_creation",
            Path(settings.BASE_DIR) / "uploads" / "vendor_creation",
        ],
        "bank": [
            LEGACY_UPLOAD_BASE / "vendor_creation",
            Path(settings.MEDIA_ROOT) / "vendor_creation",
            Path(settings.BASE_DIR) / "uploads" / "vendor_creation",
        ],
    }

    return [root / safe_name for root in folder_map.get(file_kind, [])]

def parsed_date_expr(col):
    if not col:
        return "NULL"
    return (
        f"COALESCE("
        f"STR_TO_DATE(CAST({col} AS CHAR), '%%Y-%%m-%%d %%H:%%i:%%s'), "
        f"STR_TO_DATE(CAST({col} AS CHAR), '%%d-%%m-%%Y %%H:%%i:%%s'), "
        f"STR_TO_DATE(CAST({col} AS CHAR), '%%Y-%%m-%%d'), "
        f"STR_TO_DATE(CAST({col} AS CHAR), '%%d-%%m-%%Y'), "
        f"STR_TO_DATE(CAST({col} AS CHAR), '%%d/%%m/%%Y'), "
        f"DATE({col})"
        f")"
    )


def status_text(sts):
    if str(sts) == "1":
        return "Approved"
    if str(sts) == "2":
        return "Rejected"
    return "Pending"


def soft_delete_filter(alias, columns):
    delete_col = pick_col(columns, "is_delete")
    if not delete_col:
        return ""
    return f" AND COALESCE({alias}.{delete_col}, 0) = 0"


def max_expr(col, alias_name, cast_text=False):
    if not col:
        return f"'' AS {alias_name}"
    if cast_text:
        return f"MAX(CAST({col} AS CHAR)) AS {alias_name}"
    return f"MAX({col}) AS {alias_name}"


def max_date_expr(col, alias_name, fmt):
    if not col:
        return f"'' AS {alias_name}"
    safe_fmt = fmt.replace("%", "%%")
    normalized = parsed_date_expr(col)
    return (
        f"COALESCE("
        f"DATE_FORMAT(MAX({normalized}), '{safe_fmt}'), "
        f"MAX(CAST({col} AS CHAR)), "
        f"''"
        f") AS {alias_name}"
    )


def date_expr(col, alias_name, fmt):
    if not col:
        return f"'' AS {alias_name}"
    safe_fmt = fmt.replace("%", "%%")
    normalized = parsed_date_expr(col)
    return (
        f"COALESCE("
        f"DATE_FORMAT({normalized}, '{safe_fmt}'), "
        f"CAST({col} AS CHAR), "
        f"''"
        f") AS {alias_name}"
    )


def staff_name_max_expr(col, alias_name):
    if not col:
        return f"'' AS {alias_name}"
    return (
        f"CASE WHEN MAX({col}) IS NULL OR MAX({col}) = '' THEN '' "
        f"ELSE COALESCE(get_staff_name(MAX({col})), CAST(MAX({col}) AS CHAR), '') END AS {alias_name}"
    )


def staff_name_expr(col, alias_name):
    if not col:
        return f"'' AS {alias_name}"
    return (
        f"CASE WHEN {col} IS NULL OR {col} = '' THEN '' "
        f"ELSE COALESCE(get_staff_name({col}), CAST({col} AS CHAR), '') END AS {alias_name}"
    )


def get_user_id(request):
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        return auth_header.split(" ", 1)[1].strip()
    return safe_str(request.data.get("user_id"))


class VendorBillListView(APIView):
    # /vendor-bill-approval/list/?tab=pending&from_date=&to_date=&search=&page=1&length=10
    def get(self, request):
        tab = request.query_params.get("tab", "pending").lower()
        from_date = request.query_params.get("from_date", "")
        to_date = request.query_params.get("to_date", "")
        search = request.query_params.get("search", "")

        try:
            page = max(1, int(request.query_params.get("page", 1)))
        except Exception:
            page = 1
        try:
            length = max(1, int(request.query_params.get("length", 10)))
        except Exception:
            length = 10

        offset = (page - 1) * length

        vpd_cols = table_columns("vendor_payment_details")
        vc_cols = table_columns("vendor_creation")
        if not vpd_cols or "bill_no" not in vpd_cols:
            return Response({"data": [], "total": 0, "page": page, "pages": 1})

        status_col = pick_col(vpd_cols, "vendor_bill_app_status")
        bill_date_col = pick_col(vpd_cols, "bill_date", "vendor_bill_created_date", "created")
        invoice_date_col = pick_col(vpd_cols, "vendor_inv_attach_approval_date", "invoice_date")
        invoice_no_col = pick_col(vpd_cols, "veninvverifyid", "invoice_no")
        vendor_name_col = pick_col(vpd_cols, "vendor_name")
        dc_col = pick_col(vpd_cols, "dc_num")
        amount_col = pick_col(vpd_cols, "amount", "total_amount")
        additional_charges_col = pick_col(vpd_cols, "additionalcharges", "additional_charges")
        vendor_id_col = pick_col(vpd_cols, "vendor_id")
        remark_col = pick_col(vpd_cols, "bill_remark")
        reject_reason_col = pick_col(vpd_cols, "vendor_bill_reject_reason", "reject_reason")
        approve_by_col = pick_col(vpd_cols, "vendor_bill_approval", "bill_approved")
        approve_date_col = pick_col(vpd_cols, "vendor_bill_approval_date")
        rejected_by_col = pick_col(vpd_cols, "vendor_bill_rejected_by")
        acc_entry_status_col = pick_col(vpd_cols, "acc_ent_sts", "accstatus")
        acc_entry_reject_col = pick_col(vpd_cols, "acc_ent_rej_reason")
        acc_entry_rejected_by_col = pick_col(vpd_cols, "acc_rejected_by", "vendor_account_approved_by")
        order_col = pick_col(vpd_cols, "id", "unique_id", "bill_date")

        join_vendor = ""
        vendor_name_expr = f"vb.{vendor_name_col}" if vendor_name_col else "''"
        vendor_address_expr = "''"
        vendor_phone_expr = "''"
        if vc_cols and vendor_id_col and "unique_id" in vc_cols:
            vc_name_col = pick_col(vc_cols, "company_name", "vendor_name", "name")
            vc_address_col = pick_col(vc_cols, "address")
            vc_phone_col = pick_col(vc_cols, "contact_no")
            join_vendor = f"LEFT JOIN vendor_creation v ON v.unique_id = vb.{vendor_id_col} AND COALESCE(v.is_delete, 0) = 0"
            if vc_name_col and vendor_name_col:
                vendor_name_expr = f"COALESCE(v.{vc_name_col}, vb.{vendor_name_col})"
            elif vc_name_col:
                vendor_name_expr = f"v.{vc_name_col}"
            if vc_address_col:
                vendor_address_expr = f"COALESCE(v.{vc_address_col}, '')"
            if vc_phone_col:
                vendor_phone_expr = f"COALESCE(v.{vc_phone_col}, '')"

        rejected_actor_expr = None
        if rejected_by_col and approve_by_col:
            rejected_actor_expr = f"COALESCE(NULLIF(vb.{rejected_by_col}, ''), vb.{approve_by_col})"
        elif rejected_by_col:
            rejected_actor_expr = f"vb.{rejected_by_col}"
        elif approve_by_col:
            rejected_actor_expr = f"vb.{approve_by_col}"

        status_expr = f"COALESCE(vb.{status_col}, 0)" if status_col else "0"
        reject_reason_expr = f"COALESCE(vb.{reject_reason_col}, '')" if reject_reason_col else "''"
        rejected_actor_display_expr = rejected_actor_expr or "''"

        if tab == "pending" and acc_entry_reject_col:
            status_expr = (
                f"CASE "
                f"WHEN COALESCE(vb.{acc_entry_reject_col}, '') <> '' THEN 2 "
                f"ELSE {status_expr} "
                f"END"
            )
            reject_reason_expr = (
                f"CASE "
                f"WHEN COALESCE(vb.{acc_entry_reject_col}, '') <> '' THEN COALESCE(vb.{acc_entry_reject_col}, '') "
                f"ELSE {reject_reason_expr} "
                f"END"
            )
            if acc_entry_rejected_by_col:
                rejected_actor_display_expr = (
                    f"CASE "
                    f"WHEN COALESCE(vb.{acc_entry_reject_col}, '') <> '' "
                    f"THEN COALESCE(NULLIF(vb.{acc_entry_rejected_by_col}, ''), '') "
                    f"ELSE {rejected_actor_display_expr} "
                    f"END"
                )

        bill_date_filter_expr = parsed_date_expr(f"vb.{bill_date_col}") if bill_date_col else None

        where_parts = ["1=1"]
        params = []

        delete_col = pick_col(vpd_cols, "is_delete")
        if delete_col:
            where_parts.append(f"COALESCE(vb.{delete_col}, 0) = 0")

        if status_col:
            if tab == "pending":
                where_parts.append(f"COALESCE(vb.{status_col}, 0) = 0")
            else:
                where_parts.append(f"COALESCE(vb.{status_col}, 0) IN (1, 2)")

        if from_date and bill_date_filter_expr:
            where_parts.append(f"{bill_date_filter_expr} >= %s")
            params.append(from_date)
        if to_date and bill_date_filter_expr:
            where_parts.append(f"{bill_date_filter_expr} <= %s")
            params.append(to_date)

        if search:
            search_like = f"%{search}%"
            search_parts = ["vb.bill_no LIKE %s"]
            params.append(search_like)
            if invoice_no_col:
                search_parts.append(f"CAST(vb.{invoice_no_col} AS CHAR) LIKE %s")
                params.append(search_like)
            if vendor_name_expr != "''":
                search_parts.append(f"CAST({vendor_name_expr} AS CHAR) LIKE %s")
                params.append(search_like)
            where_parts.append("(" + " OR ".join(search_parts) + ")")

        where_sql = " AND ".join(where_parts)

        count_sql = f"""
            SELECT COUNT(DISTINCT vb.bill_no)
            FROM vendor_payment_details vb
            {join_vendor}
            WHERE {where_sql}
        """

        with connection.cursor() as cur:
            cur.execute(count_sql, params)
            total = cur.fetchone()[0] or 0

        grand_total_expr = "0"
        if amount_col and additional_charges_col:
            grand_total_expr = (
                f"ROUND(IFNULL(SUM(vb.{amount_col}), 0) + MAX(IFNULL(vb.{additional_charges_col}, 0)), 2)"
            )
        elif amount_col:
            grand_total_expr = f"ROUND(IFNULL(SUM(vb.{amount_col}), 0), 2)"
        elif additional_charges_col:
            grand_total_expr = f"ROUND(MAX(IFNULL(vb.{additional_charges_col}, 0)), 2)"

        data_sql = f"""
            SELECT
                vb.bill_no,
                {max_date_expr(f'vb.{bill_date_col}', 'bill_date', '%d-%m-%Y') if bill_date_col else "'' AS bill_date"},
                {max_date_expr(f'vb.{invoice_date_col}', 'invoice_date', '%d-%m-%Y') if invoice_date_col else "'' AS invoice_date"},
                {max_expr(f'vb.{invoice_no_col}', 'invoice_no', cast_text=True) if invoice_no_col else "'' AS invoice_no"},
                {f"MAX(CAST({vendor_name_expr} AS CHAR)) AS vendor_name" if vendor_name_expr != "''" else "'' AS vendor_name"},
                {f"MAX(CAST({vendor_address_expr} AS CHAR)) AS vendor_address" if vendor_address_expr != "''" else "'' AS vendor_address"},
                {f"MAX(CAST({vendor_phone_expr} AS CHAR)) AS vendor_phone" if vendor_phone_expr != "''" else "'' AS vendor_phone"},
                {f"COUNT(DISTINCT vb.{dc_col}) AS dc_count" if dc_col else "0 AS dc_count"},
                {grand_total_expr} AS total_amount,
                MAX({status_expr}) AS vendor_bill_app_status,
                {staff_name_max_expr(f'vb.{approve_by_col}', 'approved_by') if approve_by_col else "'' AS approved_by"},
                {max_date_expr(f'vb.{approve_date_col}', 'approved_date', '%d-%m-%Y') if approve_date_col else "'' AS approved_date"},
                {staff_name_max_expr(rejected_actor_display_expr, 'rejected_by') if rejected_actor_display_expr else "'' AS rejected_by"},
                MAX(CAST({reject_reason_expr} AS CHAR)) AS reject_reason,
                {max_expr(f'vb.{remark_col}', 'bill_remark', cast_text=True) if remark_col else "'' AS bill_remark"},
                {max_expr(f'vb.{vendor_id_col}', 'vendor_id', cast_text=True) if vendor_id_col else "'' AS vendor_id"}
            FROM vendor_payment_details vb
            {join_vendor}
            WHERE {where_sql}
            GROUP BY vb.bill_no
            ORDER BY {f'MAX(vb.{order_col}) DESC' if order_col else 'vb.bill_no DESC'}
            LIMIT %s OFFSET %s
        """

        rows = db_fetch(data_sql, params + [length, offset])

        bills = []
        for row in rows:
            sts = str(row.get("vendor_bill_app_status", "0"))
            bills.append(
                {
                    "bill_no": safe_str(row.get("bill_no")),
                    "bill_date": safe_str(row.get("bill_date")),
                    "invoice_date": safe_str(row.get("invoice_date")),
                    "invoice_no": safe_str(row.get("invoice_no")),
                    "vendor_name": safe_str(row.get("vendor_name")),
                    "vendor_address": safe_str(row.get("vendor_address")),
                    "vendor_phone": safe_str(row.get("vendor_phone")),
                    "dc_count": int(row.get("dc_count") or 0),
                    "amount": safe_float(row.get("total_amount")),
                    "status": status_text(sts),
                    "vendor_id": safe_str(row.get("vendor_id")),
                    "approved_by": safe_str(row.get("approved_by")),
                    "approved_date": safe_str(row.get("approved_date")),
                    "rejected_by": safe_str(row.get("rejected_by")),
                    "reject_reason": safe_str(row.get("reject_reason")),
                    "bill_remark": safe_str(row.get("bill_remark")),
                }
            )

        return Response(
            {
                "data": bills,
                "total": total,
                "page": page,
                "pages": max(1, -(-total // length)),
            }
        )


class VendorBillDetailView(APIView):
    # /vendor-bill-approval/detail/?bill_no=VEN-xxxx
    def get(self, request):
        bill_no = request.query_params.get("bill_no", "")
        if not bill_no:
            return Response({"error": "bill_no required"}, status=status.HTTP_400_BAD_REQUEST)

        vpd_cols = table_columns("vendor_payment_details")
        vc_cols = table_columns("vendor_creation")

        if "bill_no" not in vpd_cols:
            return Response({"error": "Bill not found"}, status=status.HTTP_404_NOT_FOUND)

        vendor_id_col = pick_col(vpd_cols, "vendor_id")

        # Vendor fields (fallback-safe)
        vc_vendor_name = pick_col(vc_cols, "vendor_name", "company_name", "name")
        vc_contact_person = pick_col(vc_cols, "contact_person", "name")
        vc_address = pick_col(vc_cols, "address")
        vc_gst = pick_col(vc_cols, "gst_no")
        vc_pan = pick_col(vc_cols, "pan_no")
        vc_mail = pick_col(vc_cols, "mail_id")
        vc_phone = pick_col(vc_cols, "contact_no")
        vc_bank_name = pick_col(vc_cols, "bank_name")
        vc_branch = pick_col(vc_cols, "branch_name")
        vc_account = pick_col(vc_cols, "account_no")
        vc_ifsc = pick_col(vc_cols, "ifsc_code")
        vc_holder = pick_col(vc_cols, "account_holder_name", "acc_holder_name")
        vc_pan_copy = pick_col(vc_cols, "pan_copy", "pan_attach_file_name")
        vc_bank_proof = pick_col(vc_cols, "bank_proof")

        vb_bill_date = pick_col(vpd_cols, "bill_date", "vendor_bill_created_date", "created")
        vb_invoice_no = pick_col(vpd_cols, "veninvverifyid", "invoice_no")
        vb_invoice_date = pick_col(vpd_cols, "vendor_inv_attach_approval_date", "invoice_date")
        vb_invoice_file = pick_col(vpd_cols, "vendor_inv_attach", "inv_verfiy_attach")
        vb_po_file = pick_col(vpd_cols, "po_ven_filename", "po_file", "vendor_po_attach")
        vb_remark = pick_col(vpd_cols, "bill_remark", "vendor_bill_reject_reason")
        vb_status = pick_col(vpd_cols, "vendor_bill_app_status")
        vb_approved_date = pick_col(vpd_cols, "vendor_bill_approval_date")
        vb_approved_by = pick_col(vpd_cols, "bill_approved", "vendor_bill_approval")
        vb_created_by = pick_col(vpd_cols, "bill_created_by", "vendor_bill_created_by")
        vb_created_date = pick_col(vpd_cols, "bill_created_date", "vendor_bill_created_date")
        vb_reject_reason = pick_col(vpd_cols, "reject_reason", "vendor_bill_reject_reason")
        additional_charges_col = pick_col(vpd_cols, "additionalcharges", "additional_charges")

        join_vendor = ""
        if vc_cols and vendor_id_col and "unique_id" in vc_cols:
            join_vendor = f"LEFT JOIN vendor_creation v ON v.unique_id = vb.{vendor_id_col}"

        bill_sql = f"""
            SELECT
                {f'COALESCE(v.{vc_vendor_name}, vb.vendor_name, "")' if (vc_vendor_name and join_vendor and "vendor_name" in vpd_cols) else f'v.{vc_vendor_name}' if (vc_vendor_name and join_vendor) else "COALESCE(vb.vendor_name, '')" if "vendor_name" in vpd_cols else "''"} AS vendor_name,
                {f'COALESCE(v.{vc_contact_person}, vb.vendor_name, "")' if (vc_contact_person and join_vendor and "vendor_name" in vpd_cols) else f'v.{vc_contact_person}' if (vc_contact_person and join_vendor) else "COALESCE(vb.vendor_name, '')" if "vendor_name" in vpd_cols else "''"} AS contact_person,
                {f'v.{vc_address}' if (vc_address and join_vendor) else "''"} AS address,
                {f'v.{vc_gst}' if (vc_gst and join_vendor) else "''"} AS gst_no,
                {f'v.{vc_pan}' if (vc_pan and join_vendor) else "''"} AS pan_no,
                {f'v.{vc_mail}' if (vc_mail and join_vendor) else "''"} AS mail_id,
                {f'v.{vc_phone}' if (vc_phone and join_vendor) else "''"} AS contact_no,
                vb.bill_no,
                {date_expr(f'vb.{vb_bill_date}', 'bill_date', '%d-%m-%Y') if vb_bill_date else "'' AS bill_date"},
                {f'vb.{vb_invoice_no}' if vb_invoice_no else "''"} AS invoice_no,
                {date_expr(f'vb.{vb_invoice_date}', 'invoice_date', '%d-%m-%Y') if vb_invoice_date else "'' AS invoice_date"},
                {f'vb.{vb_invoice_file}' if vb_invoice_file else "''"} AS invoice_file,
                {f'vb.{vb_po_file}' if vb_po_file else "''"} AS po_file,
                {f'vb.{vb_remark}' if vb_remark else "''"} AS bill_remark,
                {f'vb.{vb_status}' if vb_status else "0"} AS vendor_bill_app_status,
                {date_expr(f'vb.{vb_approved_date}', 'approved_date', '%d-%m-%Y') if vb_approved_date else "'' AS approved_date"},
                {staff_name_expr(f'vb.{vb_approved_by}', 'bill_approved') if vb_approved_by else "'' AS bill_approved"},
                {staff_name_expr(f'vb.{vb_created_by}', 'bill_created_by') if vb_created_by else "'' AS bill_created_by"},
                {date_expr(f'vb.{vb_created_date}', 'bill_created_date', '%d-%m-%Y %H:%i:%s') if vb_created_date else "'' AS bill_created_date"},
                {f'v.{vc_bank_name}' if (vc_bank_name and join_vendor) else "''"} AS bank_name,
                {f'v.{vc_branch}' if (vc_branch and join_vendor) else "''"} AS branch_name,
                {f'v.{vc_account}' if (vc_account and join_vendor) else "''"} AS account_no,
                {f'v.{vc_ifsc}' if (vc_ifsc and join_vendor) else "''"} AS ifsc_code,
                {f'v.{vc_holder}' if (vc_holder and join_vendor) else "''"} AS account_holder_name,
                {f'v.{vc_pan_copy}' if (vc_pan_copy and join_vendor) else "''"} AS pan_copy,
                {f'v.{vc_bank_proof}' if (vc_bank_proof and join_vendor) else "''"} AS bank_proof,
                {f'vb.{vb_reject_reason}' if vb_reject_reason else "''"} AS reject_reason,
                {f'IFNULL(vb.{additional_charges_col}, 0)' if additional_charges_col else '0'} AS additional_charges
            FROM vendor_payment_details vb
            {join_vendor}
            WHERE vb.bill_no = %s {soft_delete_filter('vb', vpd_cols)}
            LIMIT 1
        """

        bill_row = db_fetchone(bill_sql, [bill_no])
        if not bill_row:
            return Response({"error": "Bill not found"}, status=status.HTTP_404_NOT_FOUND)

        # DC line items
        dc_no_col = pick_col(vpd_cols, "dc_num")
        dc_date_col = pick_col(vpd_cols, "dc_date")
        inv_no_col = pick_col(vpd_cols, "invoice_no")
        inv_date_col = pick_col(vpd_cols, "invoice_date", "vendor_inv_attach_approval_date")
        po_no_col = pick_col(vpd_cols, "po_num")
        po_date_col = pick_col(vpd_cols, "po_date")
        inv_qty_col = pick_col(vpd_cols, "invoice_qty")
        consignee_address_col = pick_col(vpd_cols, "consignee_address")
        consignee_alt_col = pick_col(vpd_cols, "con_address")
        po_form_unique_id_col = pick_col(vpd_cols, "po_form_unique_id", "form_main_unique_id")
        row_vendor_id_col = pick_col(vpd_cols, "vendor_id")
        unit_price_col = pick_col(vpd_cols, "unit_price", "rate")
        basic_amount_col = pick_col(vpd_cols, "basic_amount", "amount")
        gst_col = pick_col(vpd_cols, "gst")
        gst_amount_col = pick_col(vpd_cols, "gst_amount")
        total_amount_col = pick_col(vpd_cols, "amount", "total_amount")
        order_col = pick_col(vpd_cols, "id", "unique_id")

        consignee_fallback_filters = []
        if po_form_unique_id_col:
            consignee_fallback_filters.append(f"vv.form_main_unique_id = vpd.{po_form_unique_id_col}")
        if dc_no_col:
            if row_vendor_id_col:
                consignee_fallback_filters.append(
                    f"(vv.dc_number = vpd.{dc_no_col} AND vv.engineer_name = vpd.{row_vendor_id_col})"
                )
            else:
                consignee_fallback_filters.append(f"vv.dc_number = vpd.{dc_no_col}")

        consignee_fallback_expr = "''"
        if consignee_fallback_filters:
            consignee_fallback_expr = (
                "(" 
                "SELECT COALESCE(get_address(vv.consignee_unique_id), '') "
                "FROM view_outsource_vendor_verified_invoice vv "
                f"WHERE {' OR '.join(consignee_fallback_filters)} "
                "ORDER BY vv.dc_date DESC, vv.unique_id DESC "
                "LIMIT 1"
                ")"
            )

        consignee_expr_parts = []
        if consignee_address_col:
            consignee_expr_parts.append(f"NULLIF(vpd.{consignee_address_col}, '')")
        if consignee_alt_col and consignee_alt_col != consignee_address_col:
            consignee_expr_parts.append(f"NULLIF(vpd.{consignee_alt_col}, '')")
        if consignee_fallback_filters:
            consignee_expr_parts.append(consignee_fallback_expr)
        consignee_expr_parts.append("''")
        consignee_expr = f"COALESCE({', '.join(consignee_expr_parts)})"

        dc_sql = f"""
            SELECT
                {f'vpd.{dc_no_col}' if dc_no_col else "''"} AS dc_no,
                {f'vpd.{dc_date_col}' if dc_date_col else "''"} AS dc_date,
                {f'vpd.{inv_no_col}' if inv_no_col else "''"} AS inv_no,
                {f'vpd.{inv_date_col}' if inv_date_col else "''"} AS inv_date,
                {f'vpd.{po_no_col}' if po_no_col else "''"} AS po_no,
                {f'vpd.{po_date_col}' if po_date_col else "''"} AS po_date,
                {f'vpd.{inv_qty_col}' if inv_qty_col else "0"} AS inv_qty,
                {consignee_expr} AS consignee,
                {f'vpd.{unit_price_col}' if unit_price_col else "0"} AS unit_price,
                {f'vpd.{basic_amount_col}' if basic_amount_col else "0"} AS basic_amount,
                {f'vpd.{gst_col}' if gst_col else "0"} AS gst,
                {f'vpd.{gst_amount_col}' if gst_amount_col else "0"} AS gst_amount,
                {f'vpd.{total_amount_col}' if total_amount_col else "0"} AS total_amount
            FROM vendor_payment_details vpd
            WHERE vpd.bill_no = %s {soft_delete_filter('vpd', vpd_cols)}
            ORDER BY {f'vpd.{order_col}' if order_col else 'vpd.bill_no'}
        """

        dc_rows = db_fetch(dc_sql, [bill_no])
        dc_list = []
        for i, row in enumerate(dc_rows, 1):
            dc_list.append(
                {
                    "sno": i,
                    "dc_no": safe_str(row.get("dc_no")),
                    "dc_date": safe_str(row.get("dc_date")),
                    "inv_no": safe_str(row.get("inv_no")),
                    "inv_date": safe_str(row.get("inv_date")),
                    "po_no": safe_str(row.get("po_no")),
                    "po_date": safe_str(row.get("po_date")),
                    "inv_qty": safe_float(row.get("inv_qty")),
                    "consignee": safe_str(row.get("consignee")),
                    "unit_price": safe_float(row.get("unit_price")),
                    "basic_amount": safe_float(row.get("basic_amount")),
                    "gst": safe_str(row.get("gst")),
                    "gst_amount": safe_float(row.get("gst_amount")),
                    "total_amount": safe_float(row.get("total_amount")),
                }
            )

        total_amount = sum(item["total_amount"] for item in dc_list)
        additional_charges = safe_float(bill_row.get("additional_charges"))
        grand_total_amount = total_amount + additional_charges

        # Approval timeline
        app_bill_created_by = pick_col(vpd_cols, "bill_created_by", "vendor_bill_created_by")
        app_bill_created_at = pick_col(vpd_cols, "bill_created_date", "vendor_bill_created_date")
        app_op_by = pick_col(vpd_cols, "bill_approved", "vendor_bill_approval")
        app_op_at = pick_col(vpd_cols, "vendor_bill_approval_date")
        app_op_sts = pick_col(vpd_cols, "vendor_bill_app_status")
        app_acc_by = pick_col(vpd_cols, "vendor_account_approved_by", "acc_approve")
        app_acc_at = pick_col(vpd_cols, "vendor_account_approval_date")
        app_acc_sts = pick_col(vpd_cols, "acc_ent_sts")
        app_fin_by = pick_col(vpd_cols, "finance_approved_by", "finance_approve")
        app_fin_at = pick_col(vpd_cols, "finance_approved_date")
        app_fin_sts = pick_col(vpd_cols, "finance_approval_sts", "finance_approval")
        app_mgmt_by = pick_col(vpd_cols, "managment_team_approvedby", "managed_approve")
        app_mgmt_at = pick_col(vpd_cols, "managment_team_approvaldate")
        app_mgmt_sts = pick_col(vpd_cols, "managment_team_approval_sts")
        app_pay_ref = pick_col(vpd_cols, "transaction_id", "transfer_ref")
        app_pay_date = pick_col(vpd_cols, "transaction_date", "transfer_date")
        app_pay_amount = pick_col(vpd_cols, "acctotalpaybleamount", "transfer_amount")
        app_pay_status = pick_col(vpd_cols, "accounts_approval", "payment_status")

        approval_sql = f"""
            SELECT
                {staff_name_max_expr(app_bill_created_by, 'bill_created_by') if app_bill_created_by else "'' AS bill_created_by"},
                {max_date_expr(app_bill_created_at, 'bill_created_at', '%d-%m-%Y %H:%i:%s') if app_bill_created_at else "'' AS bill_created_at"},
                {staff_name_max_expr(app_op_by, 'op_by') if app_op_by else "'' AS op_by"},
                {max_date_expr(app_op_at, 'op_at', '%d-%m-%Y %H:%i:%s') if app_op_at else "'' AS op_at"},
                {max_expr(app_op_sts, 'op_sts') if app_op_sts else "0 AS op_sts"},
                {staff_name_max_expr(app_acc_by, 'acc_entry_by') if app_acc_by else "'' AS acc_entry_by"},
                {max_date_expr(app_acc_at, 'acc_entry_at', '%d-%m-%Y %H:%i:%s') if app_acc_at else "'' AS acc_entry_at"},
                {max_expr(app_acc_sts, 'acc_ent_sts') if app_acc_sts else "0 AS acc_ent_sts"},
                {staff_name_max_expr(app_fin_by, 'finance_by') if app_fin_by else "'' AS finance_by"},
                {max_date_expr(app_fin_at, 'finance_at', '%d-%m-%Y %H:%i:%s') if app_fin_at else "'' AS finance_at"},
                {max_expr(app_fin_sts, 'finance_approval_sts') if app_fin_sts else "0 AS finance_approval_sts"},
                {staff_name_max_expr(app_mgmt_by, 'mgmt_by') if app_mgmt_by else "'' AS mgmt_by"},
                {max_date_expr(app_mgmt_at, 'mgmt_at', '%d-%m-%Y %H:%i:%s') if app_mgmt_at else "'' AS mgmt_at"},
                {max_expr(app_mgmt_sts, 'managment_team_approval_sts') if app_mgmt_sts else "0 AS managment_team_approval_sts"},
                {max_expr(app_pay_ref, 'pay_ref', cast_text=True)},
                {max_date_expr(app_pay_date, 'pay_date', '%d-%m-%Y') if app_pay_date else "'' AS pay_date"},
                {max_expr(app_pay_amount, 'pay_amount') if app_pay_amount else "0 AS pay_amount"},
                {max_expr(app_pay_status, 'payment_status') if app_pay_status else "0 AS payment_status"}
            FROM vendor_payment_details
            WHERE bill_no = %s {soft_delete_filter('vendor_payment_details', vpd_cols)}
        """

        approval = db_fetchone(approval_sql, [bill_no])
        approval_data = None
        if approval:
            approval_data = {
                "bill_created_by": safe_str(approval.get("bill_created_by")),
                "bill_created_at": safe_str(approval.get("bill_created_at")),
                "bill_created_status": "Bill Created",
                "op_by": safe_str(approval.get("op_by")),
                "op_at": safe_str(approval.get("op_at")),
                "op_status": status_text(approval.get("op_sts")),
                "acc_entry_by": safe_str(approval.get("acc_entry_by")),
                "acc_entry_at": safe_str(approval.get("acc_entry_at")),
                "acc_entry_status": status_text(approval.get("acc_ent_sts")),
                "finance_by": safe_str(approval.get("finance_by")),
                "finance_at": safe_str(approval.get("finance_at")),
                "finance_status": status_text(approval.get("finance_approval_sts")),
                "mgmt_by": safe_str(approval.get("mgmt_by")),
                "mgmt_at": safe_str(approval.get("mgmt_at")),
                "mgmt_status": status_text(approval.get("managment_team_approval_sts")),
                "pay_ref": safe_str(approval.get("pay_ref")),
                "pay_date": safe_str(approval.get("pay_date")),
                "pay_amount": safe_str(approval.get("pay_amount")),
                "pay_status": "Paid" if str(approval.get("payment_status")) == "1" else "Pending",
            }

        return Response(
            {
                "vendor": {
                    "name": safe_str(bill_row.get("vendor_name")),
                    "contact_person": safe_str(bill_row.get("contact_person")),
                    "address": safe_str(bill_row.get("address")),
                    "gst_no": safe_str(bill_row.get("gst_no")),
                    "pan_no": safe_str(bill_row.get("pan_no")),
                    "mail": safe_str(bill_row.get("mail_id")),
                    "phone": safe_str(bill_row.get("contact_no")),
                    "approved_by": safe_str(bill_row.get("bill_approved")),
                },
                "bill": {
                    "bill_no": safe_str(bill_row.get("bill_no")),
                    "bill_date": safe_str(bill_row.get("bill_date")),
                    "invoice_no": safe_str(bill_row.get("invoice_no")),
                    "invoice_date": safe_str(bill_row.get("invoice_date")),
                    "invoice_file": vendor_bill_file_url(request, "invoice", bill_row.get("invoice_file")),
                    "invoice_file_name": safe_str(bill_row.get("invoice_file")),
                    "po_file": vendor_bill_file_url(request, "po", bill_row.get("po_file")),
                    "po_file_name": safe_str(bill_row.get("po_file")),
                    "bill_created_by": safe_str(bill_row.get("bill_created_by")),
                    "bill_created_date": safe_str(bill_row.get("bill_created_date")),
                    "bill_remark": safe_str(bill_row.get("bill_remark")),
                    "approved_date": safe_str(bill_row.get("approved_date")),
                },
                "bank": {
                    "bank_name": safe_str(bill_row.get("bank_name")),
                    "branch": safe_str(bill_row.get("branch_name")),
                    "account_no": safe_str(bill_row.get("account_no")),
                    "ifsc_code": safe_str(bill_row.get("ifsc_code")),
                    "holder": safe_str(bill_row.get("account_holder_name")),
                    "pan_copy": vendor_bill_file_url(request, "pan", bill_row.get("pan_copy")),
                    "pan_copy_name": safe_str(bill_row.get("pan_copy")),
                    "bank_proof": vendor_bill_file_url(request, "bank", bill_row.get("bank_proof")),
                    "bank_proof_name": safe_str(bill_row.get("bank_proof")),
                    "reject_reason": safe_str(bill_row.get("reject_reason")),
                },
                "dc_items": dc_list,
                "total_amount": total_amount,
                "additional_charges": additional_charges,
                "grand_total_amount": grand_total_amount,
                "approval": approval_data,
            }
        )


class VendorBillApprovalFileView(APIView):
    def get(self, request, file_kind, filename):
        if file_kind not in {"invoice", "po", "pan", "bank"}:
            raise Http404("Invalid file type")

        safe_name = Path(safe_str(filename)).name
        if not safe_name:
            raise Http404("File not found")

        for candidate in vendor_bill_file_candidates(file_kind, safe_name):
            try:
                if candidate.exists() and candidate.is_file():
                    return FileResponse(open(candidate, "rb"), as_attachment=False, filename=safe_name)
            except Exception:
                continue

        raise Http404("File not found")


class VendorBillUpdateRemarkView(APIView):
    # POST /vendor-bill-approval/update-remark/
    def post(self, request):
        bill_no = request.data.get("bill_no", "")
        remark = request.data.get("remark", "")

        if not bill_no:
            return Response({"error": "bill_no required"}, status=status.HTTP_400_BAD_REQUEST)

        vpd_cols = table_columns("vendor_payment_details")
        vpm_cols = table_columns("vendor_payment_details_main")

        remark_col = pick_col(vpd_cols, "bill_remark", "vendor_bill_reject_reason")
        if not remark_col:
            return Response(
                {"status": 0, "msg": "Remark field not available"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        where_clause = "bill_no = %s"
        if "is_delete" in vpd_cols:
            where_clause += " AND COALESCE(is_delete, 0) = 0"

        with connection.cursor() as cur:
            cur.execute(
                f"UPDATE vendor_payment_details SET {remark_col} = %s WHERE {where_clause}",
                [remark, bill_no],
            )

            main_remark_col = pick_col(vpm_cols, remark_col, "bill_remark", "vendor_bill_reject_reason")
            if main_remark_col:
                main_where = "bill_no = %s"
                if "is_delete" in vpm_cols:
                    main_where += " AND COALESCE(is_delete, 0) = 0"
                cur.execute(
                    f"UPDATE vendor_payment_details_main SET {main_remark_col} = %s WHERE {main_where}",
                    [remark, bill_no],
                )

        return Response({"status": 1, "msg": "Remark updated"})


class VendorBillApproveView(APIView):
    # POST /vendor-bill-approval/approve/
    def post(self, request):
        bill_no = safe_str(request.data.get("bill_no", "")).strip()
        approval_status = safe_str(request.data.get("status", "Approved"))
        remark = safe_str(request.data.get("remark", "")).strip()
        reject_reason = safe_str(request.data.get("reject_reason", remark)).strip()
        user_id = get_user_id(request)

        if not bill_no:
            return Response({"error": "bill_no required"}, status=status.HTTP_400_BAD_REQUEST)

        status_map = {"Approved": "1", "Rejected": "2", "Pending": "0"}
        status_value = status_map.get(approval_status, "0")

        if status_value == "2" and not reject_reason:
            return Response(
                {"status": 0, "msg": "Reject reason required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        vpd_cols = table_columns("vendor_payment_details")
        vpm_cols = table_columns("vendor_payment_details_main")

        status_col = pick_col(vpd_cols, "vendor_bill_app_status")
        if not status_col:
            return Response(
                {"status": 0, "msg": "Status field not available"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        bill_remark_col = pick_col(vpd_cols, "bill_remark")
        reject_reason_col = pick_col(vpd_cols, "vendor_bill_reject_reason", "reject_reason")
        date_col = pick_col(vpd_cols, "vendor_bill_approval_date")
        approve_by_col = pick_col(vpd_cols, "vendor_bill_approval", "bill_approved")
        rejected_by_col = pick_col(vpd_cols, "vendor_bill_rejected_by")

        set_parts = [f"{status_col} = %s"]
        values = [status_value]

        if bill_remark_col:
            set_parts.append(f"{bill_remark_col} = %s")
            values.append(remark)

        if reject_reason_col:
            set_parts.append(f"{reject_reason_col} = %s")
            values.append(reject_reason if status_value == "2" else "")

        if approve_by_col:
            set_parts.append(f"{approve_by_col} = %s")
            values.append(user_id if status_value == "1" else "")

        if rejected_by_col:
            set_parts.append(f"{rejected_by_col} = %s")
            values.append(user_id if status_value == "2" else "")

        if date_col:
            set_parts.append(f"{date_col} = {'NOW()' if status_value == '1' else 'NULL'}")

        where_clause = "bill_no = %s"
        if "is_delete" in vpd_cols:
            where_clause += " AND COALESCE(is_delete, 0) = 0"

        with connection.cursor() as cur:
            cur.execute(
                f"UPDATE vendor_payment_details SET {', '.join(set_parts)} WHERE {where_clause}",
                values + [bill_no],
            )

            main_status_col = pick_col(vpm_cols, status_col)
            main_bill_remark_col = pick_col(vpm_cols, bill_remark_col, "bill_remark")
            main_reject_reason_col = pick_col(vpm_cols, reject_reason_col, "vendor_bill_reject_reason", "reject_reason")
            main_date_col = pick_col(vpm_cols, date_col)
            main_approve_by_col = pick_col(vpm_cols, approve_by_col, "vendor_bill_approval", "bill_approved")
            main_rejected_by_col = pick_col(vpm_cols, rejected_by_col, "vendor_bill_rejected_by")

            if main_status_col:
                main_parts = [f"{main_status_col} = %s"]
                main_values = [status_value]
                if main_bill_remark_col:
                    main_parts.append(f"{main_bill_remark_col} = %s")
                    main_values.append(remark)
                if main_reject_reason_col:
                    main_parts.append(f"{main_reject_reason_col} = %s")
                    main_values.append(reject_reason if status_value == "2" else "")
                if main_approve_by_col:
                    main_parts.append(f"{main_approve_by_col} = %s")
                    main_values.append(user_id if status_value == "1" else "")
                if main_rejected_by_col:
                    main_parts.append(f"{main_rejected_by_col} = %s")
                    main_values.append(user_id if status_value == "2" else "")
                if main_date_col:
                    main_parts.append(f"{main_date_col} = {'NOW()' if status_value == '1' else 'NULL'}")

                main_where = "bill_no = %s"
                if "is_delete" in vpm_cols:
                    main_where += " AND COALESCE(is_delete, 0) = 0"

                cur.execute(
                    f"UPDATE vendor_payment_details_main SET {', '.join(main_parts)} WHERE {main_where}",
                    main_values + [bill_no],
                )

            # When rejected, reset invoices back to Stage 1 (Vendor Bill Creation)
            if status_value == "2":
                try:
                    cur.execute(
                        """
                        UPDATE invoice_verfication_table iv
                        JOIN vendor_payment_details vpd ON vpd.dc_num = iv.dc_number
                        SET iv.vendor_payment_allocated = 0,
                            iv.vendor_bill_no = ''
                        WHERE vpd.bill_no = %s AND COALESCE(vpd.is_delete, 0) = 0
                        """,
                        [bill_no],
                    )
                except Exception:
                    pass

        return Response({"status": 1, "msg": "Bill " + approval_status.lower()})


class VendorBillExportView(APIView):
    # GET /vendor-bill-approval/export/
    def get(self, request):
        vpd_cols = table_columns("vendor_payment_details")
        if not vpd_cols or "bill_no" not in vpd_cols:
            rows = []
        else:
            bill_date_col = pick_col(vpd_cols, "bill_date", "vendor_bill_created_date", "created")
            vendor_name_col = pick_col(vpd_cols, "vendor_name")
            dc_col = pick_col(vpd_cols, "dc_num")
            amount_col = pick_col(vpd_cols, "amount", "total_amount")
            additional_charges_col = pick_col(vpd_cols, "additionalcharges", "additional_charges")
            status_col = pick_col(vpd_cols, "vendor_bill_app_status")
            order_col = pick_col(vpd_cols, "id", "unique_id", "bill_date")

            where_parts = ["1=1"]
            if "is_delete" in vpd_cols:
                where_parts.append("COALESCE(is_delete, 0) = 0")
            where_sql = " AND ".join(where_parts)

            grand_total_expr = "0"
            if amount_col and additional_charges_col:
                grand_total_expr = f"ROUND(IFNULL(SUM({amount_col}), 0) + MAX(IFNULL({additional_charges_col}, 0)), 2)"
            elif amount_col:
                grand_total_expr = f"ROUND(IFNULL(SUM({amount_col}), 0), 2)"
            elif additional_charges_col:
                grand_total_expr = f"ROUND(MAX(IFNULL({additional_charges_col}, 0)), 2)"

            sql = f"""
                SELECT
                    bill_no,
                    {max_date_expr(bill_date_col, 'bill_date', '%d-%m-%Y')},
                    {max_expr(vendor_name_col, 'vendor_name', cast_text=True)},
                    {f"COUNT(DISTINCT {dc_col}) AS dc_count" if dc_col else "0 AS dc_count"},
                    {grand_total_expr} AS total_amount,
                    {max_expr(status_col, 'vendor_bill_app_status') if status_col else "0 AS vendor_bill_app_status"}
                FROM vendor_payment_details
                WHERE {where_sql}
                GROUP BY bill_no
                ORDER BY {f'MAX({order_col}) DESC' if order_col else 'bill_no DESC'}
            """
            rows = db_fetch(sql)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vendor Bill Approval"

        ws.merge_cells("A1:G1")
        ws["A1"] = "Vendor Bill Approval"
        ws["A1"].font = Font(bold=True, size=13)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["S.No", "Bill No", "Bill Date", "Vendor Name", "DC Count", "Amount", "Status"]
        for ci, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=ci, value=header)
            cell.font = Font(bold=True)

        for i, row in enumerate(rows, 1):
            sts = "Pending" if str(row.get("vendor_bill_app_status", "0")) == "0" else "Complete"
            ws.append(
                [
                    i,
                    safe_str(row.get("bill_no")),
                    safe_str(row.get("bill_date")),
                    safe_str(row.get("vendor_name")),
                    int(row.get("dc_count") or 0),
                    safe_float(row.get("total_amount")),
                    sts,
                ]
            )

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(len(str(c.value or "")) for c in col) + 4,
                50,
            )

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="VendorBillApproval.xlsx"'
        return response

