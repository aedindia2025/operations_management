import io
from datetime import datetime

import openpyxl
from django.db import connection
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView


def db_fetch(sql, params=None):
    with connection.cursor() as cursor:
        cursor.execute(sql, params or [])
        columns = [col[0] for col in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]


def fetch_completed_po_rows(from_date, to_date):
    return db_fetch(
        """
        SELECT DISTINCT
            vpl.po_num,
            DATE_FORMAT(vpl.po_date, '%%d-%%m-%%Y') AS po_date,
            get_department_name(vpl.department) AS department_name,
            COALESCE(
                NULLIF(get_ledger_name(icm_latest.ledger_name), ''),
                NULLIF(icm_latest.ledger_name, ''),
                NULLIF(dc.ledger_name, ''),
                ''
            ) AS ledger_name,
            COALESCE(NULLIF(icm_latest.ledger_no, ''), NULLIF(dc.ledger_no, ''), '') AS ledger_id,
            get_executive_name(vpl.executive_name) AS executive_name,
            vpl.billing_gst_no,
            vpl.region,
            vpl.zone_code,
            vpl.zone,
            vpl.con_branch_code,
            vpl.con_branch,
            vpl.cons_billing_address,
            vpl.consignee_gst,
            vpl.con_name,
            vpl.con_address,
            get_district_name(vpl.con_district) AS district,
            get_state_name(vpl.con_state) AS state,
            vpl.con_pincode,
            vpl.contact_name,
            vpl.alter_contact_name,
            vpl.con_contact_no,
            vpl.alter_number,
            vpl.cons_email_id,
            vpl.item_code,
            vpl.product,
            vpl.assign_qty,
            vpl.invoice_qty,
            vpl.qty,
            vpl.item_warranty,
            vpl.warranty_duration,
            vpl.invoice_no,
            DATE_FORMAT(vpl.invoice_date, '%%d-%%m-%%Y') AS invoice_date,
            vpl.dc_num,
            DATE_FORMAT(vpl.dc_date, '%%d-%%m-%%Y') AS dc_date,
            vpl.serial_no,
            vpl.courier_name,
            vpl.pod_no,
            DATE_FORMAT(vpl.dispatch_date, '%%d-%%m-%%Y') AS dispatch_date,
            vpl.delivery_status,
            DATE_FORMAT(vpl.delivery_date, '%%d-%%m-%%Y') AS delivery_date,
            vpl.dc_received_sts AS dc_status,
            DATE_FORMAT(vpl.dc_cus_signed_date, '%%d-%%m-%%Y') AS dc_sign_date,
            vpl.ir_rec_status AS ir_status,
            DATE_FORMAT(vpl.ir_cus_signed_date, '%%d-%%m-%%Y') AS ir_date,
            vpl.snr_rec_status AS snr_status,
            vpl.snr_cus_signed_date AS snr_date,
            CASE
                WHEN vpl.eng_type = 'own-engineer' AND vpl.eng_name = '689dd2899d78831162' THEN 'Service Team'
                WHEN vpl.eng_type = 'own-engineer' AND vpl.eng_name = '689dd241c69a654753' THEN 'Inhouse Team'
                WHEN vpl.eng_type = 'own-engineer' AND vpl.eng_name NOT IN ('689dd2899d78831162', '689dd241c69a654753') THEN get_service_engineer_name(vpl.eng_name)
                ELSE vpl.eng_name
            END AS eng_name,
            vpl.vendor_bulk_rate,
            vpl.vendor_bulk_gst,
            vpl.bulk_total_amount,
            vpl.cons_followed_by
        FROM view_po_wise_list vpl
        LEFT JOIN department_creation dc
          ON dc.unique_id = vpl.department
         AND COALESCE(dc.is_delete, 0) = 0
        LEFT JOIN (
            SELECT
                po_num,
                invoice_no,
                dc_number,
                SUBSTRING_INDEX(GROUP_CONCAT(COALESCE(ledger_name, '') ORDER BY id DESC SEPARATOR '\t'), '\t', 1) AS ledger_name,
                SUBSTRING_INDEX(GROUP_CONCAT(COALESCE(ledger_no, '') ORDER BY id DESC SEPARATOR '\t'), '\t', 1) AS ledger_no
            FROM invoice_creation_main
            WHERE COALESCE(is_delete, 0) = 0
            GROUP BY po_num, invoice_no, dc_number
        ) icm_latest
          ON icm_latest.po_num = vpl.po_num
         AND icm_latest.invoice_no = vpl.invoice_no
         AND icm_latest.dc_number = vpl.dc_num
        WHERE DATE(vpl.po_date) BETWEEN %s AND %s
          AND COALESCE(vpl.delivery_status, 0) = 1
        ORDER BY vpl.po_date, vpl.po_num, vpl.con_branch, vpl.item_code
        """,
        [from_date, to_date],
    )


class CompletedPoReportDownloadView(APIView):
    def get(self, request):
        from_date = str(request.query_params.get("from_date", "") or "").strip()
        to_date = str(request.query_params.get("to_date", "") or "").strip()

        if not from_date or not to_date:
            return Response(
                {"error": "from_date and to_date are required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            parsed_from = datetime.strptime(from_date, "%Y-%m-%d").date()
            parsed_to = datetime.strptime(to_date, "%Y-%m-%d").date()
        except ValueError:
            return Response(
                {"error": "Invalid date format. Use YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if parsed_from > parsed_to:
            return Response(
                {"error": "From date cannot be greater than To date."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        rows = fetch_completed_po_rows(from_date, to_date)
        if not rows:
            return Response({"error": "No completed PO data found"}, status=status.HTTP_404_NOT_FOUND)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Completed PO"

        headers = [
            "S.No",
            "PO No",
            "PO Date",
            "Department",
            "Billing Address",
            "Billing GST No",
            "Branch Name",
            "Branch Code",
            "Consignee Address",
            "District",
            "Region",
            "Zone Code",
            "Zone",
            "Pincode",
            "State",
            "Consignee GST No",
            "Contact Person Name",
            "Contact Number",
            "Alternate Contact Name",
            "Alternate Contact Number",
            "Consignee Email ID",
            "Ledger Name",
            "Ledger ID",
            "Executive",
            "QTY",
            "Item Code",
            "Product",
            "Assigned Qty",
            "Invoice Qty",
            "PO Qty",
            "Warranty",
            "DC No",
            "DC Date",
            "Invoice No",
            "Invoice Date",
            "Serial Number",
            "Courier Name",
            "Dispatch Date",
            "POD Number",
            "Delivery Date",
            "DC Received Date",
            "Installation Date",
            "SNR Date",
            "Vendor",
            "Vendor Rate",
            "Vendor GST %",
            "Bulk Total Amount",
            "Followed By",
            "Purchase Order Status",
            "Stock Position Status",
            "Consignee Stock Assign Status",
            "Invoice Status",
            "Operation Approval Status",
            "Accounts Approval Status",
            "Material QC Status",
            "Dispatch Status",
            "Delivery Confirmation Status",
        ]

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws["A1"] = "Completed PO Report"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws["A2"] = (
            f"From Date: {parsed_from.strftime('%d-%m-%Y')}  |  "
            f"To Date: {parsed_to.strftime('%d-%m-%Y')}  |  "
            f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
        )
        ws["A2"].font = Font(italic=True)
        ws["A2"].alignment = Alignment(horizontal="center")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        completed_text = "Completed"

        for index, row in enumerate(rows, start=1):
            warranty = row.get("item_warranty") or row.get("warranty_duration") or ""
            data_row = [
                index,
                row.get("po_num", ""),
                row.get("po_date", ""),
                row.get("department_name", ""),
                row.get("cons_billing_address", ""),
                row.get("billing_gst_no", ""),
                row.get("con_branch", ""),
                row.get("con_branch_code", ""),
                row.get("con_address", ""),
                row.get("district", ""),
                row.get("region", ""),
                row.get("zone_code", ""),
                row.get("zone", ""),
                row.get("con_pincode", ""),
                row.get("state", ""),
                row.get("consignee_gst", ""),
                row.get("contact_name", ""),
                row.get("con_contact_no", ""),
                row.get("alter_contact_name", ""),
                row.get("alter_number", ""),
                row.get("cons_email_id", ""),
                row.get("ledger_name", ""),
                row.get("ledger_id", ""),
                row.get("executive_name", ""),
                row.get("qty", ""),
                row.get("item_code", ""),
                row.get("product", ""),
                row.get("assign_qty", ""),
                row.get("invoice_qty", ""),
                row.get("qty", ""),
                warranty,
                row.get("dc_num", ""),
                row.get("dc_date", ""),
                row.get("invoice_no", ""),
                row.get("invoice_date", ""),
                row.get("serial_no", ""),
                row.get("courier_name", ""),
                row.get("dispatch_date", ""),
                row.get("pod_no", ""),
                row.get("delivery_date", ""),
                row.get("dc_sign_date", ""),
                row.get("ir_date", ""),
                row.get("snr_date", ""),
                row.get("eng_name", ""),
                row.get("vendor_bulk_rate", ""),
                row.get("vendor_bulk_gst", ""),
                row.get("bulk_total_amount", ""),
                row.get("cons_followed_by", ""),
                completed_text,
                completed_text,
                completed_text,
                completed_text,
                completed_text,
                completed_text,
                completed_text,
                completed_text,
                completed_text,
            ]

            for col_idx, value in enumerate(data_row, start=1):
                ws.cell(row=index + 3, column=col_idx, value=value)

        for col_idx in range(1, ws.max_column + 1):
            max_len = 10
            for row_idx in range(1, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                max_len = max(max_len, len(str(value or "")))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 35)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Completed_PO_{from_date}_to_{to_date}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
