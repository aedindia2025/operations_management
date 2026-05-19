# master/viewsets/po_report/po_report_viewset.py

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from django.db import connection
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import io


def disdate(date_val):
    if not date_val:
        return ""
    try:
        if isinstance(date_val, str):
            parts = date_val.split("-")
            return "-".join(reversed(parts))
        return date_val.strftime("%d-%m-%Y")
    except Exception:
        return str(date_val)


def db_lookup(table, id_col, val_col, uid):
    if not uid:
        return ""
    with connection.cursor() as cursor:
        cursor.execute(f"SELECT {val_col} FROM {table} WHERE {id_col} = %s", [uid])
        row = cursor.fetchone()
        return row[0] if row else ""


def db_fetch(sql, params=None):
    with connection.cursor() as cursor:
        cursor.execute(sql, params or [])
        columns = [col[0] for col in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]


def resolve_department_name(customer_value):
    value = str(customer_value or "").strip()
    if not value:
        return ""
    rows = db_fetch(
        """
        SELECT department
        FROM department_creation
        WHERE is_delete = 0
          AND is_active = 1
          AND (unique_id = %s OR department = %s)
        ORDER BY id DESC
        LIMIT 1
        """,
        [value, value],
    )
    return rows[0]["department"] if rows else value


def resolve_po_number(po_value):
    value = str(po_value or "").strip()
    if not value:
        return ""
    rows = db_fetch(
        """
        SELECT po_num
        FROM po_form
        WHERE is_delete = 0
          AND is_active = 1
          AND (unique_id = %s OR po_num = %s)
        ORDER BY id DESC
        LIMIT 1
        """,
        [value, value],
    )
    return rows[0]["po_num"] if rows else value


def resolve_po_lookup_values(po_value):
    value = str(po_value or "").strip()
    if not value:
        return []
    rows = db_fetch(
        """
        SELECT unique_id, po_unique_id, po_num
        FROM po_form
        WHERE is_delete = 0
          AND is_active = 1
          AND (unique_id = %s OR po_unique_id = %s OR po_num = %s)
        ORDER BY id DESC
        LIMIT 1
        """,
        [value, value, value],
    )
    values = [value]
    if rows:
        values.extend([rows[0].get("unique_id"), rows[0].get("po_unique_id"), rows[0].get("po_num")])
    return [str(item).strip() for item in dict.fromkeys(values) if str(item or "").strip()]


def resolve_customer_lookup_values(customer_value):
    value = str(customer_value or "").strip()
    if not value:
        return []
    rows = db_fetch(
        """
        SELECT unique_id, department
        FROM department_creation
        WHERE is_delete = 0
          AND is_active = 1
          AND (unique_id = %s OR department = %s)
        ORDER BY id DESC
        LIMIT 1
        """,
        [value, value],
    )
    values = [value]
    if rows:
        values.extend([rows[0].get("unique_id"), rows[0].get("department")])
    return [str(item).strip() for item in dict.fromkeys(values) if str(item or "").strip()]


def fetch_po_report_rows(customer_value="", po_value=""):
    where = []
    params = []
    if po_value:
        po_values = resolve_po_lookup_values(po_value)
        placeholders = ", ".join(["%s"] * len(po_values))
        where.append(f"(po_unique_id IN ({placeholders}) OR po_num IN ({placeholders}))")
        params.extend(po_values)
        params.extend(po_values)
    if customer_value:
        customer_values = resolve_customer_lookup_values(customer_value)
        placeholders = ", ".join(["%s"] * len(customer_values))
        where.append(f"(department IN ({placeholders}) OR get_department_name(department) IN ({placeholders}))")
        params.extend(customer_values)
        params.extend(customer_values)

    where_sql = " AND ".join(where) if where else "1=1"
    return db_fetch(
        f"""
        SELECT
            po_num,
            DATE_FORMAT(po_date, '%%d-%%m-%%Y') AS po_date,
            get_department_name(department) AS department_name,
            gst,
            con_name,
            DATE_FORMAT(consignee_batch_date, '%%d-%%m-%%Y') AS consignee_batch_date,
            con_branch,
            con_branch_code,
            cons_billing_address,
            con_address,
            get_district_name(con_district) AS district,
            get_state_name(con_state) AS state,
            con_pincode,
            zone,
            zone_code,
            con_contact_no,
            con_lan_num,
            consignee_gst,
            contact_name,
            department,
            alter_contact_name,
            alter_number,
            cons_email_id,
            item_code,
            dc_num,
            DATE_FORMAT(dc_date, '%%d-%%m-%%Y') AS dc_date,
            invoice_no,
            DATE_FORMAT(invoice_date, '%%d-%%m-%%Y') AS invoice_date,
            serial_no,
            courier_name,
            pod_no,
            DATE_FORMAT(dispatch_date, '%%d-%%m-%%Y') AS dispatch_date,
            delivery_status,
            DATE_FORMAT(delivery_date, '%%d-%%m-%%Y') AS delivery_date,
            dc_received_sts AS dc_status,
            DATE_FORMAT(dc_cus_signed_date, '%%d-%%m-%%Y') AS dc_sign_date,
            ir_rec_status AS ir_status,
            DATE_FORMAT(ir_cus_signed_date, '%%d-%%m-%%Y') AS ir_date,
            snr_rec_status AS snr_status,
            snr_cus_signed_date AS snr_date,
            assign_qty,
            invoice_qty,
            qty,
            product,
            get_executive_name(executive_name) AS executive_name,
            eng_type,
            CASE
                WHEN eng_type = 'own-engineer' AND eng_name = '689dd2899d78831162' THEN 'Service Team'
                WHEN eng_type = 'own-engineer' AND eng_name = '689dd241c69a654753' THEN 'Inhouse Team'
                WHEN eng_type = 'own-engineer' AND eng_name NOT IN ('689dd2899d78831162', '689dd241c69a654753') THEN get_service_engineer_name(eng_name)
                ELSE eng_name
            END AS eng_name,
            vendor_bulk_rate,
            vendor_bulk_gst,
            bulk_total_amount,
            cons_followed_by,
            warranty_duration,
            billing_gst_no,
            region,
            item_warranty
        FROM view_po_wise_list
        WHERE {where_sql}
        ORDER BY po_num, con_branch, item_code
        """,
        params,
    )


# ── /po-report/customers/ ─────────────────────────────────────────────────────
class PoReportCustomerOptionsView(APIView):
    def get(self, request):
        rows = db_fetch(
            """
            SELECT DISTINCT dc.unique_id, dc.department
            FROM department_creation dc
            INNER JOIN po_form pf
                ON pf.department = dc.unique_id
               AND pf.is_delete = 0
               AND pf.is_active = 1
            WHERE dc.is_delete = 0
              AND dc.is_active = 1
            ORDER BY dc.department
            """
        )
        return Response(
            [{"label": row["department"], "value": row["unique_id"]} for row in rows if row.get("department")]
        )


# ── /po-report/po-numbers/ ────────────────────────────────────────────────────
class PoReportPoNumberOptionsView(APIView):
    def get(self, request):
        customer = str(request.query_params.get("customer", "")).strip()
        sql = """
            SELECT DISTINCT unique_id, po_num
            FROM po_form
            WHERE is_delete = 0
              AND is_active = 1
        """
        params = []
        if customer:
            sql += " AND department = %s"
            params.append(customer)
        sql += " ORDER BY po_num"
        rows = db_fetch(sql, params)
        return Response(
            [{"label": row["po_num"], "value": row["unique_id"]} for row in rows if row.get("po_num")]
        )


# ── /po-report/download/ ──────────────────────────────────────────────────────
class PoReportDownloadView(APIView):
    def get(self, request):
        customer_value = str(request.query_params.get("customer", "")).strip()
        po_value = str(request.query_params.get("po_no", "")).strip()
        customer = resolve_department_name(customer_value)
        po_no = resolve_po_number(po_value)

        if not customer and not po_no:
            return Response(
                {"error": "customer or po_no is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        rows = fetch_po_report_rows(customer_value=customer_value, po_value=po_value)

        if not rows:
            return Response({"error": "No data found"}, status=status.HTTP_404_NOT_FOUND)

        # ── Build Excel ───────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PO Report"

        # Row 1 - Title
        ws.merge_cells("A1:BT1")
        ws["A1"] = "Elcot Tracker"
        ws["A1"].font      = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Row 2 - Info
        ws.merge_cells("A2:BT2")
        info_customer = customer or "All"
        info_po = po_no or "All"
        ws["A2"] = f"Customer: {info_customer}  |  PO No: {info_po}  |  Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
        ws["A2"].font      = Font(italic=True)
        ws["A2"].alignment = Alignment(horizontal="center")

        # Row 3 - Headers
        headers = [
            "S.No", "PO NUMBER", "PO DATE", "CONSIGNEE BATCH DATE", "DEPARTMENT NAME",
            "BILLING GST NUMBER", "REGION", "ZONE CODE", "ZONE NAME", "BRANCH CODE",
            "BRANCH NAME", "BILLING ADDRESS", "CONSIGNEE GST NO", "CONSIGNEE ADDRESS",
            "DISTRICT", "PINCODE", "STATE", "CONTACT PERSON NAME", "ALTER PERSON NAME",
            "CONTACT NUMBER", "ALTERNATE CONTACT NUMBER", "CONS EMAIL ID", "ITEM CODE",
            "ITEM DETAILS", "QTY", "WARRANTY", "INVOICE NO", "INVOICE DATE", "DC NO",
            "DC DATE", "SERIAL NUMBER", "COURIER NAME", "POD NUMBER", "DISPATCH DATE",
            "DELIVERY STATUS", "DELIVERY DATE", "DC STATUS", "DC RECEIVED DATE",
            "INSTALLED QTY", "SNR STATUS", "SNR QTY", "INSTALLATION STATUS",
            "INSTALLATION DATE", "VENDOR", "VENDOR PRICE", "GST %", "GST PER UNIT",
            "TOTAL PER UNIT", "TOTAL PRICE WITHOUT GST(QTY)", "TOTAL PRICE WITH GST(QTY)",
            "EXECUTIVE", "FOLLOWED BY"
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell      = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)

        # Data rows
        for sno, row in enumerate(rows, start=1):
            assign_qty = row.get("invoice_qty") or row.get("assign_qty") or ""
            try:
                rate = float(row.get("vendor_bulk_rate") or 0)
            except (TypeError, ValueError):
                rate = 0.0
            try:
                gst_percent = float(row.get("vendor_bulk_gst") or 0)
            except (TypeError, ValueError):
                gst_percent = 0.0

            gst_amount = rate * (gst_percent / 100) if gst_percent > 0 else 0.0
            total_per_unit = rate + gst_amount
            if gst_percent == 0:
                total_without_gst = row.get("bulk_total_amount") or ""
                total_with_gst = ""
            else:
                try:
                    total_without_gst = rate * float(assign_qty or 0)
                except (TypeError, ValueError):
                    total_without_gst = ""
                total_with_gst = row.get("bulk_total_amount") or ""

            installed_qty = row.get("invoice_qty") if str(row.get("ir_status")) == "1" else 0
            snr_qty = row.get("invoice_qty") if str(row.get("snr_status")) == "1" else ""

            data_row = [
                sno,
                row.get("po_num"),
                row.get("po_date"),
                row.get("consignee_batch_date"),
                row.get("department_name"),
                row.get("billing_gst_no"),
                row.get("region"),
                row.get("zone_code"),
                row.get("zone"),
                row.get("con_branch_code"),
                row.get("con_branch"),
                row.get("cons_billing_address"),
                row.get("consignee_gst"),
                row.get("con_address"),
                row.get("district"),
                row.get("con_pincode"),
                row.get("state"),
                row.get("contact_name"),
                row.get("alter_contact_name"),
                row.get("con_contact_no"),
                row.get("alter_number"),
                row.get("cons_email_id"),
                row.get("item_code"),
                row.get("product"),
                assign_qty,
                row.get("item_warranty") or row.get("warranty_duration"),
                row.get("invoice_no"),
                row.get("invoice_date"),
                row.get("dc_num"),
                row.get("dc_date"),
                row.get("serial_no"),
                row.get("courier_name"),
                row.get("pod_no"),
                row.get("dispatch_date"),
                "Yes" if str(row.get("delivery_status")) == "1" else "No",
                row.get("delivery_date"),
                "Yes" if str(row.get("dc_status")) == "1" else "No",
                row.get("dc_sign_date"),
                installed_qty,
                "Yes" if str(row.get("snr_status")) == "1" else "",
                snr_qty,
                "Yes" if str(row.get("ir_status")) == "1" else "No",
                row.get("ir_date"),
                row.get("eng_name"),
                row.get("vendor_bulk_rate"),
                row.get("vendor_bulk_gst"),
                gst_amount,
                total_per_unit,
                total_without_gst,
                total_with_gst,
                row.get("executive_name"),
                row.get("cons_followed_by"),
            ]

            for col_idx, val in enumerate(data_row, start=1):
                ws.cell(row=sno + 3, column=col_idx, value=val)

        # Auto-size columns
        for col_idx in range(1, ws.max_column + 1):
            max_len = 10
            for row_idx in range(1, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                max_len = max(max_len, len(str(value or "")))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        # Stream response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename_ref = po_no or customer or "report"
        filename = f"PO_Report_{filename_ref.replace('/', '_')}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
