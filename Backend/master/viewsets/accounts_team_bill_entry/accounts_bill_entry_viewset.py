import io

import openpyxl
from django.db import connection
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView


_COLUMN_CACHE = {}


def db_fetch(sql, params=None):
    with connection.cursor() as cursor:
        cursor.execute(sql, params or [])
        cols = [c[0] for c in cursor.description]
        return [dict(zip(cols, row)) for row in cursor.fetchall()]


def db_fetchone(sql, params=None):
    rows = db_fetch(sql, params)
    return rows[0] if rows else None


def table_columns(table_name):
    cached = _COLUMN_CACHE.get(table_name)
    if cached is not None:
        return cached

    try:
        with connection.cursor() as cursor:
            cursor.execute(f"SHOW COLUMNS FROM `{table_name}`")
            cols = {row[0] for row in cursor.fetchall()}
    except Exception:
        cols = set()

    _COLUMN_CACHE[table_name] = cols
    return cols


def ensure_decimal_column(table_name, *candidates):
    cols = table_columns(table_name)
    existing = pick_col(cols, *candidates)
    if existing:
        return existing

    target = candidates[0] if candidates else None
    if not target:
        return None

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                f"ALTER TABLE `{table_name}` ADD COLUMN `{target}` DECIMAL(15,2) NOT NULL DEFAULT 0"
            )
    except Exception:
        pass

    _COLUMN_CACHE.pop(table_name, None)
    return pick_col(table_columns(table_name), *candidates)


def pick_col(columns, *candidates):
    for col in candidates:
        if col and col in columns:
            return col
    return None


def safe_str(value):
    return "" if value is None else str(value)


def safe_float(value):
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def safe_int(value):
    try:
        return int(value or 0)
    except Exception:
        return 0


def parsed_date_expr(col):
    if not col:
        return "NULL"
    return (
        f"COALESCE("
        f"STR_TO_DATE(CAST({col} AS CHAR), '%%Y-%%m-%%d %%H:%%i:%%s'), "
        f"STR_TO_DATE(CAST({col} AS CHAR), '%%d-%%m-%%Y %%H:%%i:%%s'), "
        f"STR_TO_DATE(CAST({col} AS CHAR), '%%Y-%%m-%%d'), "
        f"STR_TO_DATE(CAST({col} AS CHAR), '%%d-%%m-%%Y'), "
        f"STR_TO_DATE(CAST({col} AS CHAR), '%%d/%%m/%%Y'), "
        f"DATE({col})"
        f")"
    )


def sql_date(col, alias_name, fmt="%d-%m-%Y"):
    if not col:
        return f"'' AS {alias_name}"
    safe_fmt = fmt.replace("%", "%%")
    normalized = parsed_date_expr(col)
    return (
        f"COALESCE("
        f"DATE_FORMAT({normalized}, '{safe_fmt}'), "
        f"CAST({col} AS CHAR), "
        f"''"
        f") AS {alias_name}"
    )


def sql_max_date(col, alias_name, fmt="%d-%m-%Y"):
    if not col:
        return f"'' AS {alias_name}"
    safe_fmt = fmt.replace("%", "%%")
    normalized = parsed_date_expr(col)
    return (
        f"COALESCE("
        f"DATE_FORMAT(MAX({normalized}), '{safe_fmt}'), "
        f"MAX(CAST({col} AS CHAR)), "
        f"''"
        f") AS {alias_name}"
    )


def status_text(sts):
    sts = str(sts or "0")
    if sts == "1":
        return "Approved"
    if sts == "2":
        return "Rejected"
    return "Pending"


def payment_status_text(sts):
    return "Paid" if str(sts or "0") == "1" else "Pending"


def auth_user_id(request):
    auth_header = request.headers.get("Authorization", "")
    token = auth_header.replace("Bearer", "", 1).strip()
    if token:
        return token
    return safe_str(request.data.get("user_id"))


def staff_name_sql(col, alias_name):
    if not col:
        return f"'' AS {alias_name}"
    return (
        f"CASE WHEN {col} IS NULL OR {col} = '' "
        f"THEN '' ELSE COALESCE(get_staff_name({col}), CAST({col} AS CHAR), '') END AS {alias_name}"
    )


def vendor_name_expr(vc_cols):
    for col in ("company_name", "vendor_name", "name"):
        if col in vc_cols:
            return f"v.{col}"
    return "''"


def build_join_vendor(vpd_cols, vc_cols):
    vendor_id_col = pick_col(vpd_cols, "vendor_id")
    if vendor_id_col and "unique_id" in vc_cols:
        return f"LEFT JOIN vendor_creation v ON v.unique_id = vb.{vendor_id_col} AND COALESCE(v.is_delete, 0) = 0"
    return ""


def list_context():
    vpd_cols = table_columns("vendor_payment_details")
    vc_cols = table_columns("vendor_creation")
    return vpd_cols, vc_cols


def build_list_query_parts(tab, search):
    vpd_cols, vc_cols = list_context()

    if not vpd_cols or "bill_no" not in vpd_cols:
        return None

    bill_date_col = pick_col(vpd_cols, "bill_date", "vendor_bill_created_date", "created")
    inv_date_col = pick_col(vpd_cols, "vendor_inv_attach_approval_date", "invoice_date")
    inv_no_col = pick_col(vpd_cols, "veninvverifyid", "invoice_no")
    amount_col = pick_col(vpd_cols, "total_amount", "bulk_total_amount", "amount")
    additional_charges_col = pick_col(vpd_cols, "additionalcharges", "additional_charges")
    dc_col = pick_col(vpd_cols, "dc_num", "dc_number")
    vendor_id_col = pick_col(vpd_cols, "vendor_id")
    bill_approval_col = pick_col(vpd_cols, "vendor_bill_approval")
    bill_approval_date_col = pick_col(vpd_cols, "vendor_bill_approval_date")
    accstatus_col = pick_col(vpd_cols, "accstatus")
    acc_reason_col = pick_col(vpd_cols, "acc_ent_rej_reason", "finance_reject_reason")
    acc_by_col = pick_col(vpd_cols, "vendor_account_approved_by", "finance_approved_by")
    mgmt_status_col = pick_col(vpd_cols, "managment_team_approval_sts")
    mgmt_reason_col = pick_col(vpd_cols, "managment_team_reject_reason")
    mgmt_by_col = pick_col(vpd_cols, "managment_team_approvedby")
    vendor_bill_status_col = pick_col(vpd_cols, "vendor_bill_app_status")
    tds_col = pick_col(vpd_cols, "acctdsvalue")
    other_col = pick_col(vpd_cols, "accotherdeduction")
    advance_col = pick_col(vpd_cols, "advancepayment")
    payable_col = pick_col(vpd_cols, "acctotalpaybleamount")
    order_col = pick_col(vpd_cols, "unique_id", "id", "created", "bill_date")

    join_vendor = build_join_vendor(vpd_cols, vc_cols)
    vendor_name_sql_expr = vendor_name_expr(vc_cols) if join_vendor else ("vb.vendor_name" if "vendor_name" in vpd_cols else "''")
    vendor_address_expr = "v.address" if join_vendor and "address" in vc_cols else "''"
    vendor_phone_expr = "v.contact_no" if join_vendor and "contact_no" in vc_cols else "''"

    status_expr = "0"
    reject_reason_expr = "''"
    rejected_by_expr = "''"

    if tab == "pending":
        if mgmt_status_col and accstatus_col:
            status_expr = f"CASE WHEN COALESCE(vb.{mgmt_status_col}, 0) = 2 THEN 2 ELSE COALESCE(vb.{accstatus_col}, 0) END"
        elif accstatus_col:
            status_expr = f"COALESCE(vb.{accstatus_col}, 0)"

        if mgmt_status_col and mgmt_reason_col and acc_reason_col:
            reject_reason_expr = (
                f"CASE WHEN COALESCE(vb.{mgmt_status_col}, 0) = 2 "
                f"THEN COALESCE(vb.{mgmt_reason_col}, '') ELSE COALESCE(vb.{acc_reason_col}, '') END"
            )
        elif acc_reason_col:
            reject_reason_expr = f"COALESCE(vb.{acc_reason_col}, '')"

        if mgmt_status_col and mgmt_by_col and acc_by_col:
            rejected_by_expr = (
                f"CASE WHEN COALESCE(vb.{mgmt_status_col}, 0) = 2 "
                f"THEN COALESCE(get_staff_name(vb.{mgmt_by_col}), CAST(vb.{mgmt_by_col} AS CHAR), '') "
                f"ELSE COALESCE(get_staff_name(vb.{acc_by_col}), CAST(vb.{acc_by_col} AS CHAR), '') END"
            )
        elif acc_by_col:
            rejected_by_expr = f"COALESCE(get_staff_name(vb.{acc_by_col}), CAST(vb.{acc_by_col} AS CHAR), '')"
    else:
        if accstatus_col:
            status_expr = f"COALESCE(vb.{accstatus_col}, 0)"

    where_parts = ["1=1"]
    params = []

    if "is_delete" in vpd_cols:
        where_parts.append("COALESCE(vb.is_delete, 0) = 0")

    if tab == "pending":
        if vendor_bill_status_col:
            where_parts.append(f"COALESCE(vb.{vendor_bill_status_col}, 0) = 1")
        if accstatus_col:
            where_parts.append(f"COALESCE(vb.{accstatus_col}, 0) = 0")
        if mgmt_status_col:
            where_parts.append(f"COALESCE(vb.{mgmt_status_col}, 0) = 0")
    else:
        if accstatus_col:
            where_parts.append(f"COALESCE(vb.{accstatus_col}, 0) IN (1, 2)")

    if search:
        like = f"%{search}%"
        search_parts = ["vb.bill_no LIKE %s"]
        params.append(like)

        if inv_no_col:
            search_parts.append(f"CAST(MAX(vb.{inv_no_col}) AS CHAR) LIKE %s")
            params.append(like)

        if vendor_name_sql_expr != "''":
            search_parts.append(f"MAX(CAST({vendor_name_sql_expr} AS CHAR)) LIKE %s")
            params.append(like)

        search_sql = " OR ".join(search_parts)
    else:
        search_sql = ""

    base_where = " AND ".join(where_parts)
    having_sql = f"HAVING ({search_sql})" if search_sql else ""

    grouped_sql = f"""
        FROM vendor_payment_details vb
        {join_vendor}
        WHERE {base_where}
        GROUP BY vb.bill_no{', vb.' + vendor_id_col if vendor_id_col else ''}
        {having_sql}
    """

    bill_value_expr = "0"
    if amount_col and additional_charges_col:
        bill_value_expr = f"ROUND(IFNULL(SUM(vb.{amount_col}), 0) + MAX(IFNULL(vb.{additional_charges_col}, 0)), 2)"
    elif amount_col:
        bill_value_expr = f"ROUND(IFNULL(SUM(vb.{amount_col}), 0), 2)"
    elif additional_charges_col:
        bill_value_expr = f"ROUND(MAX(IFNULL(vb.{additional_charges_col}, 0)), 2)"

    select_sql = f"""
        SELECT
            vb.bill_no,
            {sql_max_date(f'vb.{bill_date_col}', 'vendor_bill_date') if bill_date_col else "'' AS vendor_bill_date"},
            {sql_max_date(f'vb.{inv_date_col}', 'vendor_invoice_date') if inv_date_col else "'' AS vendor_invoice_date"},
            {f"MAX(CAST(vb.{inv_no_col} AS CHAR)) AS invoice_no" if inv_no_col else "'' AS invoice_no"},
            {f"MAX(CAST({vendor_name_sql_expr} AS CHAR)) AS vendor_name" if vendor_name_sql_expr != "''" else "'' AS vendor_name"},
            {f"MAX(CAST({vendor_address_expr} AS CHAR)) AS vendor_address" if vendor_address_expr != "''" else "'' AS vendor_address"},
            {f"MAX(CAST({vendor_phone_expr} AS CHAR)) AS vendor_phone" if vendor_phone_expr != "''" else "'' AS vendor_phone"},
            {f"MAX(CAST({vendor_name_sql_expr} AS CHAR)) AS vendor_details" if vendor_name_sql_expr != "''" else "'' AS vendor_details"},
            {f"COUNT(DISTINCT vb.{dc_col}) AS dc_count" if dc_col else "0 AS dc_count"},
            {bill_value_expr} AS bill_value,
            {f"ROUND(MAX(IFNULL(vb.{tds_col}, 0)), 2) AS tds_deduction" if tds_col else "0 AS tds_deduction"},
            {f"ROUND(MAX(IFNULL(vb.{other_col}, 0)), 2) AS other_deduction" if other_col else "0 AS other_deduction"},
            {f"ROUND(MAX(IFNULL(vb.{advance_col}, 0)), 2) AS advance_amount" if advance_col else "0 AS advance_amount"},
            {f"ROUND(MAX(IFNULL(vb.{payable_col}, 0)), 2) AS deduction_payable_bill_value" if payable_col else "0 AS deduction_payable_bill_value"},
            {staff_name_sql(f'MAX(vb.{bill_approval_col})', 'approved_by') if bill_approval_col else "'' AS approved_by"},
            {sql_max_date(f'vb.{bill_approval_date_col}', 'approved_date') if bill_approval_date_col else "'' AS approved_date"},
            MAX({status_expr}) AS status_code,
            MAX(CAST({reject_reason_expr} AS CHAR)) AS reject_reason,
            MAX(CAST({rejected_by_expr} AS CHAR)) AS rejected_by
            {f", MAX(CAST(vb.{vendor_id_col} AS CHAR)) AS vendor_id" if vendor_id_col else ", '' AS vendor_id"}
        {grouped_sql}
        ORDER BY {f'MAX(vb.{order_col}) DESC' if order_col else 'vb.bill_no DESC'}
    """

    return {
        "sql": select_sql,
        "count_sql": f"SELECT COUNT(*) AS total FROM ({select_sql}) grouped_rows",
        "params": params,
    }


def fetch_list_rows(tab, search, page=None, length=None):
    parts = build_list_query_parts(tab, search)
    if not parts:
        return [], 0

    total_row = db_fetchone(parts["count_sql"], parts["params"]) or {}
    total = safe_int(total_row.get("total"))

    sql = parts["sql"]
    params = list(parts["params"])

    if page is not None and length is not None:
        offset = max(page - 1, 0) * length
        sql += " LIMIT %s OFFSET %s"
        params.extend([length, offset])

    rows = db_fetch(sql, params)
    return rows, total


class AccountsBillEntryListView(APIView):
    def get(self, request):
        tab = request.query_params.get("tab", "pending").lower()
        search = safe_str(request.query_params.get("search")).strip()

        try:
            page = max(1, int(request.query_params.get("page", 1)))
        except Exception:
            page = 1

        try:
            length = max(1, int(request.query_params.get("length", 10)))
        except Exception:
            length = 10

        rows, total = fetch_list_rows(tab, search, page, length)

        data = []
        for idx, row in enumerate(rows, start=((page - 1) * length) + 1):
            data.append(
                {
                    "id": f"{safe_str(row.get('bill_no'))}-{safe_str(row.get('vendor_id')) or idx}",
                    "vendor_bill_date": safe_str(row.get("vendor_bill_date")),
                    "bill_no": safe_str(row.get("bill_no")),
                    "vendor_invoice_date": safe_str(row.get("vendor_invoice_date")),
                    "invoice_no": safe_str(row.get("invoice_no")),
                    "vendor_name": safe_str(row.get("vendor_name")),
                    "vendor_address": safe_str(row.get("vendor_address")),
                    "vendor_phone": safe_str(row.get("vendor_phone")),
                    "vendor_details": safe_str(row.get("vendor_details")),
                    "dc_count": safe_int(row.get("dc_count")),
                    "bill_value": safe_float(row.get("bill_value")),
                    "tds_deduction": safe_float(row.get("tds_deduction")),
                    "other_deduction": safe_float(row.get("other_deduction")),
                    "advance_amount": safe_float(row.get("advance_amount")),
                    "deduction_payable_bill_value": safe_float(row.get("deduction_payable_bill_value")),
                    "approved_by": safe_str(row.get("approved_by")),
                    "approved_date": safe_str(row.get("approved_date")),
                    "status": status_text(row.get("status_code")),
                    "reject_reason": safe_str(row.get("reject_reason")),
                    "rejected_by": safe_str(row.get("rejected_by")),
                }
            )

        return Response(
            {
                "data": data,
                "total": total,
                "page": page,
                "pages": max(1, -(-total // length)),
            }
        )


class AccountsBillEntryDetailView(APIView):
    def get(self, request):
        bill_no = safe_str(request.query_params.get("bill_no")).strip()
        if not bill_no:
            return Response({"error": "bill_no required"}, status=status.HTTP_400_BAD_REQUEST)

        vpd_cols, vc_cols = list_context()
        if not vpd_cols or "bill_no" not in vpd_cols:
            return Response({"error": "Bill not found"}, status=status.HTTP_404_NOT_FOUND)

        join_vendor = build_join_vendor(vpd_cols, vc_cols)

        vc_company = pick_col(vc_cols, "company_name", "vendor_name", "name")
        vc_address = pick_col(vc_cols, "address")
        vc_gst = pick_col(vc_cols, "gst_no")
        vc_pan = pick_col(vc_cols, "pan_no")
        vc_mail = pick_col(vc_cols, "mail_id")
        vc_phone = pick_col(vc_cols, "contact_no")
        vc_bank = pick_col(vc_cols, "bank_name")
        vc_branch = pick_col(vc_cols, "branch_name")
        vc_account = pick_col(vc_cols, "account_no")
        vc_ifsc = pick_col(vc_cols, "ifsc_code")
        vc_holder = pick_col(vc_cols, "acc_holder_name", "account_holder_name")
        vc_pan_copy = pick_col(vc_cols, "pan_attach_file_name", "pan_copy")
        vc_bank_proof = pick_col(vc_cols, "bank_proof")

        bill_date_col = pick_col(vpd_cols, "bill_date", "vendor_bill_created_date", "created")
        invoice_no_col = pick_col(vpd_cols, "veninvverifyid", "invoice_no")
        invoice_date_col = pick_col(vpd_cols, "vendor_inv_attach_approval_date", "invoice_date")
        invoice_file_col = pick_col(vpd_cols, "inv_verfiy_attach", "vendor_inv_attach")
        po_file_col = pick_col(vpd_cols, "po_ven_filename")
        tds_col = pick_col(vpd_cols, "acctdsvalue")
        other_deduction_col = pick_col(vpd_cols, "accotherdeduction")
        advance_col = pick_col(vpd_cols, "advancepayment")
        additional_charges_col = pick_col(vpd_cols, "additionalcharges", "additional_charges")
        payable_col = pick_col(vpd_cols, "acctotalpaybleamount")
        total_tds_col = pick_col(vpd_cols, "Totaltdsamount")
        remarks_col = pick_col(vpd_cols, "accdetuctionremarks")
        account_bill_id_col = pick_col(vpd_cols, "accountbillid")

        bill_row_sql = f"""
            SELECT
                vb.bill_no,
                {sql_date(f'vb.{bill_date_col}', 'vendor_bill_date') if bill_date_col else "'' AS vendor_bill_date"},
                {f"CAST(vb.{invoice_no_col} AS CHAR) AS vendor_invoice_no" if invoice_no_col else "'' AS vendor_invoice_no"},
                {sql_date(f'vb.{invoice_date_col}', 'vendor_invoice_date') if invoice_date_col else "'' AS vendor_invoice_date"},
                {f"CAST(vb.{invoice_file_col} AS CHAR) AS invoice_attach_url" if invoice_file_col else "'#' AS invoice_attach_url"},
                {f"CAST(vb.{po_file_col} AS CHAR) AS po_attach_url" if po_file_col else "'#' AS po_attach_url"},
                {f"IFNULL(vb.{tds_col}, 0) AS tds_deduction" if tds_col else "0 AS tds_deduction"},
                {f"IFNULL(vb.{other_deduction_col}, 0) AS others_deduction" if other_deduction_col else "0 AS others_deduction"},
                {f"IFNULL(vb.{advance_col}, 0) AS advance_amount" if advance_col else "0 AS advance_amount"},
                {f"IFNULL(vb.{additional_charges_col}, 0) AS additional_charges" if additional_charges_col else "0 AS additional_charges"},
                {f"IFNULL(vb.{payable_col}, 0) AS total_payable" if payable_col else "0 AS total_payable"},
                {f"IFNULL(vb.{total_tds_col}, 0) AS total_tds_amount" if total_tds_col else "0 AS total_tds_amount"},
                {f"CAST(vb.{remarks_col} AS CHAR) AS remarks" if remarks_col else "'' AS remarks"},
                {f"CAST(vb.{account_bill_id_col} AS CHAR) AS account_bill_id" if account_bill_id_col else "'' AS account_bill_id"},
                {f"CAST(v.{vc_company} AS CHAR) AS vendor_name" if (join_vendor and vc_company) else "'' AS vendor_name"},
                {f"CAST(v.{vc_address} AS CHAR) AS vendor_address" if (join_vendor and vc_address) else "'' AS vendor_address"},
                {f"CAST(v.{vc_gst} AS CHAR) AS vendor_gst" if (join_vendor and vc_gst) else "'' AS vendor_gst"},
                {f"CAST(v.{vc_pan} AS CHAR) AS vendor_pan" if (join_vendor and vc_pan) else "'' AS vendor_pan"},
                {f"CAST(v.{vc_mail} AS CHAR) AS vendor_email" if (join_vendor and vc_mail) else "'' AS vendor_email"},
                {f"CAST(v.{vc_phone} AS CHAR) AS vendor_phone" if (join_vendor and vc_phone) else "'' AS vendor_phone"},
                {f"CAST(v.{vc_bank} AS CHAR) AS bank_name" if (join_vendor and vc_bank) else "'' AS bank_name"},
                {f"CAST(v.{vc_branch} AS CHAR) AS branch" if (join_vendor and vc_branch) else "'' AS branch"},
                {f"CAST(v.{vc_account} AS CHAR) AS account_no" if (join_vendor and vc_account) else "'' AS account_no"},
                {f"CAST(v.{vc_ifsc} AS CHAR) AS ifsc_code" if (join_vendor and vc_ifsc) else "'' AS ifsc_code"},
                {f"CAST(v.{vc_holder} AS CHAR) AS account_holder" if (join_vendor and vc_holder) else "'' AS account_holder"},
                {f"CAST(v.{vc_pan_copy} AS CHAR) AS pan_copy_url" if (join_vendor and vc_pan_copy) else "'#' AS pan_copy_url"},
                {f"CAST(v.{vc_bank_proof} AS CHAR) AS bank_proof_url" if (join_vendor and vc_bank_proof) else "'#' AS bank_proof_url"}
            FROM vendor_payment_details vb
            {join_vendor}
            WHERE vb.bill_no = %s
            {"AND COALESCE(vb.is_delete, 0) = 0" if "is_delete" in vpd_cols else ""}
            LIMIT 1
        """
        bill_row = db_fetchone(bill_row_sql, [bill_no])
        if not bill_row:
            return Response({"error": "Bill not found"}, status=status.HTTP_404_NOT_FOUND)

        dc_no_col = pick_col(vpd_cols, "dc_num", "dc_number")
        dc_date_col = pick_col(vpd_cols, "dc_date")
        item_invoice_no_col = pick_col(vpd_cols, "invoice_no")
        item_invoice_date_col = pick_col(vpd_cols, "invoice_date", "vendor_inv_attach_approval_date")
        po_no_col = pick_col(vpd_cols, "po_num")
        po_date_col = pick_col(vpd_cols, "po_date")
        consignee_address_col = pick_col(vpd_cols, "consignee_address")
        consignee_alt_col = pick_col(vpd_cols, "con_address")
        po_form_unique_id_col = pick_col(vpd_cols, "po_form_unique_id", "form_main_unique_id")
        row_vendor_id_col = pick_col(vpd_cols, "vendor_id")
        qty_col = pick_col(vpd_cols, "invoice_qty")
        unit_price_col = pick_col(vpd_cols, "rate", "unit_price", "vendor_bulk_rate")
        gst_col = pick_col(vpd_cols, "gst", "vendor_bulk_gst")
        gst_amount_col = pick_col(vpd_cols, "gst_amount")
        total_amount_col = pick_col(vpd_cols, "total_amount", "bulk_total_amount", "amount")
        order_col = pick_col(vpd_cols, "unique_id", "id")

        consignee_fallback_filters = []
        if po_form_unique_id_col:
            consignee_fallback_filters.append(f"vv.form_main_unique_id = vb.{po_form_unique_id_col}")
        if dc_no_col:
            if row_vendor_id_col:
                consignee_fallback_filters.append(
                    f"(vv.dc_number = vb.{dc_no_col} AND vv.engineer_name = vb.{row_vendor_id_col})"
                )
            else:
                consignee_fallback_filters.append(f"vv.dc_number = vb.{dc_no_col}")

        consignee_fallback_expr = "''"
        if consignee_fallback_filters:
            consignee_fallback_expr = (
                "("
                "SELECT COALESCE(get_address(vv.consignee_unique_id), '') "
                "FROM view_outsource_vendor_verified_invoice vv "
                f"WHERE {' OR '.join(consignee_fallback_filters)} "
                "ORDER BY vv.dc_date DESC, vv.unique_id DESC "
                "LIMIT 1"
                ")"
            )

        consignee_expr_parts = []
        if consignee_address_col:
            consignee_expr_parts.append(f"NULLIF(vb.{consignee_address_col}, '')")
        if consignee_alt_col and consignee_alt_col != consignee_address_col:
            consignee_expr_parts.append(f"NULLIF(vb.{consignee_alt_col}, '')")
        if consignee_fallback_filters:
            consignee_expr_parts.append(consignee_fallback_expr)
        consignee_expr_parts.append("''")
        consignee_expr = f"COALESCE({', '.join(consignee_expr_parts)})"

        dc_sql = f"""
            SELECT
                {f"CAST(vb.{dc_no_col} AS CHAR) AS dc_no" if dc_no_col else "'' AS dc_no"},
                {sql_date(f'vb.{dc_date_col}', 'dc_date') if dc_date_col else "'' AS dc_date"},
                {f"CAST(vb.{item_invoice_no_col} AS CHAR) AS invoice_no" if item_invoice_no_col else "'' AS invoice_no"},
                {sql_date(f'vb.{item_invoice_date_col}', 'invoice_date') if item_invoice_date_col else "'' AS invoice_date"},
                {f"CAST(vb.{po_no_col} AS CHAR) AS po_no" if po_no_col else "'' AS po_no"},
                {sql_date(f'vb.{po_date_col}', 'po_date') if po_date_col else "'' AS po_date"},
                {consignee_expr} AS consignee_address,
                {f"IFNULL(vb.{qty_col}, 0) AS invoice_qty" if qty_col else "0 AS invoice_qty"},
                {f"IFNULL(vb.{unit_price_col}, 0) AS unit_price" if unit_price_col else "0 AS unit_price"},
                {f"IFNULL(vb.{gst_col}, 0) AS gst" if gst_col else "0 AS gst"},
                {f"IFNULL(vb.{gst_amount_col}, 0) AS gst_amount" if gst_amount_col else "0 AS gst_amount"},
                {f"IFNULL(vb.{total_amount_col}, 0) AS total_amount" if total_amount_col else "0 AS total_amount"}
            FROM vendor_payment_details vb
            WHERE vb.bill_no = %s
            {"AND COALESCE(vb.is_delete, 0) = 0" if "is_delete" in vpd_cols else ""}
            ORDER BY {f'vb.{order_col}' if order_col else 'vb.bill_no'}
        """
        dc_rows = db_fetch(dc_sql, [bill_no])

        dc_items = []
        total_amount = 0.0
        total_basic_amount = 0.0
        total_gst_amount = 0.0
        gst_labels = []
        for idx, row in enumerate(dc_rows, start=1):
            qty = safe_float(row.get("invoice_qty"))
            unit_price = safe_float(row.get("unit_price"))
            gst = safe_float(row.get("gst"))
            line_total = safe_float(row.get("total_amount"))
            basic_amount = qty * unit_price
            gst_amount = safe_float(row.get("gst_amount"))
            if not gst_amount and basic_amount and gst:
                gst_amount = (basic_amount * gst) / 100
            if not line_total and basic_amount:
                line_total = basic_amount + gst_amount
            gst_label = f"{gst:g} %" if gst else "0 %"
            total_amount += line_total
            total_basic_amount += basic_amount
            total_gst_amount += gst_amount
            if gst_label not in gst_labels:
                gst_labels.append(gst_label)

            dc_items.append(
                {
                    "s_no": idx,
                    "dc_no": safe_str(row.get("dc_no")),
                    "dc_date": safe_str(row.get("dc_date")),
                    "invoice_no": safe_str(row.get("invoice_no")),
                    "invoice_date": safe_str(row.get("invoice_date")),
                    "po_no": safe_str(row.get("po_no")),
                    "po_date": safe_str(row.get("po_date")),
                    "consignee_address": safe_str(row.get("consignee_address")),
                    "unit_price": unit_price,
                    "basic_amount": basic_amount,
                    "gst": gst_label,
                    "gst_amount": gst_amount,
                    "total_amount": line_total,
                }
            )

        tds_deduction = safe_float(bill_row.get("tds_deduction"))
        others_deduction = safe_float(bill_row.get("others_deduction"))
        advance_amount = safe_float(bill_row.get("advance_amount"))
        additional_charges = safe_float(bill_row.get("additional_charges"))
        total_payable = safe_float(bill_row.get("total_payable")) or (total_amount - tds_deduction - others_deduction - advance_amount + additional_charges)
        total_tds_amount = safe_float(bill_row.get("total_tds_amount"))
        if not total_tds_amount:
            total_tds_amount = max(total_amount - tds_deduction, 0.0)
        gst_percentage = ", ".join(gst_labels) if gst_labels else "0 %"
        with_gst_total_amount = total_basic_amount + total_gst_amount

        app_bill_created_by = pick_col(vpd_cols, "vendor_bill_created_by", "bill_created_by")
        app_bill_created_at = pick_col(vpd_cols, "vendor_bill_created_date", "bill_created_date")
        app_op_by = pick_col(vpd_cols, "vendor_bill_approval", "bill_approved")
        app_op_at = pick_col(vpd_cols, "vendor_bill_approval_date")
        app_op_sts = pick_col(vpd_cols, "vendor_bill_app_status")
        app_acc_by = pick_col(vpd_cols, "vendor_account_approved_by", "acc_approve")
        app_acc_at = pick_col(vpd_cols, "vendor_account_approval_date")
        app_acc_sts = pick_col(vpd_cols, "acc_ent_sts")
        app_accounts_by = pick_col(vpd_cols, "finance_approved_by", "finance_approve")
        app_accounts_at = pick_col(vpd_cols, "finance_approved_date")
        app_accounts_sts = pick_col(vpd_cols, "finance_approval", "finance_approval_sts")
        app_mgmt_by = pick_col(vpd_cols, "managment_team_approvedby", "managed_approve")
        app_mgmt_at = pick_col(vpd_cols, "managment_team_approvaldate")
        app_mgmt_sts = pick_col(vpd_cols, "managment_team_approval_sts")
        app_pay_amount = pick_col(vpd_cols, "acctotalpaybleamount", "transfer_amount")
        app_pay_status = pick_col(vpd_cols, "accounts_approval", "payment_status")

        approval_sql = f"""
            SELECT
                {staff_name_sql(f'MAX(vb.{app_bill_created_by})', 'bill_created_by') if app_bill_created_by else "'' AS bill_created_by"},
                {sql_max_date(f'vb.{app_bill_created_at}', 'bill_created_at', '%d-%m-%Y %H:%i:%s') if app_bill_created_at else "'' AS bill_created_at"},
                {staff_name_sql(f'MAX(vb.{app_op_by})', 'operation_by') if app_op_by else "'' AS operation_by"},
                {sql_max_date(f'vb.{app_op_at}', 'operation_at', '%d-%m-%Y %H:%i:%s') if app_op_at else "'' AS operation_at"},
                {f"MAX(IFNULL(vb.{app_op_sts}, 0)) AS operation_status" if app_op_sts else "0 AS operation_status"},
                {staff_name_sql(f'MAX(vb.{app_acc_by})', 'account_entry_by') if app_acc_by else "'' AS account_entry_by"},
                {sql_max_date(f'vb.{app_acc_at}', 'account_entry_at', '%d-%m-%Y %H:%i:%s') if app_acc_at else "'' AS account_entry_at"},
                {f"MAX(IFNULL(vb.{app_acc_sts}, 0)) AS account_entry_status" if app_acc_sts else "0 AS account_entry_status"},
                {staff_name_sql(f'MAX(vb.{app_accounts_by})', 'accounts_approval_by') if app_accounts_by else "'' AS accounts_approval_by"},
                {sql_max_date(f'vb.{app_accounts_at}', 'accounts_approval_at', '%d-%m-%Y %H:%i:%s') if app_accounts_at else "'' AS accounts_approval_at"},
                {f"MAX(IFNULL(vb.{app_accounts_sts}, 0)) AS accounts_approval_status" if app_accounts_sts else "0 AS accounts_approval_status"},
                {staff_name_sql(f'MAX(vb.{app_mgmt_by})', 'management_by') if app_mgmt_by else "'' AS management_by"},
                {sql_max_date(f'vb.{app_mgmt_at}', 'management_at', '%d-%m-%Y %H:%i:%s') if app_mgmt_at else "'' AS management_at"},
                {f"MAX(IFNULL(vb.{app_mgmt_sts}, 0)) AS management_status" if app_mgmt_sts else "0 AS management_status"},
                {f"MAX(IFNULL(vb.{app_pay_amount}, 0)) AS payment_amount" if app_pay_amount else "0 AS payment_amount"},
                {f"MAX(IFNULL(vb.{app_pay_status}, 0)) AS payment_status" if app_pay_status else "0 AS payment_status"}
            FROM vendor_payment_details vb
            WHERE vb.bill_no = %s
            {"AND COALESCE(vb.is_delete, 0) = 0" if "is_delete" in vpd_cols else ""}
        """
        approval_row = db_fetchone(approval_sql, [bill_no]) or {}

        return Response(
            {
                "id": safe_str(bill_row.get("bill_no")),
                "vendor_name": safe_str(bill_row.get("vendor_name")),
                "vendor_company_name": safe_str(bill_row.get("vendor_name")),
                "vendor_address": safe_str(bill_row.get("vendor_address")),
                "vendor_gst": safe_str(bill_row.get("vendor_gst")),
                "vendor_pan": safe_str(bill_row.get("vendor_pan")),
                "vendor_email": safe_str(bill_row.get("vendor_email")),
                "vendor_phone": safe_str(bill_row.get("vendor_phone")),
                "vendor_bill_no": safe_str(bill_row.get("bill_no")),
                "vendor_bill_date": safe_str(bill_row.get("vendor_bill_date")),
                "vendor_invoice_no": safe_str(bill_row.get("vendor_invoice_no")),
                "vendor_invoice_date": safe_str(bill_row.get("vendor_invoice_date")),
                "account_bill_id": safe_str(bill_row.get("account_bill_id")),
                "remarks": safe_str(bill_row.get("remarks")),
                "invoice_attach_url": safe_str(bill_row.get("invoice_attach_url")) or "#",
                "po_attach_url": safe_str(bill_row.get("po_attach_url")) or "#",
                "bank_name": safe_str(bill_row.get("bank_name")),
                "branch": safe_str(bill_row.get("branch")),
                "account_no": safe_str(bill_row.get("account_no")),
                "ifsc_code": safe_str(bill_row.get("ifsc_code")),
                "account_holder": safe_str(bill_row.get("account_holder")),
                "pan_copy_url": safe_str(bill_row.get("pan_copy_url")) or "#",
                "bank_proof_url": safe_str(bill_row.get("bank_proof_url")) or "#",
                "dc_items": dc_items,
                "basic_amount": total_basic_amount,
                "gst_percentage": gst_percentage,
                "gst_amount": total_gst_amount,
                "with_gst_total_amount": with_gst_total_amount,
                "total_amount": total_amount,
                "tds_deduction": tds_deduction,
                "total_tds_amount": total_tds_amount,
                "others_deduction": others_deduction,
                "advance_amount": advance_amount,
                "additional_charges": additional_charges,
                "total_payable": total_payable,
                "approvals": [
                    {
                        "s_no": 1,
                        "bill_created_by": safe_str(approval_row.get("bill_created_by")),
                        "bill_created_at": safe_str(approval_row.get("bill_created_at")),
                        "bill_created_status": "Bill Created",
                        "operation_by": safe_str(approval_row.get("operation_by")),
                        "operation_at": safe_str(approval_row.get("operation_at")),
                        "operation_status": status_text(approval_row.get("operation_status")),
                        "account_entry_by": safe_str(approval_row.get("account_entry_by")),
                        "account_entry_at": safe_str(approval_row.get("account_entry_at")),
                        "account_entry_status": status_text(approval_row.get("account_entry_status")),
                        "accounts_approval_by": safe_str(approval_row.get("accounts_approval_by")),
                        "accounts_approval_at": safe_str(approval_row.get("accounts_approval_at")),
                        "accounts_approval_status": status_text(approval_row.get("accounts_approval_status")),
                        "management_by": safe_str(approval_row.get("management_by")),
                        "management_at": safe_str(approval_row.get("management_at")),
                        "management_status": status_text(approval_row.get("management_status")),
                        "payment_amount": f"{safe_float(approval_row.get('payment_amount')):,.2f}",
                        "payment_status": payment_status_text(approval_row.get("payment_status")),
                    }
                ],
            }
        )


class AccountsBillEntrySaveView(APIView):
    def post(self, request):
        bill_no = safe_str(request.data.get("bill_no")).strip()
        if not bill_no:
            return Response({"status": 0, "msg": "bill_no required"}, status=status.HTTP_400_BAD_REQUEST)

        raw_tds_percentage = request.data.get("tds_percentage")
        raw_total_tds_amount = request.data.get("total_tds_amount")
        tds_percentage = safe_float(raw_tds_percentage)
        tds_deduction = safe_float(request.data.get("tds_deduction"))
        others_deduction = safe_float(request.data.get("others_deduction"))
        advance_amount = safe_float(request.data.get("advance_amount"))
        additional_charges = safe_float(request.data.get("additional_charges"))
        total_payable = safe_float(request.data.get("total_payable"))
        total_tds_amount = (
            safe_float(raw_total_tds_amount)
            if safe_str(raw_total_tds_amount).strip()
            else total_payable + others_deduction + advance_amount - additional_charges
        )
        remarks = safe_str(request.data.get("remarks")).strip()
        account_bill_id = safe_str(request.data.get("account_bill_id")).strip()
        user_id = auth_user_id(request)

        additional_charges_col = ensure_decimal_column("vendor_payment_details", "additionalcharges", "additional_charges")
        additional_charges_main_col = ensure_decimal_column("vendor_payment_details_main", "additionalcharges", "additional_charges")
        vpd_cols = table_columns("vendor_payment_details")
        vpm_cols = table_columns("vendor_payment_details_main")
        if not vpd_cols:
            return Response({"status": 0, "msg": "Table not available"}, status=status.HTTP_400_BAD_REQUEST)

        update_map = {}
        if "tds" in vpd_cols and safe_str(raw_tds_percentage).strip():
            update_map["tds"] = tds_percentage
        if "acctdsvalue" in vpd_cols:
            update_map["acctdsvalue"] = tds_deduction
        if "accotherdeduction" in vpd_cols:
            update_map["accotherdeduction"] = others_deduction
        if "advancepayment" in vpd_cols:
            update_map["advancepayment"] = advance_amount
        if additional_charges_col and additional_charges_col in vpd_cols:
            update_map[additional_charges_col] = additional_charges
        if "acctotalpaybleamount" in vpd_cols:
            update_map["acctotalpaybleamount"] = total_payable
        if "accdetuctionremarks" in vpd_cols:
            update_map["accdetuctionremarks"] = remarks
        if "accstatus" in vpd_cols:
            update_map["accstatus"] = 1
        if "acc_ent_sts" in vpd_cols:
            update_map["acc_ent_sts"] = 1
        if "acc_ent_rej_reason" in vpd_cols:
            update_map["acc_ent_rej_reason"] = ""
        if "finance_approval" in vpd_cols:
            update_map["finance_approval"] = 0
        if "finance_reject_reason" in vpd_cols:
            update_map["finance_reject_reason"] = ""
        if "finance_remark" in vpd_cols:
            update_map["finance_remark"] = ""
        if "finance_approved_by" in vpd_cols:
            update_map["finance_approved_by"] = ""
        if "vendor_account_approved_by" in vpd_cols:
            update_map["vendor_account_approved_by"] = user_id
        if "accountbillid" in vpd_cols:
            update_map["accountbillid"] = account_bill_id or user_id

        if not update_map:
            return Response({"status": 0, "msg": "No updatable fields found"}, status=status.HTTP_400_BAD_REQUEST)

        set_parts = []
        values = []
        for col, value in update_map.items():
            set_parts.append(f"{col} = %s")
            values.append(value)

        if "vendor_account_approval_date" in vpd_cols:
            set_parts.append("vendor_account_approval_date = NOW()")
        if "finance_approved_date" in vpd_cols:
            set_parts.append("finance_approved_date = NULL")
        if "Totaltdsamount" in vpd_cols:
            set_parts.append("Totaltdsamount = %s")
            values.append(total_tds_amount)

        where_sql = "bill_no = %s"
        values.append(bill_no)
        if "is_delete" in vpd_cols:
            where_sql += " AND COALESCE(is_delete, 0) = 0"

        with connection.cursor() as cursor:
            cursor.execute(
                f"UPDATE vendor_payment_details SET {', '.join(set_parts)} WHERE {where_sql}",
                values,
            )

            main_parts = []
            main_values = []
            for col, value in update_map.items():
                if col in vpm_cols:
                    main_parts.append(f"{col} = %s")
                    main_values.append(value)

            if (
                additional_charges_main_col
                and additional_charges_main_col != additional_charges_col
                and additional_charges_main_col in vpm_cols
            ):
                main_parts.append(f"{additional_charges_main_col} = %s")
                main_values.append(additional_charges)

            if "vendor_account_approval_date" in vpm_cols:
                main_parts.append("vendor_account_approval_date = NOW()")
            if "finance_approved_date" in vpm_cols:
                main_parts.append("finance_approved_date = NULL")
            if "Totaltdsamount" in vpm_cols:
                main_parts.append("Totaltdsamount = %s")
                main_values.append(total_tds_amount)

            if main_parts:
                main_where = "bill_no = %s"
                main_values.append(bill_no)
                if "is_delete" in vpm_cols:
                    main_where += " AND COALESCE(is_delete, 0) = 0"
                cursor.execute(
                    f"UPDATE vendor_payment_details_main SET {', '.join(main_parts)} WHERE {main_where}",
                    main_values,
                )

        return Response({"status": 1, "msg": "Accounts bill entry saved"})


class AccountsBillEntryRejectView(APIView):
    def post(self, request):
        bill_no = safe_str(request.data.get("bill_no")).strip()
        reject_reason = safe_str(request.data.get("reject_reason")).strip()
        if not bill_no:
            return Response({"status": 0, "msg": "bill_no required"}, status=status.HTTP_400_BAD_REQUEST)
        if not reject_reason:
            return Response({"status": 0, "msg": "reject_reason required"}, status=status.HTTP_400_BAD_REQUEST)

        user_id = auth_user_id(request)
        vpd_cols = table_columns("vendor_payment_details")
        vpm_cols = table_columns("vendor_payment_details_main")
        if not vpd_cols:
            return Response({"status": 0, "msg": "Table not available"}, status=status.HTTP_400_BAD_REQUEST)

        update_map = {}
        if "accstatus" in vpd_cols:
            update_map["accstatus"] = 2
        if "acc_ent_sts" in vpd_cols:
            update_map["acc_ent_sts"] = 2
        if "acc_ent_rej_reason" in vpd_cols:
            update_map["acc_ent_rej_reason"] = reject_reason
        if "finance_reject_reason" in vpd_cols:
            update_map["finance_reject_reason"] = reject_reason
        if "finance_approval" in vpd_cols:
            update_map["finance_approval"] = 0
        if "finance_remark" in vpd_cols:
            update_map["finance_remark"] = ""
        if "finance_approved_by" in vpd_cols:
            update_map["finance_approved_by"] = ""
        if "vendor_account_approved_by" in vpd_cols:
            update_map["vendor_account_approved_by"] = user_id

        if not update_map:
            return Response({"status": 0, "msg": "No updatable fields found"}, status=status.HTTP_400_BAD_REQUEST)

        set_parts = []
        values = []
        for col, value in update_map.items():
            set_parts.append(f"{col} = %s")
            values.append(value)

        if "vendor_account_approval_date" in vpd_cols:
            set_parts.append("vendor_account_approval_date = NOW()")
        if "finance_approved_date" in vpd_cols:
            set_parts.append("finance_approved_date = NULL")

        where_sql = "bill_no = %s"
        values.append(bill_no)
        if "is_delete" in vpd_cols:
            where_sql += " AND COALESCE(is_delete, 0) = 0"

        with connection.cursor() as cursor:
            cursor.execute(
                f"UPDATE vendor_payment_details SET {', '.join(set_parts)} WHERE {where_sql}",
                values,
            )

            main_parts = []
            main_values = []
            for col, value in update_map.items():
                if col in vpm_cols:
                    main_parts.append(f"{col} = %s")
                    main_values.append(value)

            if "vendor_account_approval_date" in vpm_cols:
                main_parts.append("vendor_account_approval_date = NOW()")
            if "finance_approved_date" in vpm_cols:
                main_parts.append("finance_approved_date = NULL")

            if main_parts:
                main_where = "bill_no = %s"
                main_values.append(bill_no)
                if "is_delete" in vpm_cols:
                    main_where += " AND COALESCE(is_delete, 0) = 0"
                cursor.execute(
                    f"UPDATE vendor_payment_details_main SET {', '.join(main_parts)} WHERE {main_where}",
                    main_values,
                )

            # Reset DCs back to Stage 1 (Vendor Bill Creation) pending
            try:
                cursor.execute(
                    """
                    UPDATE invoice_verfication_table iv
                    JOIN vendor_payment_details vpd ON vpd.dc_num = iv.dc_number
                    SET iv.vendor_payment_allocated = 0,
                        iv.vendor_bill_no = ''
                    WHERE vpd.bill_no = %s AND COALESCE(vpd.is_delete, 0) = 0
                    """,
                    [bill_no],
                )
            except Exception:
                pass

        return Response({"status": 1, "msg": "Accounts bill entry rejected"})


class AccountsBillEntryExportView(APIView):
    def get(self, request):
        tab = request.query_params.get("tab", "pending").lower()
        search = safe_str(request.query_params.get("search")).strip()
        rows, _ = fetch_list_rows(tab, search)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Accounts Bill Entry"

        ws.merge_cells("A1:K1")
        ws["A1"] = "Accounts Team Bill Entry"
        ws["A1"].font = Font(bold=True, size=13)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = [
            "S.No",
            "Vendor Bill Date",
            "Bill No",
            "Vendor Invoice Date",
            "Invoice No",
            "Vendor Details",
            "DC Count",
            "Bill Value",
            "Approved By",
            "Approved Date",
            "Status",
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = Font(bold=True)

        for idx, row in enumerate(rows, start=1):
            ws.append(
                [
                    idx,
                    safe_str(row.get("vendor_bill_date")),
                    safe_str(row.get("bill_no")),
                    safe_str(row.get("vendor_invoice_date")),
                    safe_str(row.get("invoice_no")),
                    safe_str(row.get("vendor_details")),
                    safe_int(row.get("dc_count")),
                    safe_float(row.get("bill_value")),
                    safe_str(row.get("approved_by")),
                    safe_str(row.get("approved_date")),
                    status_text(row.get("status_code")),
                ]
            )

        for col in ws.columns:
            letter = col[0].column_letter
            width = max(len(str(cell.value or "")) for cell in col) + 4
            ws.column_dimensions[letter].width = min(width, 42)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="AccountsBillEntry.xlsx"'
        return response


