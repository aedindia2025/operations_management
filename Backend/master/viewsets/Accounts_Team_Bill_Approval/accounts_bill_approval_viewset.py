import io
import os
from pathlib import Path
from urllib.parse import quote

import openpyxl
from django.conf import settings
from django.db import connection
from django.http import FileResponse, Http404, HttpResponse
from openpyxl.styles import Alignment, Font
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from master.viewsets.Management_Team_Bill_Approval.management_viewset import (
    db_fetch,
    db_fetchone,
    get_user_id,
    payment_status_text,
    pick_col,
    safe_float,
    safe_int,
    safe_str,
    status_text,
    table_columns,
)
from master.viewsets.accounts_team_bill_entry.accounts_bill_entry_viewset import (
    parsed_date_expr,
    sql_date,
    sql_max_date,
)


LEGACY_UPLOAD_BASE = Path(
    os.environ.get('OTM_UPLOAD_BASE', r"Z:\xampp\htdocs\otm_beta\uploads")
)


def max_expr(col, alias_name, cast_text=False):
    if not col:
        return f"'' AS {alias_name}" if cast_text else f"0 AS {alias_name}"
    if cast_text:
        return f"MAX(CAST({col} AS CHAR)) AS {alias_name}"
    return f"MAX({col}) AS {alias_name}"


def staff_name_expr(col, alias_name):
    if not col:
        return f"'' AS {alias_name}"
    return (
        f"CASE WHEN MAX({col}) IS NULL OR MAX({col}) = '' THEN '' "
        f"ELSE COALESCE(get_staff_name(MAX({col})), CAST(MAX({col}) AS CHAR), '') END AS {alias_name}"
    )


def vendor_join(alias, source_cols, vendor_cols):
    vendor_id_col = pick_col(source_cols, 'vendor_id')
    if vendor_id_col and 'unique_id' in vendor_cols:
        return f"LEFT JOIN vendor_creation v ON v.unique_id = {alias}.{vendor_id_col} AND COALESCE(v.is_delete, 0) = 0"
    return ''


def vendor_name_sql(alias, source_cols, vendor_cols):
    source_name = pick_col(source_cols, 'vendor_name')
    company = pick_col(vendor_cols, 'company_name', 'vendor_name', 'name')
    if company and source_name:
        return f"COALESCE(v.{company}, {alias}.{source_name}, '')"
    if company:
        return f"COALESCE(v.{company}, '')"
    if source_name:
        return f"COALESCE({alias}.{source_name}, '')"
    return "''"


def vendor_field_sql(vendor_cols, *names):
    col = pick_col(vendor_cols, *names)
    return f"COALESCE(v.{col}, '')" if col else "''"


def bill_file_url(request, file_kind, filename):
    value = safe_str(filename).strip()
    if not value:
        return '#'
    if value.startswith(('http://', 'https://', '/api/', '/media/')):
        if request and value.startswith('/'):
            return request.build_absolute_uri(value)
        return value
    safe_name = Path(value).name
    base = f"/api/master/accounts-bill-approval/files/{file_kind}/{quote(safe_name)}/"
    return request.build_absolute_uri(base) if request else base


def bill_file_candidates(file_kind, filename):
    safe_name = Path(safe_str(filename)).name
    if not safe_name:
        return []

    folder_map = {
        'invoice': [
            LEGACY_UPLOAD_BASE / 'vendorpayment',
            Path(settings.MEDIA_ROOT) / 'vendorpayment',
            Path(settings.BASE_DIR) / 'uploads' / 'vendorpayment',
        ],
        'po': [
            LEGACY_UPLOAD_BASE / 'po_form' / 'PO copy',
            LEGACY_UPLOAD_BASE / 'po_form',
            LEGACY_UPLOAD_BASE / 'purchase_order' / 'PO copy',
            LEGACY_UPLOAD_BASE / 'purchase_order',
            LEGACY_UPLOAD_BASE / 'vendorpayment' / 'PO copy',
            LEGACY_UPLOAD_BASE / 'vendorpayment',
            Path(settings.MEDIA_ROOT) / 'po_form' / 'PO copy',
            Path(settings.MEDIA_ROOT) / 'po_form',
            Path(settings.MEDIA_ROOT) / 'purchase_order' / 'PO copy',
            Path(settings.MEDIA_ROOT) / 'purchase_order',
            Path(settings.MEDIA_ROOT) / 'vendorpayment' / 'PO copy',
            Path(settings.MEDIA_ROOT) / 'vendorpayment',
            Path(settings.BASE_DIR) / 'uploads' / 'po_form' / 'PO copy',
            Path(settings.BASE_DIR) / 'uploads' / 'po_form',
            Path(settings.BASE_DIR) / 'uploads' / 'purchase_order' / 'PO copy',
            Path(settings.BASE_DIR) / 'uploads' / 'purchase_order',
            Path(settings.BASE_DIR) / 'uploads' / 'vendorpayment' / 'PO copy',
            Path(settings.BASE_DIR) / 'uploads' / 'vendorpayment',
        ],
        'pan': [
            Path(settings.MEDIA_ROOT) / 'vendor_creation',
            LEGACY_UPLOAD_BASE / 'vendor_creation',
            Path(settings.BASE_DIR) / 'uploads' / 'vendor_creation',
        ],
        'bank': [
            Path(settings.MEDIA_ROOT) / 'vendor_creation',
            LEGACY_UPLOAD_BASE / 'vendor_creation',
            Path(settings.BASE_DIR) / 'uploads' / 'vendor_creation',
        ],
    }

    return [root / safe_name for root in folder_map.get(file_kind, [])]


def list_rows(tab, search='', from_date='', to_date='', page=1, length=10):
    vpd_cols = table_columns('vendor_payment_details')
    vc_cols = table_columns('vendor_creation')
    if not vpd_cols or 'bill_no' not in vpd_cols:
        return [], 0

    bill_date_col = pick_col(vpd_cols, 'bill_date', 'vendor_bill_created_date', 'created')
    invoice_date_col = pick_col(vpd_cols, 'vendor_inv_attach_approval_date', 'invoice_date')
    invoice_no_col = pick_col(vpd_cols, 'veninvverifyid', 'invoice_no')
    vendor_id_col = pick_col(vpd_cols, 'vendor_id')
    dc_col = pick_col(vpd_cols, 'dc_num', 'dc_number')
    amount_col = pick_col(vpd_cols, 'amount', 'total_amount')
    additional_charges_col = pick_col(vpd_cols, 'additionalcharges', 'additional_charges')
    advance_col = pick_col(vpd_cols, 'advancepayment')
    payable_col = pick_col(vpd_cols, 'acctotalpaybleamount')
    vendor_by_col = pick_col(vpd_cols, 'vendor_bill_approval', 'bill_approved')
    vendor_at_col = pick_col(vpd_cols, 'vendor_bill_approval_date')
    finance_by_col = pick_col(vpd_cols, 'finance_approved_by')
    finance_at_col = pick_col(vpd_cols, 'finance_approved_date')
    finance_status_col = pick_col(vpd_cols, 'finance_approval', 'finance_approval_sts')
    finance_remark_col = pick_col(vpd_cols, 'finance_remark')
    finance_reject_col = pick_col(vpd_cols, 'finance_reject_reason')
    tds_col = pick_col(vpd_cols, 'acctdsvalue')
    others_col = pick_col(vpd_cols, 'accotherdeduction')
    accstatus_col = pick_col(vpd_cols, 'accstatus')
    mgmt_status_col = pick_col(vpd_cols, 'managment_team_approval_sts')
    mgmt_reason_col = pick_col(vpd_cols, 'managment_team_reject_reason')
    mgmt_by_col = pick_col(vpd_cols, 'managment_team_approvedby')
    mgmt_at_col = pick_col(vpd_cols, 'managment_team_approvaldate')
    order_col = pick_col(vpd_cols, 'unique_id', 'id', 'created')

    join_vendor = vendor_join('vb', vpd_cols, vc_cols)
    vendor_name_expr = vendor_name_sql('vb', vpd_cols, vc_cols)
    vendor_address_expr = vendor_field_sql(vc_cols, 'address')
    vendor_phone_expr = vendor_field_sql(vc_cols, 'contact_no')

    where = ['1=1']
    params = []
    if 'is_delete' in vpd_cols:
        where.append('COALESCE(vb.is_delete, 0) = 0')
    if accstatus_col:
        where.append(f'COALESCE(vb.{accstatus_col}, 0) = 1')
    if finance_status_col:
        where.append(
            f"COALESCE(vb.{finance_status_col}, 0) = 0"
            if tab == 'pending'
            else f"COALESCE(vb.{finance_status_col}, 0) IN (1, 2)"
        )

    bill_date_expr = parsed_date_expr(f'vb.{bill_date_col}') if bill_date_col else None
    if from_date and bill_date_expr:
        where.append(f'{bill_date_expr} >= %s')
        params.append(from_date)
    if to_date and bill_date_expr:
        where.append(f'{bill_date_expr} <= %s')
        params.append(to_date)

    having_parts = []
    if search:
        like = f'%{search}%'
        having_parts.append('CAST(bill_no AS CHAR) LIKE %s')
        params.append(like)
        if invoice_no_col:
            having_parts.append('invoice_no LIKE %s')
            params.append(like)
        having_parts.append('vendor_name LIKE %s')
        params.append(like)

    amount_expr = "0"
    if amount_col and additional_charges_col:
        amount_expr = f"ROUND(IFNULL(SUM(vb.{amount_col}), 0) + MAX(IFNULL(vb.{additional_charges_col}, 0)), 2)"
    elif amount_col:
        amount_expr = f"ROUND(IFNULL(SUM(vb.{amount_col}), 0), 2)"
    elif additional_charges_col:
        amount_expr = f"ROUND(MAX(IFNULL(vb.{additional_charges_col}, 0)), 2)"

    base_sql = f"""
        SELECT
            vb.bill_no,
            {sql_max_date(f'vb.{bill_date_col}', 'bill_date') if bill_date_col else "'' AS bill_date"},
            {sql_max_date(f'vb.{invoice_date_col}', 'invoice_date') if invoice_date_col else "'' AS invoice_date"},
            {max_expr(f'vb.{invoice_no_col}', 'invoice_no', cast_text=True) if invoice_no_col else "'' AS invoice_no"},
            {f"MAX(CAST({vendor_name_expr} AS CHAR)) AS vendor_name" if vendor_name_expr != "''" else "'' AS vendor_name"},
            {f"MAX(CAST({vendor_address_expr} AS CHAR)) AS vendor_address" if vendor_address_expr != "''" else "'' AS vendor_address"},
            {f"MAX(CAST({vendor_phone_expr} AS CHAR)) AS vendor_phone" if vendor_phone_expr != "''" else "'' AS vendor_phone"},
            {f"COUNT(DISTINCT vb.{dc_col}) AS dc_count" if dc_col else "0 AS dc_count"},
            {amount_expr} AS amount,
            {f"ROUND(IFNULL(MAX(vb.{advance_col}), 0), 2) AS advance_amount" if advance_col else "0 AS advance_amount"},
            {f"ROUND(IFNULL(MAX(vb.{payable_col}), 0), 2) AS total_payable" if payable_col else "0 AS total_payable"},
            {staff_name_expr(f'vb.{vendor_by_col}', 'vendor_approved_by') if vendor_by_col else "'' AS vendor_approved_by"},
            {sql_max_date(f'vb.{vendor_at_col}', 'vendor_approved_date') if vendor_at_col else "'' AS vendor_approved_date"},
            {staff_name_expr(f'vb.{finance_by_col}', 'finance_approved_by') if finance_by_col else "'' AS finance_approved_by"},
            {sql_max_date(f'vb.{finance_at_col}', 'finance_approved_date') if finance_at_col else "'' AS finance_approved_date"},
            {staff_name_expr(f'vb.{mgmt_by_col}', 'management_approved_by') if mgmt_by_col else "'' AS management_approved_by"},
            {sql_max_date(f'vb.{mgmt_at_col}', 'management_approved_date') if mgmt_at_col else "'' AS management_approved_date"},
            {max_expr(f'vb.{mgmt_status_col}', 'management_status_code') if mgmt_status_col else "0 AS management_status_code"},
            {max_expr(f'vb.{mgmt_reason_col}', 'management_reject_reason', cast_text=True) if mgmt_reason_col else "'' AS management_reject_reason"},
            {max_expr(f'vb.{finance_remark_col}', 'remarks', cast_text=True) if finance_remark_col else "'' AS remarks"},
            {max_expr(f'vb.{finance_reject_col}', 'finance_reject_reason', cast_text=True) if finance_reject_col else "'' AS finance_reject_reason"},
            {f"ROUND(IFNULL(MAX(vb.{tds_col}), 0), 2) AS tds_deduction" if tds_col else "0 AS tds_deduction"},
            {f"ROUND(IFNULL(MAX(vb.{others_col}), 0), 2) AS others_deduction" if others_col else "0 AS others_deduction"},
            {max_expr(f'vb.{finance_status_col}', 'finance_status_code') if finance_status_col else "0 AS finance_status_code"}
            {f", MAX(CAST(vb.{vendor_id_col} AS CHAR)) AS vendor_id" if vendor_id_col else ", '' AS vendor_id"}
        FROM vendor_payment_details vb
        {join_vendor}
        WHERE {' AND '.join(where)}
        GROUP BY vb.bill_no{', vb.' + vendor_id_col if vendor_id_col else ''}
        {f"HAVING {' OR '.join(having_parts)}" if having_parts else ''}
        ORDER BY {f'MAX(vb.{order_col}) DESC' if order_col else 'vb.bill_no DESC'}
    """

    total = safe_int((db_fetchone(f"SELECT COUNT(*) AS total FROM ({base_sql}) finance_rows", params) or {}).get('total'))
    rows = db_fetch(f"{base_sql} LIMIT %s OFFSET %s", params + [length, max(page - 1, 0) * length])
    return rows, total
class AccountsBillApprovalListView(APIView):
    def get(self, request):
        tab = safe_str(request.query_params.get('tab', 'pending')).lower()
        search = safe_str(request.query_params.get('search')).strip()
        from_date = safe_str(request.query_params.get('from_date')).strip()
        to_date = safe_str(request.query_params.get('to_date')).strip()
        page = safe_int(request.query_params.get('page')) or 1
        length = safe_int(request.query_params.get('length')) or 10

        rows, total = list_rows(tab, search, from_date, to_date, page, length)
        data = []
        for idx, row in enumerate(rows, start=((page - 1) * length) + 1):
            data.append({
                'id': f"{safe_str(row.get('bill_no'))}-{safe_str(row.get('vendor_id')) or idx}",
                'bill_no': safe_str(row.get('bill_no')),
                'bill_date': safe_str(row.get('bill_date')),
                'invoice_no': safe_str(row.get('invoice_no')),
                'invoice_date': safe_str(row.get('invoice_date')),
                'vendor_name': safe_str(row.get('vendor_name')),
                'vendor_address': safe_str(row.get('vendor_address')),
                'vendor_phone': safe_str(row.get('vendor_phone')),
                'dc_count': safe_int(row.get('dc_count')),
                'amount': safe_float(row.get('amount')),
                'advance_amount': safe_float(row.get('advance_amount')),
                'total_payable': safe_float(row.get('total_payable')),
                'vendor_approved_by': safe_str(row.get('vendor_approved_by')),
                'vendor_approved_date': safe_str(row.get('vendor_approved_date')),
                'finance_approved_by': safe_str(row.get('finance_approved_by')),
                'finance_approved_date': safe_str(row.get('finance_approved_date')),
                'management_approved_by': safe_str(row.get('management_approved_by')),
                'management_approved_date': safe_str(row.get('management_approved_date')),
                'management_status': status_text(row.get('management_status_code')),
                'management_reject_reason': safe_str(row.get('management_reject_reason')),
                'remarks': safe_str(row.get('remarks')),
                'finance_reject_reason': safe_str(row.get('finance_reject_reason')),
                'tds_deduction': safe_float(row.get('tds_deduction')),
                'others_deduction': safe_float(row.get('others_deduction')),
                'status': status_text(row.get('finance_status_code')),
            })

        return Response({'data': data, 'total': total, 'page': page, 'pages': max(1, -(-total // length))})


class AccountsBillApprovalFileView(APIView):
    def get(self, request, file_kind, filename):
        if file_kind not in {'invoice', 'po', 'pan', 'bank'}:
            raise Http404('Invalid file type')

        safe_name = Path(safe_str(filename)).name
        if not safe_name:
            raise Http404('File not found')

        for candidate in bill_file_candidates(file_kind, safe_name):
            try:
                if candidate.exists() and candidate.is_file():
                    return FileResponse(open(candidate, 'rb'), as_attachment=False, filename=safe_name)
            except Exception:
                continue

        raise Http404('File not found')


class AccountsBillApprovalDetailView(APIView):
    def get(self, request):
        bill_no = safe_str(request.query_params.get('bill_no')).strip()
        if not bill_no:
            return Response({'error': 'bill_no required'}, status=status.HTTP_400_BAD_REQUEST)

        vpd_cols = table_columns('vendor_payment_details')
        vpm_cols = table_columns('vendor_payment_details_main')
        vc_cols = table_columns('vendor_creation')
        vv_cols = table_columns('view_outsource_vendor_verified_invoice')

        source_table = 'vendor_payment_details_main' if vpm_cols and 'bill_no' in vpm_cols else 'vendor_payment_details'
        source_cols = vpm_cols if source_table == 'vendor_payment_details_main' else vpd_cols
        alias = 'src'
        join_vendor = vendor_join(alias, source_cols, vc_cols)

        company = pick_col(vc_cols, 'company_name', 'vendor_name', 'name')
        contact_name = pick_col(vc_cols, 'name')
        address = pick_col(vc_cols, 'address')
        gst = pick_col(vc_cols, 'gst_no')
        pan = pick_col(vc_cols, 'pan_no')
        mail = pick_col(vc_cols, 'mail_id')
        phone = pick_col(vc_cols, 'contact_no')
        bank = pick_col(vc_cols, 'bank_name')
        branch = pick_col(vc_cols, 'branch_name')
        account = pick_col(vc_cols, 'account_no')
        ifsc = pick_col(vc_cols, 'ifsc_code')
        holder = pick_col(vc_cols, 'acc_holder_name', 'account_holder_name')
        pan_copy = pick_col(vc_cols, 'pan_attach_file_name', 'pan_copy')
        bank_proof = pick_col(vc_cols, 'bank_proof')

        bill_date_col = pick_col(source_cols, 'bill_date', 'vendor_bill_created_date', 'created')
        invoice_no_col = pick_col(source_cols, 'user_vendor_invoice_id', 'veninvverifyid', 'invoice_no')
        invoice_date_col = pick_col(source_cols, 'vendor_inv_attach_approval_date', 'invoice_date')
        invoice_attach_col = pick_col(source_cols, 'inv_verfiy_attach', 'vendor_inv_attach')
        po_attach_col = pick_col(source_cols, 'po_ven_filename')
        tds_col = pick_col(source_cols, 'acctdsvalue')
        others_col = pick_col(source_cols, 'accotherdeduction')
        advance_col = pick_col(source_cols, 'advancepayment')
        payable_col = pick_col(source_cols, 'acctotalpaybleamount')
        additional_charges_col = pick_col(source_cols, 'additionalcharges', 'additional_charges')
        remark_col = pick_col(source_cols, 'finance_remark')
        reject_col = pick_col(source_cols, 'finance_reject_reason')

        detail_sql = f"""
            SELECT
                {f"CAST(v.{company} AS CHAR) AS vendor_name" if (join_vendor and company) else "'' AS vendor_name"},
                {f"CAST(v.{contact_name} AS CHAR) AS vendor_contact_name" if (join_vendor and contact_name) else "'' AS vendor_contact_name"},
                {f"CAST(v.{address} AS CHAR) AS vendor_address" if (join_vendor and address) else "'' AS vendor_address"},
                {f"CAST(v.{gst} AS CHAR) AS vendor_gst" if (join_vendor and gst) else "'' AS vendor_gst"},
                {f"CAST(v.{pan} AS CHAR) AS vendor_pan" if (join_vendor and pan) else "'' AS vendor_pan"},
                {f"CAST(v.{mail} AS CHAR) AS vendor_email" if (join_vendor and mail) else "'' AS vendor_email"},
                {f"CAST(v.{phone} AS CHAR) AS vendor_phone" if (join_vendor and phone) else "'' AS vendor_phone"},
                CAST({alias}.bill_no AS CHAR) AS vendor_bill_no,
                {sql_date(f'{alias}.{bill_date_col}', 'vendor_bill_date') if bill_date_col else "'' AS vendor_bill_date"},
                {f"CAST({alias}.{invoice_no_col} AS CHAR) AS vendor_invoice_no" if invoice_no_col else "'' AS vendor_invoice_no"},
                {sql_date(f'{alias}.{invoice_date_col}', 'vendor_invoice_date') if invoice_date_col else "'' AS vendor_invoice_date"},
                {f"CAST({alias}.{invoice_attach_col} AS CHAR) AS invoice_attach_url" if invoice_attach_col else "'#' AS invoice_attach_url"},
                {f"CAST({alias}.{po_attach_col} AS CHAR) AS po_attach_url" if po_attach_col else "'#' AS po_attach_url"},
                {f"IFNULL({alias}.{tds_col}, 0) AS tds_deduction" if tds_col else "0 AS tds_deduction"},
                {f"IFNULL({alias}.{others_col}, 0) AS others_deduction" if others_col else "0 AS others_deduction"},
                {f"IFNULL({alias}.{advance_col}, 0) AS advance_amount" if advance_col else "0 AS advance_amount"},
                {f"IFNULL({alias}.{payable_col}, 0) AS total_payable" if payable_col else "0 AS total_payable"},
                {f"IFNULL({alias}.{additional_charges_col}, 0) AS additional_charges" if additional_charges_col else "0 AS additional_charges"},
                {f"CAST({alias}.{remark_col} AS CHAR) AS remarks" if remark_col else "'' AS remarks"},
                {f"CAST({alias}.{reject_col} AS CHAR) AS reject_reason" if reject_col else "'' AS reject_reason"},
                {f"CAST(v.{bank} AS CHAR) AS bank_name" if (join_vendor and bank) else "'' AS bank_name"},
                {f"CAST(v.{branch} AS CHAR) AS branch" if (join_vendor and branch) else "'' AS branch"},
                {f"CAST(v.{account} AS CHAR) AS account_no" if (join_vendor and account) else "'' AS account_no"},
                {f"CAST(v.{ifsc} AS CHAR) AS ifsc_code" if (join_vendor and ifsc) else "'' AS ifsc_code"},
                {f"CAST(v.{holder} AS CHAR) AS account_holder" if (join_vendor and holder) else "'' AS account_holder"},
                {f"CAST(v.{pan_copy} AS CHAR) AS pan_copy_url" if (join_vendor and pan_copy) else "'#' AS pan_copy_url"},
                {f"CAST(v.{bank_proof} AS CHAR) AS bank_proof_url" if (join_vendor and bank_proof) else "'#' AS bank_proof_url"}
            FROM {source_table} {alias}
            {join_vendor}
            WHERE {alias}.bill_no = %s
            {'AND COALESCE(' + alias + '.is_delete, 0) = 0' if 'is_delete' in source_cols else ''}
            LIMIT 1
        """
        bill = db_fetchone(detail_sql, [bill_no])
        if not bill:
            return Response({'error': 'Bill not found'}, status=status.HTTP_404_NOT_FOUND)

        dc_col = pick_col(vpd_cols, 'dc_num', 'dc_number')
        dc_date_col = pick_col(vpd_cols, 'dc_date')
        line_invoice_no_col = pick_col(vpd_cols, 'invoice_no')
        line_invoice_date_col = pick_col(vpd_cols, 'invoice_date', 'vendor_inv_attach_approval_date')
        po_col = pick_col(vpd_cols, 'po_num')
        po_date_col = pick_col(vpd_cols, 'po_date')
        qty_col = pick_col(vpd_cols, 'invoice_qty')
        rate_col = pick_col(vpd_cols, 'rate', 'unit_price', 'vendor_bulk_rate')
        gst_col = pick_col(vpd_cols, 'gst', 'vendor_bulk_gst')
        line_amount_col = pick_col(vpd_cols, 'amount', 'total_amount', 'bulk_total_amount')
        vendor_id_col = pick_col(vpd_cols, 'vendor_id')
        order_col = pick_col(vpd_cols, 'unique_id', 'id')

        join_verified = ''
        if dc_col:
            join_verified = f"LEFT JOIN view_outsource_vendor_verified_invoice vv ON vv.dc_number = vpd.{dc_col}"
            if vendor_id_col and 'engineer_name' in vv_cols:
                join_verified += f" AND vv.engineer_name = vpd.{vendor_id_col}"

        invoice_date_expr = f"COALESCE(vv.invoice_date, vpd.{line_invoice_date_col})" if line_invoice_date_col and 'invoice_date' in vv_cols else (f"vpd.{line_invoice_date_col}" if line_invoice_date_col else 'vv.invoice_date')
        po_date_expr = f"COALESCE(vv.po_date, vpd.{po_date_col})" if po_date_col and 'po_date' in vv_cols else (f"vpd.{po_date_col}" if po_date_col else 'vv.po_date')

        dc_sql = f"""
            SELECT
                {f"CAST(vpd.{dc_col} AS CHAR) AS dc_no" if dc_col else "'' AS dc_no"},
                {sql_date(f'vpd.{dc_date_col}', 'dc_date') if dc_date_col else "'' AS dc_date"},
                {f"CAST(vpd.{line_invoice_no_col} AS CHAR) AS invoice_no" if line_invoice_no_col else "'' AS invoice_no"},
                {sql_date(invoice_date_expr, 'invoice_date')},
                {f"CAST(vpd.{po_col} AS CHAR) AS po_no" if po_col else "'' AS po_no"},
                {sql_date(po_date_expr, 'po_date')},
                COALESCE(get_address(vv.consignee_unique_id), '') AS consignee_address,
                {f"IFNULL(vpd.{qty_col}, 0) AS invoice_qty" if qty_col else "0 AS invoice_qty"},
                {f"IFNULL(vpd.{rate_col}, 0) AS unit_price" if rate_col else "0 AS unit_price"},
                {f"IFNULL(vpd.{gst_col}, 0) AS gst" if gst_col else "0 AS gst"},
                {f"IFNULL(vpd.{line_amount_col}, 0) AS total_amount" if line_amount_col else "0 AS total_amount"}
            FROM vendor_payment_details vpd
            {join_verified}
            WHERE vpd.bill_no = %s
            {'AND COALESCE(vpd.is_delete, 0) = 0' if 'is_delete' in vpd_cols else ''}
            ORDER BY {f'vpd.{order_col}' if order_col else 'vpd.bill_no'}
        """
        dc_rows = db_fetch(dc_sql, [bill_no])
        dc_items = []
        total_amount = 0.0
        for idx, row in enumerate(dc_rows, start=1):
            qty = safe_float(row.get('invoice_qty'))
            unit_price = safe_float(row.get('unit_price'))
            gst_value = safe_float(row.get('gst'))
            basic_amount = qty * unit_price
            gst_amount = (basic_amount * gst_value) / 100 if basic_amount else 0.0
            line_total = safe_float(row.get('total_amount')) or (basic_amount + gst_amount)
            total_amount += line_total
            dc_items.append({
                's_no': idx,
                'dc_no': safe_str(row.get('dc_no')),
                'dc_date': safe_str(row.get('dc_date')),
                'invoice_no': safe_str(row.get('invoice_no')),
                'invoice_date': safe_str(row.get('invoice_date')),
                'po_no': safe_str(row.get('po_no')),
                'po_date': safe_str(row.get('po_date')),
                'consignee_address': safe_str(row.get('consignee_address')),
                'invoice_qty': qty,
                'unit_price': unit_price,
                'basic_amount': basic_amount,
                'gst': f"{gst_value:g} %",
                'gst_amount': gst_amount,
                'total_amount': line_total,
            })
        additional_charges = safe_float(bill.get('additional_charges'))
        grand_total_amount = total_amount + additional_charges

        bill_created_by_col = pick_col(vpd_cols, 'vendor_bill_created_by')
        bill_created_at_col = pick_col(vpd_cols, 'vendor_bill_created_date')
        bill_approved_by_col = pick_col(vpd_cols, 'vendor_bill_approval')
        bill_approved_at_col = pick_col(vpd_cols, 'vendor_bill_approval_date')
        bill_status_col = pick_col(vpd_cols, 'vendor_bill_app_status')
        account_entry_by_col = pick_col(vpd_cols, 'vendor_account_approved_by')
        account_entry_at_col = pick_col(vpd_cols, 'vendor_account_approval_date')
        account_entry_status_col = pick_col(vpd_cols, 'acc_ent_sts')
        accounts_approval_by_col = pick_col(vpd_cols, 'finance_approved_by')
        accounts_approval_at_col = pick_col(vpd_cols, 'finance_approved_date')
        accounts_approval_status_col = pick_col(vpd_cols, 'finance_approval', 'finance_approval_sts')
        management_by_col = pick_col(vpd_cols, 'managment_team_approvedby')
        management_at_col = pick_col(vpd_cols, 'managment_team_approvaldate')
        management_status_col = pick_col(vpd_cols, 'managment_team_approval_sts')
        payment_ref_col = pick_col(vpd_cols, 'transaction_id')
        payment_date_col = pick_col(vpd_cols, 'transaction_date')
        payment_amount_col = pick_col(vpd_cols, 'acctotalpaybleamount')
        payment_status_col = pick_col(vpd_cols, 'accounts_approval')
        account_bill_id_col = pick_col(vpd_cols, 'accountbillid')
        account_remarks_col = pick_col(vpd_cols, 'accdetuctionremarks')

        approval_sql = f"""
            SELECT
                {staff_name_expr(f'vpd.{bill_created_by_col}', 'bill_created_by') if bill_created_by_col else "'' AS bill_created_by"},
                {sql_max_date(f'vpd.{bill_created_at_col}', 'bill_created_at', '%d-%m-%Y %H:%i:%s') if bill_created_at_col else "'' AS bill_created_at"},
                {staff_name_expr(f'vpd.{bill_approved_by_col}', 'operation_by') if bill_approved_by_col else "'' AS operation_by"},
                {sql_max_date(f'vpd.{bill_approved_at_col}', 'operation_at', '%d-%m-%Y %H:%i:%s') if bill_approved_at_col else "'' AS operation_at"},
                {max_expr(f'vpd.{bill_status_col}', 'operation_status') if bill_status_col else "0 AS operation_status"},
                {staff_name_expr(f'vpd.{account_entry_by_col}', 'account_entry_by') if account_entry_by_col else "'' AS account_entry_by"},
                {sql_max_date(f'vpd.{account_entry_at_col}', 'account_entry_at', '%d-%m-%Y %H:%i:%s') if account_entry_at_col else "'' AS account_entry_at"},
                {max_expr(f'vpd.{account_entry_status_col}', 'account_entry_status') if account_entry_status_col else "0 AS account_entry_status"},
                {staff_name_expr(f'vpd.{accounts_approval_by_col}', 'accounts_approval_by') if accounts_approval_by_col else "'' AS accounts_approval_by"},
                {sql_max_date(f'vpd.{accounts_approval_at_col}', 'accounts_approval_at', '%d-%m-%Y %H:%i:%s') if accounts_approval_at_col else "'' AS accounts_approval_at"},
                {max_expr(f'vpd.{accounts_approval_status_col}', 'accounts_approval_status') if accounts_approval_status_col else "0 AS accounts_approval_status"},
                {staff_name_expr(f'vpd.{management_by_col}', 'management_by') if management_by_col else "'' AS management_by"},
                {sql_max_date(f'vpd.{management_at_col}', 'management_at', '%d-%m-%Y %H:%i:%s') if management_at_col else "'' AS management_at"},
                {max_expr(f'vpd.{management_status_col}', 'management_status') if management_status_col else "0 AS management_status"},
                {max_expr(f'vpd.{payment_ref_col}', 'payment_ref', cast_text=True) if payment_ref_col else "'' AS payment_ref"},
                {sql_max_date(f'vpd.{payment_date_col}', 'payment_date') if payment_date_col else "'' AS payment_date"},
                {max_expr(f'vpd.{payment_amount_col}', 'payment_amount') if payment_amount_col else "0 AS payment_amount"},
                {max_expr(f'vpd.{payment_status_col}', 'payment_status') if payment_status_col else "0 AS payment_status"},
                {max_expr(f'vpd.{account_bill_id_col}', 'account_bill_id', cast_text=True) if account_bill_id_col else "'' AS account_bill_id"},
                {max_expr(f'vpd.{account_remarks_col}', 'account_remarks', cast_text=True) if account_remarks_col else "'' AS account_remarks"}
            FROM vendor_payment_details vpd
            WHERE vpd.bill_no = %s AND COALESCE(vpd.is_delete, 0) = 0
        """
        approval = db_fetchone(approval_sql, [bill_no]) or {}
        return Response({
            'id': safe_str(bill.get('vendor_bill_no')),
            'vendor_name': safe_str(bill.get('vendor_name')),
            'vendor_contact_name': safe_str(bill.get('vendor_contact_name')),
            'vendor_address': safe_str(bill.get('vendor_address')),
            'vendor_gst': safe_str(bill.get('vendor_gst')),
            'vendor_pan': safe_str(bill.get('vendor_pan')),
            'vendor_email': safe_str(bill.get('vendor_email')),
            'vendor_phone': safe_str(bill.get('vendor_phone')),
            'vendor_bill_no': safe_str(bill.get('vendor_bill_no')),
            'vendor_bill_date': safe_str(bill.get('vendor_bill_date')),
            'vendor_invoice_no': safe_str(bill.get('vendor_invoice_no')),
            'vendor_invoice_date': safe_str(bill.get('vendor_invoice_date')),
            'invoice_attach_url': bill_file_url(request, 'invoice', bill.get('invoice_attach_url')),
            'po_attach_url': bill_file_url(request, 'po', bill.get('po_attach_url')),
            'bank_name': safe_str(bill.get('bank_name')),
            'branch': safe_str(bill.get('branch')),
            'account_no': safe_str(bill.get('account_no')),
            'ifsc_code': safe_str(bill.get('ifsc_code')),
            'account_holder': safe_str(bill.get('account_holder')),
            'pan_copy_url': bill_file_url(request, 'pan', bill.get('pan_copy_url')),
            'bank_proof_url': bill_file_url(request, 'bank', bill.get('bank_proof_url')),
            'remarks': safe_str(bill.get('remarks')),
            'reject_reason': safe_str(bill.get('reject_reason')),
            'account_bill_id': safe_str(approval.get('account_bill_id')),
            'account_remarks': safe_str(approval.get('account_remarks')),
            'bill_created_by': safe_str(approval.get('bill_created_by')),
            'bill_created_at': safe_str(approval.get('bill_created_at')),
            'bill_approved_by': safe_str(approval.get('operation_by')),
            'bill_approved_at': safe_str(approval.get('operation_at')),
            'account_bill_created_by': safe_str(approval.get('account_entry_by')),
            'account_bill_created_at': safe_str(approval.get('account_entry_at')),
            'accounts_team_approved_by': safe_str(approval.get('accounts_approval_by')),
            'accounts_team_approved_at': safe_str(approval.get('accounts_approval_at')),
            'dc_items': dc_items,
            'total_amount': total_amount,
            'additional_charges': additional_charges,
            'grand_total_amount': grand_total_amount,
            'tds_deduction': safe_float(bill.get('tds_deduction')),
            'others_deduction': safe_float(bill.get('others_deduction')),
            'advance_amount': safe_float(bill.get('advance_amount')),
            'total_payable': safe_float(bill.get('total_payable')) or total_amount,
            'approvals': [{
                's_no': 1,
                'bill_created_by': safe_str(approval.get('bill_created_by')),
                'bill_created_at': safe_str(approval.get('bill_created_at')),
                'bill_created_status': 'Bill Created',
                'operation_by': safe_str(approval.get('operation_by')),
                'operation_at': safe_str(approval.get('operation_at')),
                'operation_status': status_text(approval.get('operation_status')),
                'account_entry_by': safe_str(approval.get('account_entry_by')),
                'account_entry_at': safe_str(approval.get('account_entry_at')),
                'account_entry_status': status_text(approval.get('account_entry_status')),
                'accounts_approval_by': safe_str(approval.get('accounts_approval_by')),
                'accounts_approval_at': safe_str(approval.get('accounts_approval_at')),
                'accounts_approval_status': status_text(approval.get('accounts_approval_status')),
                'management_by': safe_str(approval.get('management_by')),
                'management_at': safe_str(approval.get('management_at')),
                'management_status': status_text(approval.get('management_status')),
                'payment_ref': safe_str(approval.get('payment_ref')),
                'payment_date': safe_str(approval.get('payment_date')),
                'payment_amount': f"{safe_float(approval.get('payment_amount')):,.2f}",
                'payment_status': payment_status_text(approval.get('payment_status')),
            }],
        })

class AccountsBillApprovalUpdateRemarkView(APIView):
    def post(self, request):
        bill_no = safe_str(request.data.get('bill_no')).strip()
        remark = safe_str(request.data.get('remark'))
        if not bill_no:
            return Response({'error': 'bill_no required'}, status=status.HTTP_400_BAD_REQUEST)

        vpd_cols = table_columns('vendor_payment_details')
        vpm_cols = table_columns('vendor_payment_details_main')
        remark_col = pick_col(vpd_cols, 'finance_remark')
        if not remark_col:
            return Response({'status': 0, 'msg': 'Remark field not available'}, status=status.HTTP_400_BAD_REQUEST)

        with connection.cursor() as cursor:
            cursor.execute(
                f"UPDATE vendor_payment_details SET {remark_col} = %s WHERE bill_no = %s AND COALESCE(is_delete, 0) = 0",
                [remark, bill_no],
            )
            main_remark_col = pick_col(vpm_cols, 'finance_remark')
            if main_remark_col:
                cursor.execute(
                    f"UPDATE vendor_payment_details_main SET {main_remark_col} = %s WHERE bill_no = %s AND COALESCE(is_delete, 0) = 0",
                    [remark, bill_no],
                )

        return Response({'status': 1, 'msg': 'Remark updated'})


class AccountsBillApprovalApproveView(APIView):
    def post(self, request):
        bill_no = safe_str(request.data.get('bill_no')).strip()
        action_status = safe_str(request.data.get('status', 'Approved')).strip() or 'Approved'
        remark = safe_str(request.data.get('remark')).strip()
        reject_reason = safe_str(request.data.get('reject_reason', '')).strip()
        user_id = get_user_id(request)
        if not bill_no:
            return Response({'error': 'bill_no required'}, status=status.HTTP_400_BAD_REQUEST)

        status_value = {'Approved': 1, 'Rejected': 2, 'Pending': 0}.get(action_status, 0)
        vpd_cols = table_columns('vendor_payment_details')
        vpm_cols = table_columns('vendor_payment_details_main')

        def update_parts(columns):
            status_col = pick_col(columns, 'finance_approval', 'finance_approval_sts')
            if not status_col:
                return None
            parts = [f"{status_col} = %s"]
            values = [status_value]
            remark_col = pick_col(columns, 'finance_remark')
            reject_col = pick_col(columns, 'finance_reject_reason')
            by_col = pick_col(columns, 'finance_approved_by')
            at_col = pick_col(columns, 'finance_approved_date')
            if remark_col:
                parts.append(f"{remark_col} = %s")
                values.append(remark)
            if reject_col:
                parts.append(f"{reject_col} = %s")
                values.append(reject_reason if status_value == 2 else '')
            if by_col:
                parts.append(f"{by_col} = %s")
                values.append(user_id if status_value in (1, 2) else '')
            if at_col:
                parts.append(f"{at_col} = NOW()" if status_value in (1, 2) else f"{at_col} = NULL")
            return parts, values

        main_update = update_parts(vpd_cols)
        if not main_update:
            return Response({'status': 0, 'msg': 'Finance status field not available'}, status=status.HTTP_400_BAD_REQUEST)

        with connection.cursor() as cursor:
            parts, values = main_update
            cursor.execute(
                f"UPDATE vendor_payment_details SET {', '.join(parts)} WHERE bill_no = %s AND COALESCE(is_delete, 0) = 0",
                values + [bill_no],
            )
            summary_update = update_parts(vpm_cols)
            if summary_update:
                parts, values = summary_update
                cursor.execute(
                    f"UPDATE vendor_payment_details_main SET {', '.join(parts)} WHERE bill_no = %s AND COALESCE(is_delete, 0) = 0",
                    values + [bill_no],
                )

            # When rejected (status_value == 2), reset invoices back to Stage 1 (Vendor Bill Creation)
            if status_value == 2:
                try:
                    cursor.execute(
                        """
                        UPDATE invoice_verfication_table iv
                        JOIN vendor_payment_details vpd ON vpd.dc_num = iv.dc_number
                        SET iv.vendor_payment_allocated = 0,
                            iv.vendor_bill_no = '',
                            iv.managment_team_allocated = 0,
                            iv.vendor_finance_approval = 0
                        WHERE vpd.bill_no = %s AND COALESCE(vpd.is_delete, 0) = 0
                        """,
                        [bill_no],
                    )
                except Exception:
                    pass

        return Response({'status': 1, 'msg': action_status.lower()})


class AccountsBillApprovalExportView(APIView):
    def get(self, request):
        tab = safe_str(request.query_params.get('tab', 'complete')).lower()
        search = safe_str(request.query_params.get('search')).strip()
        from_date = safe_str(request.query_params.get('from_date')).strip()
        to_date = safe_str(request.query_params.get('to_date')).strip()
        rows, _ = list_rows(tab, search, from_date, to_date, 1, 100000)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Accounts Bill Approval'
        ws.merge_cells('A1:O1')
        ws['A1'] = 'Accounts Team Bill Approval'
        ws['A1'].font = Font(bold=True, size=13)
        ws['A1'].alignment = Alignment(horizontal='center')

        headers = [
            'S.No', 'Bill No', 'Bill Date', 'Invoice No', 'Invoice Date',
            'Vendor Name', 'Vendor Address', 'Vendor Phone', 'DC Count', 'Bill Value',
            'Advance Amount', 'Total Payable', 'Finance Status', 'Management Status', 'Remarks',
        ]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col_idx, value=header).font = Font(bold=True)

        for idx, row in enumerate(rows, start=1):
            ws.append([
                idx,
                safe_str(row.get('bill_no')),
                safe_str(row.get('bill_date')),
                safe_str(row.get('invoice_no')),
                safe_str(row.get('invoice_date')),
                safe_str(row.get('vendor_name')),
                safe_str(row.get('vendor_address')),
                safe_str(row.get('vendor_phone')),
                safe_int(row.get('dc_count')),
                safe_float(row.get('amount')),
                safe_float(row.get('advance_amount')),
                safe_float(row.get('total_payable')),
                status_text(row.get('finance_status_code')),
                status_text(row.get('management_status_code')),
                safe_str(row.get('remarks')),
            ])

        for col in ws.columns:
            letter = col[0].column_letter
            ws.column_dimensions[letter].width = min(max(len(str(cell.value or '')) for cell in col) + 4, 40)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="AccountsBillApproval.xlsx"'
        return response

