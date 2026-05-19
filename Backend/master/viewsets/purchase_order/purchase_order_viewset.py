import csv
import os
import uuid
from collections import defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import TextIOWrapper
from zipfile import BadZipFile

from django.conf import settings
from django.db import connection, transaction
from django.db.models import Count, Q
from django.http import FileResponse, Http404
from django.utils.decorators import method_decorator
from django.views.decorators.clickjacking import xframe_options_exempt
from rest_framework import status
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from master.apps.account_sector.accountsectormodel import AccountSector
from master.apps.account_vertical.accountverticalmodel import AccountVertical
from master.apps.department.departmentmodel import DepartmentCreation
from master.apps.district.districtmodel import DistrictCreation
from master.apps.executive_creation.executivecreation_model import ExecutiveName
from master.apps.insurance_type.insurance_type_models import InsuranceType
from master.apps.item_creation.itemcreationmodel import ItemCreation, ItemCreationSub
from master.apps.purchase_order.purchaseordermodel import (
    PurchaseOrderAmc,
    PurchaseOrder,
    PurchaseOrderAssign,
    PurchaseOrderConsignee,
    PurchaseOrderProduct,
)
from master.apps.state.statemodel import StateCreation
from master.apps.user_type.usertypemodel import UserType
from master.serializers.purchase_order.purchaseorder_serializer import (
    PurchaseOrderAmcSerializer,
    PurchaseOrderConsigneeSerializer,
    PurchaseOrderProductSerializer,
    PurchaseOrderSerializer,
)
from master.tenant import apply_tenant_audit, request_company_id, tenant_queryset
from master.viewsets.Payment_Transaction_Form.payment_viewset import (
    _ensure_payment_notification_table,
    _get_request_user_id,
)


PO_TYPES = [
    {"value": "Product", "label": "Product"},
    {"value": "Service", "label": "Service"},
]

LD_TYPES = [
    {"value": "day", "label": "LD Per Day"},
    {"value": "week", "label": "LD Per Week"},
    {"value": "month", "label": "LD Per Month"},
]

WARRANTY_STARTS = [
    {"value": "delivery_date", "label": "Delivery Date"},
    {"value": "installation_date", "label": "Installation Date"},
    {"value": "po_date", "label": "PO Date"},
]

DATE_TYPES = [
    {"value": "po_date", "label": "From PO Date"},
    {"value": "fixed_date", "label": "Consignee Date"},
]


def _to_decimal(value):
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", ""))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _safe_int(value, default=0):
    try:
        return int(str(value or "").strip())
    except (TypeError, ValueError):
        return default


def _as_bool(value):
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def _normalize_po_num(value):
    return str(value or "").strip()


def _duplicate_po_exists(po_num, exclude_unique_id=""):
    normalized = _normalize_po_num(po_num)
    if not normalized:
        return False
    qs = PurchaseOrder.objects.filter(po_num__iexact=normalized, is_delete=0)
    if exclude_unique_id:
        qs = qs.exclude(unique_id=exclude_unique_id)
    return qs.exists()


def _normalize_ld_type(value):
    raw = str(value or "").strip()
    mapping = {
        "LD Per Day": "day",
        "LD Per Week": "week",
        "LD Per Month": "month",
        "day": "day",
        "week": "week",
        "month": "month",
    }
    return mapping.get(raw, raw[:10])


def _po_storage_roots(folder):
    base_dir = getattr(settings, "BASE_DIR", "")
    media_root = getattr(settings, "MEDIA_ROOT", os.path.join(base_dir, "media"))
    return [
        os.path.join(base_dir, "uploads", "purchase_order", folder),
        os.path.join(media_root, "purchase_order", folder),
    ]


def _po_upload_root(folder):
    target = _po_storage_roots(folder)[0]
    os.makedirs(target, exist_ok=True)
    return target


def _resolve_po_file_path(folder, filename):
    safe_name = os.path.basename(filename)
    for root in _po_storage_roots(folder):
        path = os.path.join(root, safe_name)
        if os.path.exists(path):
            return path
    return ""


def _save_upload(file_obj, folder, filename):
    upload_dir = _po_upload_root(folder)
    path = os.path.join(upload_dir, filename)
    with open(path, "wb+") as handle:
        for chunk in file_obj.chunks():
            handle.write(chunk)
    return filename


def _build_customer_text(po, data=None):
    data = data or {}
    parts = [
        data.get("department_display") or po.department,
        po.bill_address,
        data.get("state_name_display") or po.state_name,
        data.get("district_name_display") or po.district,
        f"({po.gst_value})" if po.gst_value else "",
    ]
    return "\n".join([part for part in parts if part])


def _refresh_po_summary(po):
    products = PurchaseOrderProduct.objects.filter(form_main_unique_id=po.unique_id, is_delete=0)
    consignees = PurchaseOrderConsignee.objects.filter(form_main_unique_id=po.unique_id, is_delete=0)
    assigns = PurchaseOrderAssign.objects.filter(form_main_unique_id=po.unique_id, is_delete=0)

    po.total_qty = str(sum(_safe_int(row.qty) for row in products))
    po.total_amount = str(sum((_to_decimal(row.total_value) for row in products), Decimal("0")))
    po.no_of_consignee = str(consignees.count())
    po.product_sts = 1 if products.exists() else 0
    po.consignee_sts = 1 if consignees.exists() else 0
    po.assign_sts = 1 if assigns.exists() else 0
    po.po_com_sts = 1 if products.exists() and consignees.exists() else 0
    po.save(
        update_fields=[
            "no_of_po",
            "total_qty",
            "total_amount",
            "no_of_consignee",
            "product_sts",
            "consignee_sts",
            "assign_sts",
            "po_com_sts",
            "updated_at",
        ]
    )


def _sync_consignee_team_member(consignee):
    team_member = str(consignee.assign_team_member or consignee.team_mem or "").strip()
    consignee_unique_id = str(consignee.unique_id or "").strip()
    form_main_unique_id = str(consignee.form_main_unique_id or "").strip()
    if not consignee_unique_id:
        return

    params = [team_member, consignee_unique_id]
    assign_sql = """
        UPDATE po_product_assign_details
        SET con_assign_team_member = %s
        WHERE con_unique_id = %s
          AND is_delete = 0
    """
    invoice_main_sql = """
        UPDATE invoice_creation_main
        SET team_mem = %s
        WHERE consignee_unique_id = %s
          AND is_delete = 0
    """
    invoice_main_payment_sql = """
        UPDATE invoice_creation_main_payment_data
        SET team_mem = %s
        WHERE consignee_unique_id = %s
          AND is_delete = 0
    """
    invoice_sql = """
        UPDATE invoice_creation
        SET team_mem = %s
        WHERE consignee_id = %s
          AND is_delete = 0
    """

    if form_main_unique_id:
        assign_sql += " AND form_main_unique_id = %s"
        invoice_main_sql += " AND form_main_unique_id = %s"
        invoice_main_payment_sql += " AND form_main_unique_id = %s"
        invoice_sql += " AND po_unique_id = %s"
        params.append(form_main_unique_id)

    with connection.cursor() as cur:
        cur.execute(assign_sql, params)
        cur.execute(invoice_main_sql, params)
        cur.execute(invoice_main_payment_sql, params)
        cur.execute(invoice_sql, params)


def _stores_user_type_for_request(request):
    company_id = request_company_id(request)
    rows = list(
        tenant_queryset(
            request,
            UserType.objects.filter(user_type__iexact="Stores", is_delete=0),
            include_global=True,
        ).order_by("sess_company_id", "s_no")
    )
    if not rows:
        return None
    for row in rows:
        if str(row.sess_company_id or "").strip() == company_id:
            return row
    return rows[0]


def _create_consignee_update_notification(request, row, po):
    stores_type = _stores_user_type_for_request(request)
    if not stores_type:
        return

    po_number = str(getattr(po, "po_num", "") or row.po_number or row.form_main_unique_id or "").strip()
    consignee_name = str(row.con_branch or row.con_branch_code or row.unique_id or "").strip()
    source_path = ""
    if po:
        source_path = f"/order/purchase-order/consignee-details/{po.unique_id}/{po_number}"

    _ensure_payment_notification_table()
    with connection.cursor() as cursor:
        cursor.execute(
            """
            INSERT INTO payment_transaction_notifications
                (unique_id, recipient_user_id, bill_no, notification_type, title, message, source_module, source_path, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            [
                uuid.uuid4().hex[:20],
                stores_type.unique_id,
                po_number,
                "consignee_updated",
                "Consignee Details Updated",
                f"Consignee details updated for PO {po_number}: {consignee_name}.",
                "purchase_order",
                source_path,
                _get_request_user_id(request),
            ],
        )


def _generate_batch_id():
    latest = (
        PurchaseOrderConsignee.objects
        .filter(batch_id__startswith="PO-BAT-", is_delete=0)
        .order_by("-batch_id")
        .values_list("batch_id", flat=True)
        .first()
    )
    prefix = datetime.now().strftime("PO-BAT-%y%m-")
    if latest and latest.startswith(prefix):
        try:
            seq = int(latest.split("-")[-1]) + 1
        except ValueError:
            seq = 1
    else:
        seq = 1
    return f"{prefix}{seq:04d}"


def _current_academic_year():
    now = datetime.now()
    if now.month >= 4:
        return f"{now.year}-{now.year + 1}"
    return f"{now.year - 1}-{now.year}"


def _parse_uploaded_rows(file_obj):
    extension = os.path.splitext(file_obj.name or "")[1].lower()
    rows = []

    if extension == ".csv":
        file_obj.seek(0)
        wrapper = TextIOWrapper(file_obj.file, encoding="utf-8-sig", newline="")
        reader = csv.reader(wrapper)
        for row in reader:
            values = ["" if cell is None else str(cell).strip() for cell in row]
            if any(values):
                rows.append(values)
        return rows

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to import Excel files.") from exc

    file_obj.seek(0)
    workbook = load_workbook(file_obj, data_only=True)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        values = ["" if cell is None else str(cell).strip() for cell in row]
        if any(values):
            rows.append(values)
    return rows


def _map_district_unique_id(value):
    if not value:
        return ""
    district = (
        DistrictCreation.objects
        .filter(Q(unique_id=value) | Q(district_name__iexact=value), is_delete=0)
        .order_by("-id")
        .first()
    )
    return district.unique_id if district else value


def _map_state_unique_id(value):
    if not value:
        return ""
    state = (
        StateCreation.objects
        .filter(Q(unique_id=value) | Q(state_name__iexact=value), is_delete=0)
        .order_by("-id")
        .first()
    )
    return state.unique_id if state else value


def _po_row(po, index, request):
    serializer = PurchaseOrderSerializer(po, context={"request": request, "include_children": False})
    data = serializer.data
    data["s_no"] = index
    data["customer_details"] = _build_customer_text(po, data)
    data["customer_name"] = data.get("department_display") or po.department
    data["executive_label"] = data.get("executive_name_display") or po.executive_name
    data["ld_required_display"] = "Yes" if _as_bool(po.ld_required) else "No"
    data["department"] = data.get("department_display") or po.department
    data["state_name"] = data.get("state_name_display") or po.state_name
    data["district_name"] = data.get("district_name_display") or po.district
    data["executive_name"] = data.get("executive_name_display") or po.executive_name
    data["pro_cnt"] = data.get("product_count", 0)
    data["qty"] = data.get("product_qty", 0)
    data["cons_cnt"] = data.get("consignee_count", 0)
    data["total_value"] = data.get("po_value") or data.get("total_amount") or "0"
    data["ld_required"] = "on" if _as_bool(po.ld_required) else "off"
    return data


class PurchaseOrderOptionsView(APIView):
    def get(self, request):
        tender_code = request.query_params.get("tender_code", "").strip()
        district_state = request.query_params.get("state_name", "").strip()
        company_id = request_company_id(request)
        data = {}
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT unique_id, department, acc_sector, ledger_name, ledger_no FROM department_creation WHERE is_delete = 0 AND is_active = 1 ORDER BY department"
            )
            data["customers"] = [
                {
                    "unique_id": row[0],
                    "label": row[1],
                    "acc_sector": row[2],
                    "ledger_name": row[3],
                    "ledger_no": row[4],
                }
                for row in cursor.fetchall()
            ]

            cursor.execute(
                "SELECT unique_id, account_name FROM account_vertical WHERE is_delete = '0' AND is_active = 1 ORDER BY account_name"
            )
            data["account_verticals"] = [{"unique_id": row[0], "label": row[1]} for row in cursor.fetchall()]

            cursor.execute(
                "SELECT unique_id, sector_name FROM account_sector WHERE is_delete = '0' AND is_active = 1 ORDER BY sector_name"
            )
            data["account_sectors"] = [{"unique_id": row[0], "label": row[1]} for row in cursor.fetchall()]

            executive_params = []
            executive_company_sql = ""
            if company_id:
                executive_company_sql = " AND COALESCE(sess_company_id, '') IN ('', %s)"
                executive_params.append(company_id)
            cursor.execute(
                f"""
                SELECT unique_id, executive_name
                FROM executive_name
                WHERE is_delete = 0
                  AND is_active = 1
                  {executive_company_sql}
                ORDER BY executive_name
                """,
                executive_params,
            )
            executives = [{"unique_id": row[0], "label": row[1]} for row in cursor.fetchall()]
            seen_executives = {str(row["unique_id"]) for row in executives}

            user_company_sql = ""
            user_params = ["69b0115ced3bd96390", "%executive%"]
            if company_id:
                user_company_sql = " AND COALESCE(u.sess_company_id, '') = %s"
                user_params.append(company_id)
            cursor.execute(
                f"""
                SELECT u.unique_id, u.staff_name
                FROM user u
                INNER JOIN user_type ut
                  ON ut.unique_id = u.user_type_unique_id
                 AND ut.is_delete = 0
                WHERE u.is_delete = 0
                  AND u.is_active = 1
                  AND (u.user_type_unique_id = %s OR LOWER(ut.user_type) LIKE %s)
                  {user_company_sql}
                ORDER BY u.staff_name
                """,
                user_params,
            )
            for unique_id, staff_name in cursor.fetchall():
                if str(unique_id) not in seen_executives:
                    executives.append({"unique_id": unique_id, "label": staff_name})
                    seen_executives.add(str(unique_id))
            data["executives"] = sorted(executives, key=lambda row: str(row["label"] or "").lower())

            cursor.execute(
                "SELECT staff_id, staff_name FROM user WHERE is_delete = 0 ORDER BY staff_name"
            )
            data["team_members"] = [{"unique_id": row[0], "label": row[1]} for row in cursor.fetchall()]

            cursor.execute(
                "SELECT unique_id, insurence_type FROM insurence_type WHERE is_delete = '0' AND is_active = 1 ORDER BY insurence_type"
            )
            data["insurance_types"] = [{"unique_id": row[0], "label": row[1]} for row in cursor.fetchall()]

            cursor.execute(
                "SELECT unique_id, tender_name, tender_code FROM item_creation WHERE is_delete = 0 AND is_active = 1 ORDER BY tender_name"
            )
            data["tenders"] = [{"unique_id": row[0], "label": row[1], "tender_code": row[2]} for row in cursor.fetchall()]

            item_sql = (
                "SELECT unique_id, item_code, item_description, rc_unit_price, rc_net_price, gst, warranty_in_yrs, tender_code "
                "FROM item_creation_sub WHERE is_delete = 0 AND is_active = 1"
            )
            params = []
            if tender_code:
                item_sql += " AND tender_code = %s"
                params.append(tender_code)
            item_sql += " ORDER BY item_code"
            cursor.execute(item_sql, params)
            data["items"] = [
                {
                    "unique_id": row[0],
                    "label": row[1],
                    "item_code": row[1],
                    "product": row[2],
                    "unit_price": str(row[3] or ""),
                    "net_price": str(row[4] or ""),
                    "tax": str(row[5] or ""),
                    "warranty_duration": row[6] or "",
                    "tender_code": row[7] or "",
                }
                for row in cursor.fetchall()
            ]

            cursor.execute(
                "SELECT unique_id, state_name FROM state_creation WHERE is_delete = 0 AND is_active = 1 ORDER BY state_name"
            )
            data["states"] = [{"unique_id": row[0], "label": row[1]} for row in cursor.fetchall()]

            district_sql = "SELECT unique_id, district_name, state_name FROM district_creation WHERE is_delete = 0 AND is_active = 1"
            params = []
            if district_state:
                district_sql += " AND state_name = %s"
                params.append(district_state)
            district_sql += " ORDER BY district_name"
            cursor.execute(district_sql, params)
            data["districts"] = [{"unique_id": row[0], "label": row[1], "state_name": row[2]} for row in cursor.fetchall()]

        data["po_types"] = PO_TYPES
        data["ld_types"] = LD_TYPES
        data["warranty_starts"] = WARRANTY_STARTS
        data["date_types"] = DATE_TYPES
        return Response({"status": True, "data": data})


class PurchaseOrderListView(APIView):
    def _response_raw(self, request):
        draw = int(request.data.get("draw", request.query_params.get("draw", 1)) or 1)
        start = int(request.data.get("start", request.query_params.get("start", 0)) or 0)
        length = int(request.data.get("length", request.query_params.get("length", 10)) or 10)
        search = (request.query_params.get("search") or request.data.get("search", "") or "").strip()
        from_date = (request.query_params.get("from_date") or request.data.get("from_date", "") or "").strip()
        to_date = (request.query_params.get("to_date") or request.data.get("to_date", "") or "").strip()
        month_fill = (request.query_params.get("month_fill") or request.data.get("month_fill", "") or "").strip()
        user_type_unique_id = (request.query_params.get("user_type_unique_id") or request.data.get("user_type_unique_id", "") or "").strip()
        amc_only = (request.query_params.get("amc_only") or request.data.get("amc_only", "") or "").strip()
        overdue_incomplete = (request.query_params.get("overdue_incomplete") or request.data.get("overdue_incomplete", "") or "").strip()
        current_user_id = (
            str(getattr(request.user, "unique_id", "") or getattr(request.user, "pk", "") or "")
            or (request.query_params.get("user_unique_id") or request.data.get("user_unique_id", "") or "").strip()
        )

        where = ["vp.is_delete='0'"]
        params = []
        company_id = request_company_id(request)
        if company_id:
            where.append(
                """
                EXISTS (
                    SELECT 1
                    FROM po_form pf_tenant
                    WHERE pf_tenant.unique_id = vp.unique_id
                      AND pf_tenant.is_delete = 0
                      AND pf_tenant.sess_company_id = %s
                )
                """
            )
            params.append(company_id)

        if month_fill:
            try:
                year, month = month_fill.split("-", 1)
                where.append("YEAR(vp.po_date)=%s AND MONTH(vp.po_date)=%s")
                params.extend([int(year), int(month)])
            except (TypeError, ValueError):
                pass

        if from_date:
            where.append("vp.po_date >= %s")
            params.append(from_date)
        if to_date:
            where.append("vp.po_date <= %s")
            params.append(to_date)

        if amc_only in {"1", "true", "True", "yes", "on"}:
            where.append(
                """
                EXISTS (
                    SELECT 1
                    FROM po_form pf_amc
                    WHERE pf_amc.unique_id = vp.unique_id
                      AND pf_amc.is_delete = 0
                      AND COALESCE(pf_amc.amc_required, '0') = '1'
                )
                """
            )

        if overdue_incomplete in {"1", "true", "True", "yes", "on"}:
            where.append("vp.po_date < CURDATE()")
            where.append(
                """
                (
                    COALESCE(vp.file_name, '') = ''
                    OR NOT EXISTS (
                        SELECT 1
                        FROM consignee_details_sub dis
                        WHERE dis.form_main_unique_id = vp.unique_id
                          AND dis.is_delete = 0
                    )
                    OR NOT EXISTS (
                        SELECT 1
                        FROM po_product_assign_details dis
                        WHERE dis.form_main_unique_id = vp.unique_id
                          AND dis.is_delete = 0
                    )
                    OR NOT EXISTS (
                        SELECT 1
                        FROM product_details_sub dis
                        WHERE dis.form_main_unique_id = vp.unique_id
                          AND dis.is_delete = 0
                    )
                )
                """
            )

        if user_type_unique_id == "69b0115ced3bd96390" and current_user_id:
            # Some deployments may not expose executive_name_unique_id in the view.
            has_exec_uid = False
            try:
                with connection.cursor() as cur:
                    cur.execute("SHOW COLUMNS FROM view_po_form_list LIKE 'executive_name_unique_id'")
                    has_exec_uid = cur.fetchone() is not None
            except Exception:
                has_exec_uid = False
            if has_exec_uid:
                where.append("vp.executive_name_unique_id = %s")
                params.append(current_user_id)

        if search:
            like = f"%{search}%"
            where.append(
                "("
                + " OR ".join(
                    [
                        "vp.po_num LIKE %s",
                        "vp.department LIKE %s",
                        "pf.department LIKE %s",
                        "dc.department LIKE %s",
                        "vp.executive_name LIKE %s",
                        "pf.executive_name LIKE %s",
                        "en.executive_name LIKE %s",
                        "vp.gst_value LIKE %s",
                        "vp.state_name LIKE %s",
                        "sc.state_name LIKE %s",
                        "vp.district LIKE %s",
                        "dist.district_name LIKE %s",
                        "vp.bill_address LIKE %s",
                    ]
                )
                + ")"
            )
            params.extend([like] * 13)

        where_sql = " AND ".join(where)

        has_view_id = False
        try:
            with connection.cursor() as cur:
                cur.execute("SHOW COLUMNS FROM view_po_form_list LIKE 'id'")
                has_view_id = cur.fetchone() is not None
        except Exception:
            has_view_id = False

        order_by = "vp.id DESC" if has_view_id else "vp.unique_id DESC"
        order = request.data.get("order") if isinstance(request.data, dict) else None
        if isinstance(order, list) and order:
            order_column = order[0].get("column")
            order_dir = order[0].get("dir", "desc")
            if str(order_column) == "1":
                order_by = f"vp.po_date {order_dir}"

        lookup_joins = """
        LEFT JOIN po_form pf ON pf.unique_id = vp.unique_id AND pf.is_delete = 0
        LEFT JOIN department_creation dc ON dc.unique_id = pf.department AND dc.is_delete = 0
        LEFT JOIN executive_name en ON en.unique_id = pf.executive_name AND en.is_delete = 0
        LEFT JOIN state_creation sc ON sc.unique_id = pf.state_name AND sc.is_delete = 0
        LEFT JOIN district_creation dist ON dist.unique_id = pf.district AND dist.is_delete = 0
        """

        count_sql = f"""
        SELECT COUNT(*)
        FROM view_po_form_list vp
        {lookup_joins}
        WHERE {where_sql}
        """

        data_sql = f"""
        SELECT
            vp.unique_id,
            vp.po_num,
            DATE_FORMAT(vp.po_date, '%%d-%%m-%%Y') AS po_date,
            COALESCE(NULLIF(sc.state_name, ''), NULLIF(vp.state_name, ''), NULLIF(pf.state_name, '')) AS state_name,
            COALESCE(NULLIF(dist.district_name, ''), NULLIF(vp.district, ''), NULLIF(pf.district, '')) AS district_name,
            COALESCE(NULLIF(dc.department, ''), NULLIF(vp.department, ''), NULLIF(pf.department, '')) AS department,
            vp.gst_value,
            COALESCE(NULLIF(en.executive_name, ''), NULLIF(vp.executive_name, ''), NULLIF(pf.executive_name, '')) AS executive_name,
            vp.total_amount AS total_value,
            vp.file_name,
            vp.ld_required,
            (
                SELECT COALESCE(pf.amc_required, '0')
                FROM po_form pf
                WHERE pf.unique_id = vp.unique_id
                  AND pf.is_delete = 0
                LIMIT 1
            ) AS amc_required,
            (
                SELECT COALESCE(pf.amcvalue, '0')
                FROM po_form pf
                WHERE pf.unique_id = vp.unique_id
                  AND pf.is_delete = 0
                LIMIT 1
            ) AS amcvalue,
            vp.bill_address,
            vp.delivery_due_dates,
            COALESCE(prod.cnt, 0) AS pro_cnt,
            COALESCE(prod.qty, 0) AS pro_qty,
            COALESCE(cons.cnt, 0) AS cons_cnt,
            COALESCE(assigns.cnt, 0) AS assign_cnt,
            COALESCE(stockcnt.cnt, 0) AS stock_cnt,
            (SELECT IF(EXISTS(
                SELECT 1 FROM consignee_details_sub AS dis
                WHERE dis.form_main_unique_id = vp.unique_id AND dis.is_delete = 0
            ), 1, 0)) AS consignee_exists,
            (SELECT IF(EXISTS(
                SELECT 1 FROM po_product_assign_details AS dis
                WHERE dis.form_main_unique_id = vp.unique_id AND dis.is_delete = 0
            ), 1, 0)) AS po_product_assign_exists,
            (SELECT IF(EXISTS(
                SELECT 1 FROM product_details_sub AS dis
                WHERE dis.form_main_unique_id = vp.unique_id AND dis.is_delete = 0
            ), 1, 0)) AS product_detail_exist
        FROM view_po_form_list vp
        {lookup_joins}
        LEFT JOIN (
            SELECT form_main_unique_id, COUNT(*) AS cnt, SUM(qty) AS qty
            FROM product_details_sub
            WHERE is_delete = 0
            GROUP BY form_main_unique_id
        ) prod ON prod.form_main_unique_id = vp.unique_id
        LEFT JOIN (
            SELECT form_main_unique_id, COUNT(*) AS cnt
            FROM consignee_details_sub
            WHERE is_delete = 0
            GROUP BY form_main_unique_id
        ) cons ON cons.form_main_unique_id = vp.unique_id
        LEFT JOIN (
            SELECT form_main_unique_id, COUNT(*) AS cnt
            FROM po_product_assign_details
            WHERE is_delete = 0
            GROUP BY form_main_unique_id
        ) assigns ON assigns.form_main_unique_id = vp.unique_id
        LEFT JOIN (
            SELECT form_main_unique_id, COUNT(*) AS cnt
            FROM stock_position_main
            GROUP BY form_main_unique_id
        ) stockcnt ON stockcnt.form_main_unique_id = vp.unique_id
        WHERE {where_sql}
        ORDER BY {order_by}
        """

        limit_sql = ""
        params_with_limit = list(params)
        if length != -1:
            limit_sql = " LIMIT %s OFFSET %s"
            params_with_limit.extend([length, start])

        with connection.cursor() as cursor:
            cursor.execute(count_sql, params)
            total = cursor.fetchone()[0] or 0
            cursor.execute(data_sql + limit_sql, params_with_limit)
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]

        data = []
        for idx, row in enumerate(rows, start=start + 1):
            record = dict(zip(columns, row))
            record["s_no"] = idx
            record["qty"] = record.get("pro_qty") or 0
            record["product_qty"] = record.get("pro_qty") or 0
            record["total_value"] = str(record.get("total_value") or "0")
            file_name = record.get("file_name") or ""
            if file_name:
                base = f"/api/master/purchase-order/files/po_copy/{file_name}/"
                record["file_url"] = base
            else:
                record["file_url"] = ""
            record["po_incomplete"] = (
                not record.get("file_name")
                or int(record.get("pro_cnt") or 0) == 0
                or int(record.get("cons_cnt") or 0) == 0
                or int(record.get("assign_cnt") or 0) == 0
                or int(record.get("consignee_exists") or 0) == 0
                or int(record.get("po_product_assign_exists") or 0) == 0
                or int(record.get("product_detail_exist") or 0) == 0
            )
            data.append(record)

        return Response(
            {
                "draw": draw,
                "recordsTotal": total,
                "recordsFiltered": total,
                "data": data,
            }
        )

    def get_queryset(self, request, cancel_only=False):
        qs = tenant_queryset(request, PurchaseOrder.objects.filter(is_delete=0), include_global=False)
        if cancel_only:
            qs = qs.filter(
                Q(reject_reason__isnull=False) & ~Q(reject_reason="")
                | Q(po_cancel_file__isnull=False) & ~Q(po_cancel_file="")
            )

        search = (request.query_params.get("search") or request.data.get("search", "") or "").strip()
        from_date = (request.query_params.get("from_date") or request.data.get("from_date", "") or "").strip()
        to_date = (request.query_params.get("to_date") or request.data.get("to_date", "") or "").strip()
        month_fill = (request.query_params.get("month_fill") or request.data.get("month_fill", "") or "").strip()

        if month_fill:
            try:
                year, month = month_fill.split("-", 1)
                qs = qs.filter(po_date__year=int(year), po_date__month=int(month))
            except (TypeError, ValueError):
                pass
        if from_date:
            qs = qs.filter(po_date__gte=from_date)
        if to_date:
            qs = qs.filter(po_date__lte=to_date)
        if search:
            qs = qs.filter(
                Q(po_num__icontains=search)
                | Q(department__icontains=search)
                | Q(executive_name__icontains=search)
                | Q(gst_value__icontains=search)
            )
        return qs.order_by("-po_date", "-id")

    def _response(self, request, cancel_only=False):
        if cancel_only:
            qs = self.get_queryset(request, cancel_only=cancel_only)
            draw = int(request.data.get("draw", request.query_params.get("draw", 1)) or 1)
            start = int(request.data.get("start", request.query_params.get("start", 0)) or 0)
            length = int(request.data.get("length", request.query_params.get("length", 10)) or 10)
            total = qs.count()
            page = qs if length == -1 else qs[start:start + length]
            data = [_po_row(row, idx, request) for idx, row in enumerate(page, start=start + 1)]
            return Response(
                {
                    "draw": draw,
                    "recordsTotal": total,
                    "recordsFiltered": total,
                    "data": data,
                }
            )
        return self._response_raw(request)

    def get(self, request):
        return self._response(request, cancel_only=False)

    def post(self, request):
        return self._response(request, cancel_only=False)


class PurchaseOrderCancelListView(PurchaseOrderListView):
    def get(self, request):
        return self._response(request, cancel_only=True)

    def post(self, request):
        return self._response(request, cancel_only=True)


class PurchaseOrderCreateView(APIView):
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def post(self, request):
        data = request.data.copy()
        po_num = _normalize_po_num(data.get("po_num", ""))
        if _duplicate_po_exists(po_num):
            return Response(
                {"status": False, "message": f"PO Number '{po_num}' already exists."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        po_unique_id = data.get("po_unique_id") or uuid.uuid4().hex[:18]
        unique_id = data.get("unique_id") or uuid.uuid4().hex[:18]
        po = PurchaseOrder(
            unique_id=unique_id,
            po_unique_id=po_unique_id,
            po_num=po_num,
            po_date=data.get("po_date") or datetime.now().date(),
            department=data.get("department", ""),
            gst_option=data.get("gst_option", "No"),
            gst_value=data.get("gst_value", ""),
            executive_name=data.get("executive_name", ""),
            bill_address=data.get("bill_address", ""),
            contact_name=data.get("contact_name", ""),
            contact_number=data.get("contact_number", ""),
            landline_number=data.get("landline_number", ""),
            acc_vertical=data.get("acc_vertical", ""),
            acc_sector=data.get("acc_sector", ""),
            email=data.get("email", ""),
            district=data.get("district", ""),
            pin=data.get("pin", ""),
            po_prepared_by=data.get("po_prepared_by", ""),
            po_type=data.get("po_type", "Product"),
            no_of_po=str(data.get("no_of_po", data.get("no_of_items", "0")) or "0"),
            state_name=data.get("state_name", ""),
            ld_required="1" if _as_bool(data.get("ld_required")) else "0",
            ld_date_type=data.get("ld_date_type", ""),
            ld_delivery_due_date=data.get("ld_delivery_due_date", ""),
            ld_installation_due_date=data.get("ld_installation_due_date", ""),
            insurence_required="1" if _as_bool(data.get("insurence_required")) else "0",
            insurence_types=data.get("insurence_types", ""),
            other_insurance_type=data.get("other_insurance_type", ""),
            amc_required="1" if _as_bool(data.get("amc_required")) else "0",
            start_date=data.get("start_date") or None,
            end_date=data.get("end_date") or None,
            amc_percentae=data.get("amc_percentae", ""),
            amcvalue=data.get("amcvalue", ""),
            is_active=_safe_int(data.get("is_active", 1), 1),
            acc_year=data.get("acc_year", ""),
            session_id=data.get("session_id", ""),
            sess_user_type=data.get("sess_user_type", ""),
            sess_user_id=data.get("sess_user_id", ""),
            sess_company_id=request_company_id(request),
            sess_branch_id=getattr(request.user, "sess_branch_id", ""),
        )

        po_copy = request.FILES.get("po_copy")
        if po_copy:
            filename = f"{unique_id}_po_{po_copy.name}"
            po.file_name = _save_upload(po_copy, "po_copy", filename)
            po.file_org_name = po_copy.name

        amc_file = request.FILES.get("amc_file")
        if amc_file:
            filename = f"{unique_id}_amc_{amc_file.name}"
            po.amcfile_names = _save_upload(amc_file, "amc", filename)
            po.amcfile_org_names = amc_file.name

        apply_tenant_audit(po, request)
        po.save()
        serializer = PurchaseOrderSerializer(po, context={"request": request})
        return Response({"status": True, "msg": "create", "data": serializer.data}, status=status.HTTP_201_CREATED)


class PurchaseOrderPushView(APIView):
    parser_classes = [JSONParser]

    def post(self, request):
        company_id = request_company_id(request)
        if not company_id:
            return Response(
                {"status": False, "message": "Company context missing. Login with company code and pass Bearer token."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        payload = request.data if isinstance(request.data, dict) else {}
        po_data = payload.get("po_details") or payload.get("po") or {}
        product_rows = payload.get("product_details") or payload.get("products") or []
        consignee_rows = payload.get("consignee_details") or payload.get("consignees") or []

        if not isinstance(po_data, dict):
            return Response({"status": False, "message": "po_details must be an object."}, status=status.HTTP_400_BAD_REQUEST)
        if not isinstance(product_rows, list):
            return Response({"status": False, "message": "product_details must be an array."}, status=status.HTTP_400_BAD_REQUEST)
        if not isinstance(consignee_rows, list):
            return Response({"status": False, "message": "consignee_details must be an array."}, status=status.HTTP_400_BAD_REQUEST)

        po_num = _normalize_po_num(po_data.get("po_num") or po_data.get("po_number"))
        if not po_num:
            return Response({"status": False, "message": "po_details.po_num is required."}, status=status.HTTP_400_BAD_REQUEST)
        if _duplicate_po_exists(po_num):
            return Response(
                {"status": False, "message": f"PO Number '{po_num}' already exists."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        po_unique_id = po_data.get("po_unique_id") or uuid.uuid4().hex[:18]
        unique_id = po_data.get("unique_id") or uuid.uuid4().hex[:18]

        try:
            with transaction.atomic():
                po = PurchaseOrder(
                    unique_id=unique_id,
                    po_unique_id=po_unique_id,
                    po_num=po_num,
                    po_date=po_data.get("po_date") or datetime.now().date(),
                    department=po_data.get("department", ""),
                    gst_option=po_data.get("gst_option", "No"),
                    gst_value=po_data.get("gst_value", ""),
                    executive_name=po_data.get("executive_name", ""),
                    bill_address=po_data.get("bill_address", ""),
                    contact_name=po_data.get("contact_name", ""),
                    contact_number=po_data.get("contact_number", ""),
                    landline_number=po_data.get("landline_number", ""),
                    acc_vertical=po_data.get("acc_vertical", ""),
                    acc_sector=po_data.get("acc_sector", ""),
                    email=po_data.get("email", ""),
                    district=po_data.get("district", ""),
                    pin=po_data.get("pin", ""),
                    po_prepared_by=po_data.get("po_prepared_by", ""),
                    po_type=po_data.get("po_type", "Product"),
                    no_of_po=str(po_data.get("no_of_po", len(product_rows)) or len(product_rows)),
                    state_name=po_data.get("state_name", ""),
                    ld_required="1" if _as_bool(po_data.get("ld_required")) else "0",
                    ld_date_type=po_data.get("ld_date_type", ""),
                    ld_delivery_due_date=po_data.get("ld_delivery_due_date", ""),
                    ld_installation_due_date=po_data.get("ld_installation_due_date", ""),
                    insurence_required="1" if _as_bool(po_data.get("insurence_required")) else "0",
                    insurence_types=po_data.get("insurence_types", ""),
                    other_insurance_type=po_data.get("other_insurance_type", ""),
                    amc_required="1" if _as_bool(po_data.get("amc_required")) else "0",
                    start_date=po_data.get("start_date") or None,
                    end_date=po_data.get("end_date") or None,
                    amc_percentae=po_data.get("amc_percentae", ""),
                    amcvalue=po_data.get("amcvalue", ""),
                    is_active=_safe_int(po_data.get("is_active", 1), 1),
                    acc_year=po_data.get("acc_year", ""),
                    session_id=po_data.get("session_id", ""),
                    sess_user_type=po_data.get("sess_user_type", ""),
                    sess_user_id=po_data.get("sess_user_id", ""),
                    sess_company_id=company_id,
                    sess_branch_id=getattr(request.user, "sess_branch_id", ""),
                )
                apply_tenant_audit(po, request)
                po.save()

                created_products = []
                for item in product_rows:
                    item = item or {}
                    row = PurchaseOrderProduct(
                        unique_id=item.get("unique_id") or uuid.uuid4().hex[:18],
                        form_main_unique_id=po.unique_id,
                        screen_unique_id=item.get("screen_unique_id", ""),
                        no_of_items=str(item.get("no_of_items", "")),
                        tender_code=item.get("tender_code", ""),
                        item_code=item.get("item_code", ""),
                        product=item.get("product", ""),
                        qty=str(item.get("qty", "")),
                        unit_price=str(item.get("unit_price", "")),
                        net_price=str(item.get("net_price", "")),
                        tax=str(item.get("tax", "")),
                        total_value=str(item.get("total_value", "")),
                        net_value=str(item.get("net_value", "")),
                        insta_due_days=str(item.get("insta_due_days", item.get("installation_due_days", ""))),
                        document_required=item.get("document_required", ""),
                        warranty_starts=item.get("warranty_starts", ""),
                        bg_required=item.get("bg_required", ""),
                        bg_percen=str(item.get("bg_percen", item.get("bg_percent", ""))),
                        bg_month=str(item.get("bg_month", "")),
                        rem_qty=str(item.get("rem_qty", item.get("qty", ""))),
                        assign_qty=str(item.get("assign_qty", "0")),
                        billed_qty=str(item.get("billed_qty", "0")),
                        delivery_due_dates=str(item.get("delivery_due_dates", item.get("delivery_due_days", ""))),
                        ld_type=_normalize_ld_type(item.get("ld_type", "")),
                        ld_per_day=str(item.get("ld_per_day", "")),
                        ld_maximum_val=str(item.get("ld_maximum_val", "")),
                        warranty=item.get("warranty", ""),
                        warranty_duration=str(item.get("warranty_duration", "")),
                        is_active=_safe_int(item.get("is_active", 1), 1),
                    )
                    apply_tenant_audit(row, request)
                    row.save()
                    created_products.append(row.unique_id)

                default_batch_id = payload.get("batch_id") or _generate_batch_id()
                created_consignees = []
                for item in consignee_rows:
                    item = item or {}
                    row = PurchaseOrderConsignee(
                        unique_id=item.get("unique_id") or uuid.uuid4().hex[:18],
                        form_main_unique_id=po.unique_id,
                        screen_unique_id=item.get("screen_unique_id", ""),
                        no_of_consignee=str(item.get("no_of_consignee", "")),
                        team_mem=item.get("team_mem") or item.get("assign_team_member", ""),
                        con_assign_qty=str(item.get("con_assign_qty", "")),
                        con_address=item.get("con_address", ""),
                        con_district=_map_district_unique_id(item.get("con_district", "")),
                        con_state_name=_map_state_unique_id(item.get("con_state_name", "")),
                        con_pincode=str(item.get("con_pincode", "")),
                        zone=item.get("zone", ""),
                        billing_address=item.get("billing_address", ""),
                        con_branch=item.get("con_branch", ""),
                        con_contact_name=item.get("con_contact_name", ""),
                        con_contact_number=str(item.get("con_contact_number", "")),
                        con_lan_num=str(item.get("con_lan_num", "")),
                        consignee_gst=item.get("consignee_gst", ""),
                        cons_verify_sts=str(item.get("cons_verify_sts", "0")),
                        batch_id=item.get("batch_id") or default_batch_id,
                        batch_status=_safe_int(item.get("batch_status", 0), 0),
                        po_number=po.po_num,
                        batch_entry_date=item.get("batch_entry_date") or datetime.now().date(),
                        consignee_received_date=item.get("consignee_received_date") or datetime.now().date(),
                        assign_team_member=item.get("assign_team_member", item.get("team_mem", "")),
                        con_branch_code=item.get("con_branch_code", ""),
                        alter_contact_name=item.get("alter_contact_name", ""),
                        alter_number=item.get("alter_number", ""),
                        cons_email_id=item.get("cons_email_id", ""),
                        zone_code=item.get("zone_code", ""),
                        billing_gst_no=item.get("billing_gst_no", ""),
                        region=item.get("region", ""),
                        is_active=_safe_int(item.get("is_active", 1), 1),
                    )
                    apply_tenant_audit(row, request)
                    row.save()
                    created_consignees.append(row.unique_id)

                _refresh_po_summary(po)
                serializer = PurchaseOrderSerializer(po, context={"request": request, "include_children": False})
        except Exception as exc:
            return Response({"status": False, "message": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(
            {
                "status": True,
                "msg": "push_create",
                "data": serializer.data,
                "created": {
                    "po_unique_id": po.unique_id,
                    "product_unique_ids": created_products,
                    "consignee_unique_ids": created_consignees,
                },
            },
            status=status.HTTP_201_CREATED,
        )


class PurchaseOrderDetailView(APIView):
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def _get_object(self, request, unique_id):
        return tenant_queryset(request, PurchaseOrder.objects.filter(unique_id=unique_id, is_delete=0), include_global=False).order_by("-id").first()

    def get(self, request, unique_id):
        po = self._get_object(request, unique_id)
        if not po:
            return Response({"status": False, "message": "Purchase order not found."}, status=status.HTTP_404_NOT_FOUND)
        serializer = PurchaseOrderSerializer(po, context={"request": request})
        return Response({"status": True, "data": serializer.data})

    def put(self, request, unique_id):
        po = self._get_object(request, unique_id)
        if not po:
            return Response({"status": False, "message": "Purchase order not found."}, status=status.HTTP_404_NOT_FOUND)

        data = request.data.copy()
        po_num = _normalize_po_num(data.get("po_num", po.po_num))
        if _duplicate_po_exists(po_num, exclude_unique_id=unique_id):
            return Response(
                {"status": False, "message": f"PO Number '{po_num}' already exists."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        for field in [
            "po_num", "department", "gst_option", "gst_value", "executive_name", "bill_address",
            "contact_name", "contact_number", "landline_number", "acc_vertical", "acc_sector",
            "email", "district", "pin", "po_prepared_by", "po_type", "no_of_po", "state_name", "ld_date_type",
            "ld_delivery_due_date", "ld_installation_due_date", "insurence_types",
            "other_insurance_type", "amc_percentae", "amcvalue",
        ]:
            if field in data:
                setattr(po, field, po_num if field == "po_num" else data.get(field))

        if "po_date" in data and data.get("po_date"):
            po.po_date = data.get("po_date")
        if "start_date" in data:
            po.start_date = data.get("start_date") or None
        if "end_date" in data:
            po.end_date = data.get("end_date") or None

        if "ld_required" in data:
            po.ld_required = "1" if _as_bool(data.get("ld_required")) else "0"
        if "insurence_required" in data:
            po.insurence_required = "1" if _as_bool(data.get("insurence_required")) else "0"
        if "amc_required" in data:
            po.amc_required = "1" if _as_bool(data.get("amc_required")) else "0"
        if "is_active" in data:
            po.is_active = _safe_int(data.get("is_active"), po.is_active)

        po_copy = request.FILES.get("po_copy")
        if po_copy:
            filename = f"{unique_id}_po_{po_copy.name}"
            po.file_name = _save_upload(po_copy, "po_copy", filename)
            po.file_org_name = po_copy.name

        amc_file = request.FILES.get("amc_file")
        if amc_file:
            filename = f"{unique_id}_amc_{amc_file.name}"
            po.amcfile_names = _save_upload(amc_file, "amc", filename)
            po.amcfile_org_names = amc_file.name

        apply_tenant_audit(po, request)
        po.save()
        _refresh_po_summary(po)
        serializer = PurchaseOrderSerializer(po, context={"request": request})
        return Response({"status": True, "msg": "update", "data": serializer.data})

    def delete(self, request, unique_id):
        po = self._get_object(request, unique_id)
        if not po:
            return Response({"status": False, "message": "Purchase order not found."}, status=status.HTTP_404_NOT_FOUND)
        po.is_delete = 1
        po.save(update_fields=["is_delete", "updated_at"])
        return Response({"status": True, "msg": "success_delete"})


class PurchaseOrderProductListCreateView(APIView):
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def get(self, request, unique_id):
        rows = PurchaseOrderProduct.objects.filter(form_main_unique_id=unique_id, is_delete=0).order_by("id")
        serializer = PurchaseOrderProductSerializer(rows, many=True, context={"request": request})
        return Response({"status": True, "data": serializer.data})

    def post(self, request, unique_id):
        payload = request.data.copy()
        row = PurchaseOrderProduct(
            unique_id=payload.get("unique_id") or uuid.uuid4().hex[:18],
            form_main_unique_id=unique_id,
            screen_unique_id=payload.get("screen_unique_id", ""),
            no_of_items=str(payload.get("no_of_items", "")),
            tender_code=payload.get("tender_code", ""),
            item_code=payload.get("item_code", ""),
            product=payload.get("product", ""),
            qty=str(payload.get("qty", "")),
            unit_price=str(payload.get("unit_price", "")),
            net_price=str(payload.get("net_price", "")),
            tax=str(payload.get("tax", "")),
            total_value=str(payload.get("total_value", "")),
            net_value=str(payload.get("net_value", "")),
            insta_due_days=str(payload.get("insta_due_days", payload.get("installation_due_days", ""))),
            document_required=payload.get("document_required", ""),
            warranty_starts=payload.get("warranty_starts", ""),
            bg_required=payload.get("bg_required", ""),
            bg_percen=str(payload.get("bg_percen", payload.get("bg_percent", ""))),
            bg_month=str(payload.get("bg_month", "")),
            rem_qty=str(payload.get("rem_qty", payload.get("qty", ""))),
            assign_qty=str(payload.get("assign_qty", "0")),
            billed_qty=str(payload.get("billed_qty", "0")),
            delivery_due_dates=str(payload.get("delivery_due_dates", payload.get("delivery_due_days", ""))),
            ld_type=_normalize_ld_type(payload.get("ld_type", "")),
            ld_per_day=str(payload.get("ld_per_day", "")),
            ld_maximum_val=str(payload.get("ld_maximum_val", "")),
            warranty=payload.get("warranty", ""),
            warranty_duration=str(payload.get("warranty_duration", "")),
            is_active=_safe_int(payload.get("is_active", 1), 1),
        )
        row.save()
        po = PurchaseOrder.objects.filter(unique_id=unique_id, is_delete=0).first()
        if po:
            _refresh_po_summary(po)
        return Response(
            {
                "status": True,
                "msg": "create",
                "data": {
                    "unique_id": row.unique_id,
                    "form_main_unique_id": row.form_main_unique_id,
                },
            },
            status=status.HTTP_201_CREATED,
        )


class PurchaseOrderProductDetailView(APIView):
    def _get_object(self, sub_id):
        return PurchaseOrderProduct.objects.filter(unique_id=sub_id, is_delete=0).order_by("-id").first()

    def get(self, request, sub_id):
        row = self._get_object(sub_id)
        if not row:
            return Response({"status": False, "message": "Product row not found."}, status=status.HTTP_404_NOT_FOUND)
        serializer = PurchaseOrderProductSerializer(row, context={"request": request})
        return Response({"status": True, "data": serializer.data})

    def put(self, request, sub_id):
        row = self._get_object(sub_id)
        if not row:
            return Response({"status": False, "message": "Product row not found."}, status=status.HTTP_404_NOT_FOUND)
        payload = request.data.copy()
        for field in [
            "screen_unique_id", "no_of_items", "tender_code", "item_code", "product", "qty",
            "unit_price", "net_price", "tax", "total_value", "net_value", "insta_due_days",
            "document_required", "warranty_starts", "bg_required", "bg_percen", "bg_month",
            "rem_qty", "assign_qty", "billed_qty", "delivery_due_dates", "ld_per_day",
            "ld_maximum_val", "warranty", "warranty_duration",
        ]:
            if field in payload:
                setattr(row, field, payload.get(field))
        if "ld_type" in payload:
            row.ld_type = _normalize_ld_type(payload.get("ld_type"))
        if "installation_due_days" in payload:
            row.insta_due_days = payload.get("installation_due_days")
        if "delivery_due_days" in payload:
            row.delivery_due_dates = payload.get("delivery_due_days")
        row.save()
        po = PurchaseOrder.objects.filter(unique_id=row.form_main_unique_id, is_delete=0).first()
        if po:
            _refresh_po_summary(po)
        return Response(
            {
                "status": True,
                "msg": "update",
                "data": {
                    "unique_id": row.unique_id,
                    "form_main_unique_id": row.form_main_unique_id,
                },
            }
        )

    def delete(self, request, sub_id):
        row = self._get_object(sub_id)
        if not row:
            return Response({"status": False, "message": "Product row not found."}, status=status.HTTP_404_NOT_FOUND)
        row.is_delete = 1
        row.save(update_fields=["is_delete", "updated_at"])
        po = PurchaseOrder.objects.filter(unique_id=row.form_main_unique_id, is_delete=0).first()
        if po:
            _refresh_po_summary(po)
        return Response({"status": True, "msg": "success_delete"})


class PurchaseOrderConsigneeListCreateView(APIView):
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def get(self, request, unique_id):
        rows = PurchaseOrderConsignee.objects.filter(form_main_unique_id=unique_id, is_delete=0).order_by("id")
        serializer = PurchaseOrderConsigneeSerializer(rows, many=True, context={"request": request})
        return Response({"status": True, "data": serializer.data})

    def post(self, request, unique_id):
        payload = request.data.copy()
        po = PurchaseOrder.objects.filter(unique_id=unique_id, is_delete=0).first()
        if not po:
            return Response({"status": False, "message": "Purchase order not found."}, status=status.HTTP_404_NOT_FOUND)

        batch_id = payload.get("batch_id") or _generate_batch_id()
        row = PurchaseOrderConsignee(
            unique_id=payload.get("unique_id") or uuid.uuid4().hex[:18],
            form_main_unique_id=unique_id,
            screen_unique_id=payload.get("screen_unique_id", ""),
            no_of_consignee=str(payload.get("no_of_consignee", "")),
            team_mem=payload.get("team_mem", ""),
            con_assign_qty=str(payload.get("con_assign_qty", "")),
            con_address=payload.get("con_address", ""),
            con_district=_map_district_unique_id(payload.get("con_district", "")),
            con_state_name=_map_state_unique_id(payload.get("con_state_name", "")),
            con_pincode=str(payload.get("con_pincode", "")),
            zone=payload.get("zone", ""),
            billing_address=payload.get("billing_address", ""),
            con_branch=payload.get("con_branch", ""),
            con_contact_name=payload.get("con_contact_name", ""),
            con_contact_number=str(payload.get("con_contact_number", "")),
            con_lan_num=str(payload.get("con_lan_num", "")),
            consignee_gst=payload.get("consignee_gst", ""),
            cons_verify_sts=str(payload.get("cons_verify_sts", "0")),
            batch_id=batch_id,
            batch_status=_safe_int(payload.get("batch_status", 0), 0),
            po_number=po.po_num,
            batch_entry_date=payload.get("batch_entry_date") or datetime.now().date(),
            consignee_received_date=payload.get("consignee_received_date") or datetime.now().date(),
            assign_team_member=payload.get("assign_team_member", ""),
            con_branch_code=payload.get("con_branch_code", ""),
            alter_contact_name=payload.get("alter_contact_name", ""),
            alter_number=payload.get("alter_number", ""),
            cons_email_id=payload.get("cons_email_id", ""),
            zone_code=payload.get("zone_code", ""),
            billing_gst_no=payload.get("billing_gst_no", ""),
            region=payload.get("region", ""),
            is_active=_safe_int(payload.get("is_active", 1), 1),
        )
        row.save()
        _refresh_po_summary(po)
        serializer = PurchaseOrderConsigneeSerializer(row, context={"request": request})
        return Response({"status": True, "msg": "create", "data": serializer.data}, status=status.HTTP_201_CREATED)


class PurchaseOrderConsigneeDetailView(APIView):
    def _get_object(self, sub_id):
        return PurchaseOrderConsignee.objects.filter(unique_id=sub_id, is_delete=0).order_by("-id").first()

    def get(self, request, sub_id):
        row = self._get_object(sub_id)
        if not row:
            return Response({"status": False, "message": "Consignee row not found."}, status=status.HTTP_404_NOT_FOUND)
        serializer = PurchaseOrderConsigneeSerializer(row, context={"request": request})
        return Response({"status": True, "data": serializer.data})

    def put(self, request, sub_id):
        row = self._get_object(sub_id)
        if not row:
            return Response({"status": False, "message": "Consignee row not found."}, status=status.HTTP_404_NOT_FOUND)
        payload = request.data.copy()
        team_sync_requested = "team_mem" in payload or "assign_team_member" in payload
        effective_team_member = str(payload.get("team_mem") or payload.get("assign_team_member") or "").strip()

        with transaction.atomic():
            for field in [
                "screen_unique_id", "no_of_consignee", "team_mem", "con_assign_qty", "con_address",
                "con_pincode", "zone", "billing_address", "con_branch", "con_contact_name",
                "con_contact_number", "con_lan_num", "consignee_gst", "cons_verify_sts", "batch_id",
                "batch_status", "po_number", "assign_team_member", "con_branch_code",
                "alter_contact_name", "alter_number", "cons_email_id", "zone_code", "billing_gst_no",
                "region",
            ]:
                if field in payload:
                    setattr(row, field, payload.get(field))

            if team_sync_requested:
                row.team_mem = effective_team_member
                row.assign_team_member = effective_team_member

            if "con_district" in payload:
                row.con_district = _map_district_unique_id(payload.get("con_district"))
            if "con_state_name" in payload:
                row.con_state_name = _map_state_unique_id(payload.get("con_state_name"))
            if "batch_entry_date" in payload:
                row.batch_entry_date = payload.get("batch_entry_date") or row.batch_entry_date
            if "consignee_received_date" in payload:
                row.consignee_received_date = payload.get("consignee_received_date") or row.consignee_received_date
            row.save()

            if team_sync_requested:
                _sync_consignee_team_member(row)

        po = PurchaseOrder.objects.filter(unique_id=row.form_main_unique_id, is_delete=0).first()
        if po:
            _refresh_po_summary(po)
        _create_consignee_update_notification(request, row, po)
        serializer = PurchaseOrderConsigneeSerializer(row, context={"request": request})
        return Response({"status": True, "msg": "update", "data": serializer.data})

    def delete(self, request, sub_id):
        row = self._get_object(sub_id)
        if not row:
            return Response({"status": False, "message": "Consignee row not found."}, status=status.HTTP_404_NOT_FOUND)
        row.is_delete = 1
        row.save(update_fields=["is_delete", "updated_at"])
        po = PurchaseOrder.objects.filter(unique_id=row.form_main_unique_id, is_delete=0).first()
        if po:
            _refresh_po_summary(po)
        return Response({"status": True, "msg": "success_delete"})


class PurchaseOrderConsigneeBatchListView(APIView):
    def get(self, request, unique_id):
        rows = PurchaseOrderConsignee.objects.filter(form_main_unique_id=unique_id, is_delete=0).order_by("batch_id", "id")
        products_qs = PurchaseOrderProduct.objects.filter(form_main_unique_id=unique_id, is_delete=0)
        item_count = products_qs.count()
        order_qty = sum(_safe_int(val) for val in products_qs.values_list("qty", flat=True))
        grouped = {}
        for row in rows:
            if row.batch_id not in grouped:
                grouped[row.batch_id] = {
                    "batch_id": row.batch_id,
                    "batch_insert_date": row.batch_entry_date,
                    "consignee_received_date": row.consignee_received_date,
                    "total_consignee_count": 0,
                    "item_count": item_count,
                    "order_qty": order_qty,
                    "verified": str(row.cons_verify_sts) == "1",
                    "batch_cnt": 0,
                    "form_main_unique_id": unique_id,
                }
            grouped[row.batch_id]["total_consignee_count"] += 1
        batch_ids = list(grouped.keys())
        if batch_ids:
            counts = (
                PurchaseOrderAssign.objects
                .filter(form_main_unique_id=unique_id, batch_id__in=batch_ids, is_delete=0)
                .values("batch_id")
                .annotate(total=Count("id"))
            )
            for item in counts:
                grouped[item["batch_id"]]["batch_cnt"] = item["total"]
        data = []
        for idx, item in enumerate(grouped.values(), start=1):
            data.append({"s_no": idx, **item})
        return Response({"status": True, "data": data})


class PurchaseOrderConsigneeBatchDeleteView(APIView):
    def delete(self, request, unique_id, batch_id):
        rows = PurchaseOrderConsignee.objects.filter(form_main_unique_id=unique_id, batch_id=batch_id, is_delete=0)
        if not rows.exists():
            return Response({"status": False, "message": "Batch not found."}, status=status.HTTP_404_NOT_FOUND)

        has_assign = PurchaseOrderAssign.objects.filter(form_main_unique_id=unique_id, batch_id=batch_id, is_delete=0).exists()
        if has_assign:
            return Response({"status": False, "message": "Assigned batch cannot be deleted."}, status=status.HTTP_400_BAD_REQUEST)

        rows.update(is_delete=1)
        po = PurchaseOrder.objects.filter(unique_id=unique_id, is_delete=0).first()
        if po:
            _refresh_po_summary(po)
        return Response({"status": True, "msg": "success_delete"})


class PurchaseOrderConsigneeBatchDateUpdateView(APIView):
    parser_classes = [JSONParser, FormParser]

    def post(self, request, unique_id, batch_id):
        received_date = request.data.get("consignee_received_date")
        if not received_date:
            return Response({"status": False, "message": "consignee_received_date is required."}, status=status.HTTP_400_BAD_REQUEST)

        rows = PurchaseOrderConsignee.objects.filter(form_main_unique_id=unique_id, batch_id=batch_id, is_delete=0)
        if not rows.exists():
            return Response({"status": False, "message": "Batch not found."}, status=status.HTTP_404_NOT_FOUND)

        rows.update(consignee_received_date=received_date)
        return Response({"status": True, "msg": "update"})


class PurchaseOrderConsigneeImportView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, unique_id):
        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"status": False, "message": "Excel file is required."}, status=status.HTTP_400_BAD_REQUEST)

        po = PurchaseOrder.objects.filter(unique_id=unique_id, is_delete=0).first()
        if not po:
            return Response({"status": False, "message": "Purchase order not found."}, status=status.HTTP_404_NOT_FOUND)

        batch_id = request.data.get("batch_id") or _generate_batch_id()
        received_date = request.data.get("consignee_received_date") or datetime.now().date()
        created_rows = 0

        def save_row(values):
            nonlocal created_rows
            PurchaseOrderConsignee.objects.create(
                unique_id=uuid.uuid4().hex[:18],
                form_main_unique_id=unique_id,
                no_of_consignee=str(request.data.get("no_of_consignee", "")),
                billing_address=values[0] if len(values) > 0 else "",
                billing_gst_no=values[1] if len(values) > 1 else "",
                con_branch=values[2] if len(values) > 2 else "",
                con_branch_code=values[3] if len(values) > 3 else "",
                con_address=values[4] if len(values) > 4 else "",
                consignee_gst=values[5] if len(values) > 5 else "",
                cons_email_id=values[6] if len(values) > 6 else "",
                con_state_name=_map_state_unique_id(values[7] if len(values) > 7 else ""),
                con_district=_map_district_unique_id(values[8] if len(values) > 8 else ""),
                region=values[9] if len(values) > 9 else "",
                con_contact_name=values[10] if len(values) > 10 else "",
                con_contact_number=values[11] if len(values) > 11 else "",
                alter_contact_name=values[12] if len(values) > 12 else "",
                alter_number=values[13] if len(values) > 13 else "",
                con_lan_num=values[14] if len(values) > 14 else "",
                zone=values[15] if len(values) > 15 else "",
                zone_code=values[16] if len(values) > 16 else "",
                con_pincode=values[17] if len(values) > 17 else "",
                batch_id=batch_id,
                team_mem=values[18] if len(values) > 18 else "",
                po_number=po.po_num,
                batch_entry_date=datetime.now().date(),
                consignee_received_date=received_date,
                cons_verify_sts=str(request.data.get("cons_verify_sts", "0")),
                batch_status=0,
                is_active=1,
                is_delete=0,
            )
            created_rows += 1

        try:
            first = True
            for values in _parse_uploaded_rows(file_obj):
                if first:
                    first = False
                    header_probe = " ".join(values).lower()
                    if "billing" in header_probe or "consignee" in header_probe:
                        continue
                save_row(values)
        except RuntimeError as exc:
            return Response({"status": False, "message": str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except BadZipFile:
            return Response(
                {
                    "status": False,
                    "message": "Invalid Excel file. Upload the downloaded CSV template as .csv, or upload a real .xlsx file.",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        _refresh_po_summary(po)
        return Response(
            {
                "status": True,
                "msg": "success",
                "batch_id": batch_id,
                "created_count": created_rows,
            }
        )


class PurchaseOrderAssignImportView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, unique_id):
        file_obj = request.FILES.get("file")
        batch_id = str(request.data.get("batch_id") or "").strip()
        if not file_obj:
            return Response({"status": False, "message": "Excel file is required."}, status=status.HTTP_400_BAD_REQUEST)
        if not batch_id:
            return Response({"status": False, "message": "batch_id is required."}, status=status.HTTP_400_BAD_REQUEST)

        po = PurchaseOrder.objects.filter(unique_id=unique_id, is_delete=0).first()
        if not po:
            return Response({"status": False, "message": "Purchase order not found."}, status=status.HTTP_404_NOT_FOUND)

        consignees = list(
            PurchaseOrderConsignee.objects.filter(
                form_main_unique_id=unique_id,
                batch_id=batch_id,
                is_delete=0,
            ).order_by("id")
        )
        if not consignees:
            return Response({"status": False, "message": "Consignee batch not found."}, status=status.HTTP_404_NOT_FOUND)

        products = list(PurchaseOrderProduct.objects.filter(form_main_unique_id=unique_id, is_delete=0).order_by("id"))
        if not products:
            return Response({"status": False, "message": "No products found for this PO."}, status=status.HTTP_400_BAD_REQUEST)

        consignee_by_id = {str(row.unique_id): row for row in consignees}
        consignee_by_match = {
            (
                (row.con_branch or "").strip().lower(),
                (row.con_contact_number or "").strip(),
                (row.con_address or "").strip().lower(),
            ): row
            for row in consignees
        }
        product_by_id = {str(row.unique_id): row for row in products}
        product_by_item = {str(row.item_code or "").strip().lower(): row for row in products if row.item_code}

        try:
            raw_rows = _parse_uploaded_rows(file_obj)
        except RuntimeError as exc:
            return Response({"status": False, "message": str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except BadZipFile:
            return Response(
                {"status": False, "message": "Invalid Excel file. Please upload a valid .csv or .xlsx file."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not raw_rows:
            return Response({"status": False, "message": "Import file is empty."}, status=status.HTTP_400_BAD_REQUEST)

        header_row = [str(cell or "").strip().lower() for cell in raw_rows[0]]
        first_row_probe = " ".join(header_row)
        has_header = any(token in first_row_probe for token in ["assign qty", "item code", "product unique", "con unique", "batch id"])
        data_rows = raw_rows[1:] if has_header else raw_rows

        header_index = {name: idx for idx, name in enumerate(header_row)} if has_header else {}

        def cell(values, *names, default=""):
            for name in names:
                idx = header_index.get(name)
                if idx is not None and idx < len(values):
                    return values[idx]
            return default

        parsed_rows = []
        import_totals = defaultdict(int)

        for row_index, values in enumerate(data_rows, start=2 if has_header else 1):
            if not any(values):
                continue

            if has_header:
                con_name = cell(values, "consignee name", default="")
                con_contact_no = cell(values, "contact no", default="")
                con_address = cell(values, "consignee address", default="")
                batch_id_xls = cell(values, "batch id", default=batch_id)
                po_number_xls = cell(values, "po number", default=po.po_num)
                item_code = cell(values, "item code", default="")
                product_name = cell(values, "product", default="")
                qty = _safe_int(cell(values, "qty", default="0"))
                con_unique_id = cell(values, "con unique id", default="")
                product_unique_id = cell(values, "product unique id", default="")
                unit_price = cell(values, "unit price", default="")
                assign_qty = _safe_int(cell(values, "assign qty", default="0"))
            else:
                con_name = values[0] if len(values) > 0 else ""
                con_contact_no = values[1] if len(values) > 1 else ""
                con_address = values[2] if len(values) > 2 else ""
                batch_id_xls = values[3] if len(values) > 3 else batch_id
                po_number_xls = values[4] if len(values) > 4 else po.po_num
                item_code = values[5] if len(values) > 5 else ""
                product_name = values[6] if len(values) > 6 else ""
                qty = _safe_int(values[7] if len(values) > 7 else 0)
                con_unique_id = values[8] if len(values) > 8 else ""
                product_unique_id = values[9] if len(values) > 9 else ""
                unit_price = values[10] if len(values) > 10 else ""
                assign_qty = _safe_int(values[11] if len(values) > 11 else 0)

            if batch_id_xls and batch_id_xls != batch_id:
                return Response(
                    {"status": False, "message": f"Row {row_index}: batch id does not match the selected batch."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            consignee = consignee_by_id.get(str(con_unique_id).strip())
            if not consignee:
                consignee = consignee_by_match.get(
                    (
                        (con_name or "").strip().lower(),
                        (con_contact_no or "").strip(),
                        (con_address or "").strip().lower(),
                    )
                )
            if not consignee:
                consignee = next(
                    (
                        row
                        for row in consignees
                        if (row.con_contact_number or "").strip() == (con_contact_no or "").strip()
                        and (row.con_address or "").strip().lower() == (con_address or "").strip().lower()
                    ),
                    None,
                )
            if not consignee:
                return Response(
                    {"status": False, "message": f"Row {row_index}: consignee not found in this batch."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            product = product_by_id.get(str(product_unique_id).strip())
            if not product and item_code:
                product = product_by_item.get(str(item_code).strip().lower())
            if not product:
                return Response(
                    {"status": False, "message": f"Row {row_index}: product not found for item code '{item_code}'."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            parsed_rows.append(
                {
                    "consignee": consignee,
                    "product": product,
                    "con_name": con_name or consignee.con_branch or "",
                    "con_contact_no": con_contact_no or consignee.con_contact_number or "",
                    "con_address": con_address or consignee.con_address or "",
                    "po_number": po_number_xls or po.po_num or "",
                    "qty": qty or _safe_int(product.qty),
                    "assign_qty": assign_qty,
                    "unit_price": unit_price or str(product.unit_price or ""),
                    "item_code": item_code or product.item_code or "",
                    "product_name": product_name or product.product or "",
                }
            )
            import_totals[product.unique_id] += assign_qty

        if not parsed_rows:
            return Response({"status": False, "message": "No valid rows found in import file."}, status=status.HTTP_400_BAD_REQUEST)

        for product_unique_id, current_total in import_totals.items():
            product = product_by_id[product_unique_id]
            other_batches = (
                PurchaseOrderAssign.objects.filter(
                    form_main_unique_id=unique_id,
                    product_unique_id=product_unique_id,
                    is_delete=0,
                ).exclude(batch_id=batch_id)
            )
            already_assigned = sum(item.assign_qty for item in other_batches)
            if already_assigned + current_total > _safe_int(product.qty):
                return Response(
                    {
                        "status": False,
                        "message": f"Upload failed. Assigned quantity for item '{product.item_code}' exceeds available quantity ({product.qty}).",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        created_count = 0
        updated_count = 0
        for entry in parsed_rows:
            consignee = entry["consignee"]
            product = entry["product"]
            obj = (
                PurchaseOrderAssign.objects.filter(
                    form_main_unique_id=unique_id,
                    batch_id=batch_id,
                    con_unique_id=consignee.unique_id,
                    product_unique_id=product.unique_id,
                    is_delete=0,
                )
                .order_by("-id")
                .first()
            )

            payload = {
                "con_name": entry["con_name"],
                "con_contact_no": entry["con_contact_no"],
                "con_address": entry["con_address"],
                "con_assign_team_member": consignee.assign_team_member or "",
                "unit_price": entry["unit_price"],
                "item_tax": str(product.tax or ""),
                "po_num": po.po_num,
                "po_date": str(po.po_date or ""),
                "no_of_consignee": str(po.no_of_consignee or len(consignees)),
                "no_of_item": str(po.no_of_po or len(products)),
                "executive_name": po.executive_name or "",
                "item_code": entry["item_code"],
                "product": entry["product_name"],
                "qty": entry["qty"] or _safe_int(product.qty),
                "assign_qty": entry["assign_qty"],
                "rem_qty": _safe_int(product.qty) - entry["assign_qty"],
                "assign_value": _to_decimal(entry["unit_price"] or product.unit_price) * Decimal(entry["assign_qty"]),
                "po_number": entry["po_number"],
            }

            if obj:
                for field, value in payload.items():
                    setattr(obj, field, value)
                obj.save()
                updated_count += 1
            else:
                PurchaseOrderAssign.objects.create(
                    unique_id=uuid.uuid4().hex[:18],
                    form_main_unique_id=unique_id,
                    po_unique_id=po.po_unique_id,
                    con_unique_id=consignee.unique_id,
                    product_unique_id=product.unique_id,
                    batch_id=batch_id,
                    is_active=1,
                    is_delete=0,
                    acc_year=po.acc_year,
                    session_id=po.session_id,
                    sess_user_type=po.sess_user_type,
                    sess_user_id=po.sess_user_id,
                    sess_company_id=po.sess_company_id,
                    sess_branch_id=po.sess_branch_id,
                    **payload,
                )
                created_count += 1

        _refresh_po_summary(po)
        return Response(
            {
                "status": True,
                "msg": "success",
                "created_count": created_count,
                "updated_count": updated_count,
                "batch_id": batch_id,
            }
        )


class PurchaseOrderPendingVerifyView(APIView):
    def get(self, request, unique_id):
        rows = PurchaseOrderConsignee.objects.filter(
            form_main_unique_id=unique_id,
            is_delete=0,
            cons_verify_sts="0",
        ).order_by("batch_id", "id")
        grouped = {}
        for row in rows:
            grouped.setdefault(row.batch_id, 0)
            grouped[row.batch_id] += 1
        data = [{"s_no": idx, "batch_id": batch_id, "status": "Pending", "count": count} for idx, (batch_id, count) in enumerate(grouped.items(), start=1)]
        return Response({"status": True, "data": data})


class PurchaseOrderVerifyBatchesView(APIView):
    parser_classes = [JSONParser, FormParser]

    def post(self, request):
        batch_ids = request.data.get("batch_ids") or []
        if isinstance(batch_ids, str):
            batch_ids = [item.strip() for item in batch_ids.split(",") if item.strip()]
        if not batch_ids:
            return Response({"status": False, "message": "batch_ids is required."}, status=status.HTTP_400_BAD_REQUEST)
        PurchaseOrderConsignee.objects.filter(batch_id__in=batch_ids, is_delete=0).update(cons_verify_sts="1", batch_status=1)
        return Response({"status": True, "msg": "success"})


class PurchaseOrderAssignView(APIView):
    def get(self, request, unique_id, batch_id):
        po = PurchaseOrder.objects.filter(unique_id=unique_id, is_delete=0).first()
        if not po:
            return Response({"status": False, "message": "Purchase order not found."}, status=status.HTTP_404_NOT_FOUND)

        consignees = PurchaseOrderConsignee.objects.filter(form_main_unique_id=unique_id, batch_id=batch_id, is_delete=0).order_by("id")
        products = list(PurchaseOrderProduct.objects.filter(form_main_unique_id=unique_id, is_delete=0).order_by("id"))
        assigns = PurchaseOrderAssign.objects.filter(form_main_unique_id=unique_id, batch_id=batch_id, is_delete=0)
        item_lookup = {}
        item_unique_ids = [product.item_code for product in products if str(product.item_code or "").startswith("item_")]
        if item_unique_ids:
            with connection.cursor() as cursor:
                placeholders = ",".join(["%s"] * len(item_unique_ids))
                cursor.execute(
                    f"""
                    SELECT unique_id, item_code, tender_code
                    FROM item_creation_sub
                    WHERE unique_id IN ({placeholders}) AND is_delete = 0
                    """,
                    item_unique_ids,
                )
                item_lookup = {
                    row[0]: (row[1] or row[2] or "")
                    for row in cursor.fetchall()
                }

        assign_map = {(row.con_unique_id, row.product_unique_id): row for row in assigns}
        data = []
        for consignee in consignees:
            product_rows = []
            for product in products:
                assigned = assign_map.get((consignee.unique_id, product.unique_id))
                display_item_code = (
                    product.tender_code
                    or item_lookup.get(product.item_code, "")
                    or product.item_code
                )
                product_rows.append(
                    {
                        "product_unique_id": product.unique_id,
                        "item_code": display_item_code,
                        "product": product.product,
                        "qty": _safe_int(product.qty),
                        "assign_qty": assigned.assign_qty if assigned else 0,
                        "unit_price": product.unit_price,
                        "assign_unique_id": assigned.unique_id if assigned else "",
                    }
                )
            data.append(
                {
                    "con_unique_id": consignee.unique_id,
                    "con_name": consignee.con_branch,
                    "con_contact_no": consignee.con_contact_number,
                    "con_address": consignee.con_address,
                    "batch_id": batch_id,
                    "po_num": po.po_num,
                    "po_date": po.po_date,
                    "products": product_rows,
                }
            )
        return Response({"status": True, "data": data})


class PurchaseOrderAssignSaveView(APIView):
    parser_classes = [JSONParser, FormParser]

    def post(self, request, unique_id):
        po = PurchaseOrder.objects.filter(unique_id=unique_id, is_delete=0).first()
        if not po:
            return Response({"status": False, "message": "Purchase order not found."}, status=status.HTTP_404_NOT_FOUND)

        batch_id = request.data.get("batch_id", "")
        rows = request.data.get("rows") or []
        if not batch_id or not isinstance(rows, list):
            return Response({"status": False, "message": "batch_id and rows are required."}, status=status.HTTP_400_BAD_REQUEST)

        totals = {}
        for row in rows:
            product_unique_id = row.get("product_unique_id", "")
            totals.setdefault(product_unique_id, 0)
            totals[product_unique_id] += _safe_int(row.get("assign_qty"))

        for product_unique_id, current_total in totals.items():
            product = PurchaseOrderProduct.objects.filter(unique_id=product_unique_id, form_main_unique_id=unique_id, is_delete=0).first()
            if not product:
                return Response({"status": False, "message": f"Product not found for {product_unique_id}."}, status=status.HTTP_400_BAD_REQUEST)
            other_batches = (
                PurchaseOrderAssign.objects
                .filter(form_main_unique_id=unique_id, product_unique_id=product_unique_id, is_delete=0)
                .exclude(batch_id=batch_id)
            )
            already_assigned = sum(item.assign_qty for item in other_batches)
            if already_assigned + current_total > _safe_int(product.qty):
                return Response(
                    {
                        "status": False,
                        "message": f"Assigned quantity exceeds available quantity for item {product.item_code}.",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        for row in rows:
            assign_qty = _safe_int(row.get("assign_qty"))
            con_unique_id = row.get("con_unique_id", "")
            product_unique_id = row.get("product_unique_id", "")
            product = PurchaseOrderProduct.objects.filter(unique_id=product_unique_id, form_main_unique_id=unique_id, is_delete=0).first()
            consignee = PurchaseOrderConsignee.objects.filter(unique_id=con_unique_id, form_main_unique_id=unique_id, is_delete=0).first()
            if not product or not consignee:
                continue

            obj = (
                PurchaseOrderAssign.objects
                .filter(form_main_unique_id=unique_id, batch_id=batch_id, con_unique_id=con_unique_id, product_unique_id=product_unique_id, is_delete=0)
                .order_by("-id")
                .first()
            )
            if obj:
                obj.assign_qty = assign_qty
                obj.qty = _safe_int(product.qty)
                obj.rem_qty = _safe_int(product.qty) - assign_qty
                obj.assign_value = _to_decimal(product.unit_price) * Decimal(assign_qty)
                obj.save()
            else:
                PurchaseOrderAssign.objects.create(
                    unique_id=uuid.uuid4().hex[:18],
                    form_main_unique_id=unique_id,
                    po_unique_id=po.po_unique_id,
                    con_unique_id=consignee.unique_id,
                    con_name=consignee.con_branch,
                    con_contact_no=consignee.con_contact_number,
                    con_address=consignee.con_address,
                    con_assign_team_member=consignee.assign_team_member,
                    unit_price=str(product.unit_price or ""),
                    item_tax=str(product.tax or ""),
                    po_num=po.po_num,
                    po_date=str(po.po_date),
                    no_of_consignee=str(po.no_of_consignee or ""),
                    no_of_item=str(po.no_of_po or ""),
                    executive_name=po.executive_name,
                    product_unique_id=product.unique_id,
                    item_code=product.item_code,
                    product=product.product,
                    qty=_safe_int(product.qty),
                    assign_qty=assign_qty,
                    rem_qty=_safe_int(product.qty) - assign_qty,
                    assign_value=_to_decimal(product.unit_price) * Decimal(assign_qty),
                    batch_id=batch_id,
                    po_number=po.po_num,
                )

        _refresh_po_summary(po)
        return Response({"status": True, "msg": "success"})


class PurchaseOrderAmcListCreateView(APIView):
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get(self, request, unique_id):
        po = PurchaseOrder.objects.filter(unique_id=unique_id, is_delete=0).first()
        if not po:
            return Response({"status": False, "message": "Purchase order not found."}, status=status.HTTP_404_NOT_FOUND)

        rows = PurchaseOrderAmc.objects.filter(form_main_unique_id=unique_id, is_delete=0).order_by("-id")
        serializer = PurchaseOrderAmcSerializer(rows, many=True, context={"request": request})
        return Response({"status": True, "data": serializer.data})

    def post(self, request, unique_id):
        po = PurchaseOrder.objects.filter(unique_id=unique_id, is_delete=0).first()
        if not po:
            return Response({"status": False, "message": "Purchase order not found."}, status=status.HTTP_404_NOT_FOUND)

        data = request.data
        required_fields = [
            "start_date",
            "end_date",
            "amc_percentae",
            "amcvalue",
            "amc_tax",
            "amc_unit_price",
            "amc_remarks",
        ]
        missing = [field for field in required_fields if not str(data.get(field) or "").strip()]
        if missing:
            return Response({"status": False, "message": f"Missing required fields: {', '.join(missing)}."}, status=status.HTTP_400_BAD_REQUEST)

        amc_file = request.FILES.get("amc_file")
        if not amc_file:
            return Response({"status": False, "message": "AMC attach copy is required."}, status=status.HTTP_400_BAD_REQUEST)

        if str(po.file_name or "").strip():
            po_file_name = po.file_name
            po_file_org_name = po.file_org_name
        else:
            po_copy = request.FILES.get("po_copy")
            if not po_copy:
                return Response({"status": False, "message": "PO copy is required before adding AMC sublist."}, status=status.HTTP_400_BAD_REQUEST)
            filename = f"{unique_id}_po_{po_copy.name}"
            po.file_name = _save_upload(po_copy, "po_copy", filename)
            po.file_org_name = po_copy.name
            po.save(update_fields=["file_name", "file_org_name", "updated_at"])
            po_file_name = po.file_name
            po_file_org_name = po.file_org_name

        amc_filename = f"{unique_id}_amc_sub_{uuid.uuid4().hex[:8]}_{amc_file.name}"
        saved_amc = _save_upload(amc_file, "amc", amc_filename)

        row = PurchaseOrderAmc.objects.create(
            unique_id=uuid.uuid4().hex[:18],
            form_main_unique_id=unique_id,
            po_no=po.po_num or "",
            po_unique_id=unique_id,
            start_date=data.get("start_date") or None,
            end_date=data.get("end_date") or None,
            amc_percentae=str(data.get("amc_percentae", "") or ""),
            amcvalue=str(data.get("amcvalue", "") or ""),
            amc_tax=str(data.get("amc_tax", "") or ""),
            amc_unit_price=str(data.get("amc_unit_price", "") or ""),
            amc_remarks=str(data.get("amc_remarks", "") or ""),
            amcfile_names=saved_amc,
            amcfile_org_names=amc_file.name,
            po_file_name=po_file_name,
            po_file_org_name=po_file_org_name,
            batch_id=str(data.get("batch_id", "") or ""),
            is_active=1,
            is_delete=0,
            acc_year=po.acc_year or _current_academic_year(),
            session_id=po.session_id or "",
            sess_user_type=po.sess_user_type or "",
            sess_user_id=po.sess_user_id or "",
            sess_company_id=po.sess_company_id or "",
            sess_branch_id=po.sess_branch_id or "",
        )
        serializer = PurchaseOrderAmcSerializer(row, context={"request": request})
        return Response({"status": True, "msg": "create", "data": serializer.data}, status=status.HTTP_201_CREATED)


class PurchaseOrderAmcDetailView(APIView):
    def delete(self, request, sub_id):
        row = PurchaseOrderAmc.objects.filter(unique_id=sub_id, is_delete=0).order_by("-id").first()
        if not row:
            return Response({"status": False, "message": "AMC sublist not found."}, status=status.HTTP_404_NOT_FOUND)
        row.is_delete = 1
        row.save(update_fields=["is_delete"])
        return Response({"status": True, "msg": "success_delete"})


class PurchaseOrderCancelView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, unique_id):
        po = PurchaseOrder.objects.filter(unique_id=unique_id, is_delete=0).first()
        if not po:
            return Response({"status": False, "message": "Purchase order not found."}, status=status.HTTP_404_NOT_FOUND)

        po.reject_reason = request.data.get("reject_reason", po.reject_reason)
        cancel_file = request.FILES.get("file")
        if cancel_file:
            filename = f"{unique_id}_cancel_{cancel_file.name}"
            po.po_cancel_file = _save_upload(cancel_file, "po_cancel", filename)
            po.po_cancel_file_orgname = cancel_file.name
        apply_tenant_audit(po, request)
        po.save()
        serializer = PurchaseOrderSerializer(po, context={"request": request})
        return Response({"status": True, "msg": "update", "data": serializer.data})


@method_decorator(xframe_options_exempt, name="dispatch")
class PurchaseOrderFileView(APIView):
    authentication_classes = []
    permission_classes = []

    def get(self, request, folder, filename):
        safe_folder = folder.strip().replace("..", "")
        safe_name = os.path.basename(filename)
        path = _resolve_po_file_path(safe_folder, safe_name)
        if not os.path.exists(path):
            raise Http404("File not found.")
        return FileResponse(open(path, "rb"), as_attachment=False, filename=safe_name)





