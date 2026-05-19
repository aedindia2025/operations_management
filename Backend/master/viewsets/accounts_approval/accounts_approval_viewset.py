"""
master/viewsets/accounts_approval_viewset.py
"""

from __future__ import annotations

import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import connection, transaction
from django.db.models import Q
from django.http import HttpResponse
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.viewsets import ViewSet

from master.apps.operation_approval.operationapprovalmodel import (
    InvoiceCreation,
    InvoiceCreationMain as OperationApprovalInvoiceCreationMain,
    InvoiceSublist,
    InvoiceVerificationTable,
)
from master.apps.purchase_order.purchaseordermodel import PurchaseOrder, PurchaseOrderConsignee
from master.apps.user.usermodel import UserCreation
from master.apps.vendor_allocation.vendorallocation_model import OperationInvoiceCreation as InvoiceCreationMain
from master.serializers.accounts_approval.accounts_approval_serializer import (
    AccountsApprovalDetailSerializer,
    AccountsApprovalListSerializer,
    AccountsApprovalUpdateSerializer,
    OverallApprovalSerializer,
    PendingListSerializer,
)
from master.tenant import request_company_id


def _paginate(queryset, start: int, length: int):
    total = queryset.count()
    if length and length != -1:
        page_qs = queryset[start : start + length]
    else:
        page_qs = queryset
    return page_qs, total


def _apply_date_filter(queryset, from_date: str, to_date: str, opt: str):
    if not from_date or not to_date:
        return queryset
    if opt == "4":
        po_ids = list(
            PurchaseOrder.objects.filter(po_date__gte=from_date, po_date__lte=to_date, is_delete=0).values_list(
                "unique_id", flat=True
            )
        )
        return queryset.filter(form_main_unique_id__in=po_ids)
    if opt == "5":
        return queryset.filter(invoice_date__gte=from_date, invoice_date__lte=to_date)
    return queryset


def _apply_search(queryset, search: str):
    if not search:
        return queryset

    po_ids = list(
        PurchaseOrder.objects.filter(Q(po_num__icontains=search) | Q(po_unique_id__icontains=search), is_delete=0).values_list(
            "unique_id", flat=True
        )
    )
    consignee_ids = list(
        PurchaseOrderConsignee.objects.filter(
            Q(con_address__icontains=search)
            | Q(con_contact_name__icontains=search)
            | Q(po_number__icontains=search),
            is_delete=0,
        ).values_list("unique_id", flat=True)
    )
    staff_ids = list(
        UserCreation.objects.filter(
            Q(staff_name__icontains=search) | Q(staff_id__icontains=search),
            is_delete=0,
        ).values_list("staff_id", flat=True)
    )

    return queryset.filter(
        Q(dc_number__icontains=search)
        | Q(invoice_no__icontains=search)
        | Q(form_main_unique_id__in=po_ids)
        | Q(consignee_unique_id__in=consignee_ids)
        | Q(team_mem__in=staff_ids)
    )


def _number_rows(objects, start: int):
    for i, obj in enumerate(objects, start=start + 1):
        obj._sno = i
    return objects


def _apply_company_filter(queryset, request):
    company_id = request_company_id(request)
    if not company_id:
        return queryset.none()
    po_ids = PurchaseOrder.objects.filter(
        is_delete=0,
        sess_company_id=company_id,
    ).values_list("unique_id", flat=True)
    return queryset.filter(form_main_unique_id__in=po_ids)


def _indian_money(value) -> str:
    try:
        val = float(value or 0)
        s = f"{val:.2f}"
        integer_part, decimal_part = s.split(".")
        if len(integer_part) > 3:
            last3 = integer_part[-3:]
            rest = integer_part[:-3]
            groups: list[str] = []
            while len(rest) > 2:
                groups.append(rest[-2:])
                rest = rest[:-2]
            if rest:
                groups.append(rest)
            groups.reverse()
            integer_part = ",".join(groups) + "," + last3
        return f"{integer_part}.{decimal_part}"
    except Exception:
        return str(value or "0.00")


def _status_label(invoice_doc_status):
    mapping = {4: "Approved", 2: "Not Approved", "4": "Approved", "2": "Not Approved"}
    return mapping.get(invoice_doc_status, "Pending")


def _dictfetchall(cursor):
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def _to_decimal(value) -> Decimal:
    try:
        return Decimal(str(value or "0"))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal("0")


def _get_item_rows(form_main_unique_id: str, consignee_unique_id: str, dc_number: str):
    """Fetch invoice line items for the given invoice"""
    with connection.cursor() as cur:
        cur.execute(
            """
            SELECT item_code, product, invoice_qty, invoice_qty_value, ser_no
            FROM invoice_creation
            WHERE dc_num = %s
              AND po_unique_id = %s
              AND consignee_id = %s
              AND is_delete = 0
            ORDER BY id ASC
            """,
            [dc_number, form_main_unique_id, consignee_unique_id],
        )
        rows = _dictfetchall(cur)

    data = []
    for idx, row in enumerate(rows, start=1):
        try:
            # Apply 18% GST on invoice value (using quantize for proper decimal handling)
            invoice_value = (_to_decimal(row.get("invoice_qty_value")) * Decimal("1.18")).quantize(Decimal("0.01"))
        except Exception:
            invoice_value = Decimal("0.00")
        
        data.append(
            {
                "id": idx,
                "itemName": row.get("item_code", ""),
                "itemDesc": row.get("product", ""),
                "dcQty": int(row.get("invoice_qty") or 0),
                "invoiceValue": float(invoice_value),
                "serialNo": str(row.get("ser_no") or "").replace(",", ", "),
            }
        )
    return data


class AccountsApprovalViewSet(ViewSet):
    permission_classes = []

    def retrieve(self, request, pk=None):
        try:
            record = InvoiceCreationMain.objects.get(unique_id=pk, is_delete=0)
        except InvoiceCreationMain.DoesNotExist:
            return Response(
                {"status": False, "error": "Record not found", "msg": "error"},
                status=status.HTTP_404_NOT_FOUND,
            )

        serializer = AccountsApprovalDetailSerializer(record)
        return Response({"status": True, "data": serializer.data, "error": "", "msg": "success"})

    @action(detail=False, methods=["get"], url_path="pending")
    def pending(self, request):
        search = request.query_params.get("search", "")
        length = int(request.query_params.get("length", 10))
        start = int(request.query_params.get("start", 0))
        from_date = request.query_params.get("from_date", "")
        to_date = request.query_params.get("to_date", "")
        opt = request.query_params.get("opt", "4")
        team_mem = request.query_params.get("team_mem", "")
        order_dir = request.query_params.get("order_dir", "asc")

        qs = InvoiceCreationMain.objects.filter(invoice_doc_status=1, is_delete=0)
        qs = _apply_company_filter(qs, request)
        qs = _apply_date_filter(qs, from_date, to_date, opt)
        qs = _apply_search(qs, search)
        if team_mem:
            qs = qs.filter(team_mem=team_mem)
        qs = qs.order_by("invoice_date" if order_dir == "asc" else "-invoice_date", "-id")

        page_qs, total = _paginate(qs, start, length)
        items = list(page_qs)
        _number_rows(items, start)
        serializer = AccountsApprovalListSerializer(items, many=True)
        return Response(
            {
                "draw": int(request.query_params.get("draw", 1)),
                "recordsTotal": total,
                "recordsFiltered": total,
                "data": serializer.data,
            }
        )

    @action(detail=False, methods=["get"], url_path="completed")
    def completed(self, request):
        search = request.query_params.get("search", "")
        length = int(request.query_params.get("length", 10))
        start = int(request.query_params.get("start", 0))
        from_date = request.query_params.get("from_date", "")
        to_date = request.query_params.get("to_date", "")
        opt = request.query_params.get("opt", "4")
        team_mem = request.query_params.get("team_mem", "")
        order_dir = request.query_params.get("order_dir", "asc")

        qs = InvoiceCreationMain.objects.filter(invoice_doc_status=4, is_delete=0)
        qs = _apply_company_filter(qs, request)
        qs = _apply_date_filter(qs, from_date, to_date, opt)
        qs = _apply_search(qs, search)
        if team_mem:
            qs = qs.filter(team_mem=team_mem)
        qs = qs.order_by("invoice_date" if order_dir == "asc" else "-invoice_date", "-id")

        page_qs, total = _paginate(qs, start, length)
        items = list(page_qs)
        _number_rows(items, start)
        serializer = AccountsApprovalListSerializer(items, many=True)
        return Response(
            {
                "draw": int(request.query_params.get("draw", 1)),
                "recordsTotal": total,
                "recordsFiltered": total,
                "data": serializer.data,
            }
        )

    @action(detail=False, methods=["get"], url_path="pending-list")
    def pending_list(self, request):
        search = request.query_params.get("search", "")
        length = int(request.query_params.get("length", 10))
        start = int(request.query_params.get("start", 0))
        order_dir = request.query_params.get("order_dir", "asc")

        qs = InvoiceCreationMain.objects.filter(invoice_doc_status=1, is_delete=0)
        qs = _apply_company_filter(qs, request)
        qs = _apply_search(qs, search)
        qs = qs.order_by("invoice_date" if order_dir == "asc" else "-invoice_date", "-id")

        page_qs, total = _paginate(qs, start, length)
        items = list(page_qs)
        _number_rows(items, start)
        serializer = PendingListSerializer(items, many=True)
        return Response(
            {
                "draw": int(request.query_params.get("draw", 1)),
                "recordsTotal": total,
                "recordsFiltered": total,
                "data": serializer.data,
            }
        )

    @action(detail=False, methods=["post"], url_path="approve-reject")
    def approve_reject(self, request):
        serializer = AccountsApprovalUpdateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(
                {"status": False, "error": serializer.errors, "msg": "error"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        data = serializer.validated_data
        invoice_unique_id = data["invoice_unique_id"]
        dc_number = data["dc_number"]
        ac_status = data["ac_team_verifiy_status"]
        reject_reason = data["ac_reason_reject"]
        approved_by = data["approved_by"]
        today = datetime.date.today()
        inv_doc_status = 4 if ac_status == "1" else 2

        try:
            with transaction.atomic():
                InvoiceCreationMain.objects.filter(unique_id=invoice_unique_id).update(
                    invoice_doc_status=inv_doc_status,
                    ac_team_verifiy_status=int(ac_status),
                    ac_team_approved_by=approved_by,
                    reject_reason_elcot=reject_reason,
                    approved_date=today,
                    approved_by=approved_by,
                )
                if ac_status == "2":
                    InvoiceCreationMain.objects.filter(dc_number=dc_number).update(
                        dc_number="",
                        dc_date=None,
                        invoice_no="",
                        ledger_name="",
                        ledger_no="",
                        invoice_date=None,
                    )
                    InvoiceCreation.objects.filter(dc_num=dc_number).update(
                        dc_num="",
                        dc_date=None,
                        invoice_no="",
                        ledger_name="",
                        ledger_no="",
                        invoice_date=None,
                    )
                    InvoiceSublist.objects.filter(invoice_id=invoice_unique_id).update(
                        reject_reason=reject_reason,
                        is_delete=1,
                    )
                else:
                    InvoiceSublist.objects.filter(invoice_id=invoice_unique_id).update(reject_reason=reject_reason)
                _update_invoice_verification_by_dcnumber(dc_number)
                InvoiceCreationMain.objects.filter(unique_id=invoice_unique_id).update(
                    invoice_doc_status=inv_doc_status,
                    ac_team_verifiy_status=int(ac_status),
                    ac_team_approved_by=approved_by,
                    reject_reason_elcot=reject_reason,
                    approved_date=today,
                    approved_by=approved_by,
                )
        except Exception as exc:
            return Response(
                {"status": False, "error": str(exc), "msg": "error"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response({"status": True, "msg": "update", "error": ""})

    @action(detail=False, methods=["post"], url_path="overall-approval")
    def overall_approval(self, request):
        serializer = OverallApprovalSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(
                {"status": False, "error": serializer.errors, "msg": "error"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        data = serializer.validated_data
        dc_numbers = data["dc_numbers"]
        approved_by = data["approved_by"]
        today = datetime.date.today()

        try:
            with transaction.atomic():
                InvoiceCreationMain.objects.filter(dc_number__in=dc_numbers).update(
                    approved_by=approved_by,
                    approved_date=today,
                    invoice_doc_status=4,
                )
        except Exception as exc:
            return Response(
                {"status": False, "error": str(exc), "msg": "error"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response({"status": True, "msg": "update", "error": ""})

    @action(detail=True, methods=["delete"], url_path="delete")
    def soft_delete(self, request, pk=None):
        try:
            record = InvoiceCreationMain.objects.get(unique_id=pk)
        except InvoiceCreationMain.DoesNotExist:
            return Response(
                {"status": False, "error": "Record not found", "msg": "error"},
                status=status.HTTP_404_NOT_FOUND,
            )
        
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT bill_no FROM sign_doc_verification_detail
                WHERE invoice_no = %s AND dc_number = %s AND is_delete = 0
                LIMIT 1
            """, [record.invoice_no, record.dc_number])
            bill_record = cursor.fetchone()
            
            if bill_record and bill_record[0]:  # bill_no is not empty
                return Response({
                    "status": False, 
                    "error": "Cannot delete record. Bill has already been created.",
                    "msg": "error"
                }, status=status.HTTP_400_BAD_REQUEST)
        
        with transaction.atomic():
            InvoiceCreationMain.objects.filter(unique_id=pk, is_delete=0).update(
                invoice_doc_status=0,
                ac_team_verifiy_status=0,
                ac_team_approved_by="",
                approved_by="",
                approved_date=None,
                reject_reason_elcot="",
            )
            OperationApprovalInvoiceCreationMain.objects.filter(unique_id=pk, is_delete=0).update(
                doc_approval_sts=0,
                invoice_doc_status=0,
                approved_by="",
                approved_date=None,
                reject_reason_elcot="",
            )
            InvoiceSublist.objects.filter(invoice_id=pk).update(
                doc_approval_sts=0,
                reject_reason="",
                is_delete=0,
            )
            if record.dc_number:
                InvoiceVerificationTable.objects.filter(dc_number=record.dc_number).update(
                    doc_approval_sts=0,
                    invoice_doc_status=0,
                    approved_by="",
                    approved_date=None,
                )

        return Response({"status": True, "msg": "success_delete", "error": ""})

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font
        except ImportError:
            return Response(
                {"status": False, "error": "openpyxl is not installed", "msg": "error"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        qs = InvoiceCreationMain.objects.filter(is_delete=0).exclude(invoice_doc_status=0).order_by("-invoice_doc_status", "-id")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Docs Approval AC"
        ws.merge_cells("A1:L1")
        ws["A1"] = "Docs Approval AC"
        ws["A1"].font = Font(bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = [
            "S.No",
            "PO No",
            "PO Date",
            "Consignee",
            "Location",
            "DC No",
            "DC Date",
            "Invoice No",
            "Invoice Date",
            "Status",
            "Approved By",
            "Invoice Value",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)

        for i, row in enumerate(qs, start=1):
            po = PurchaseOrder.objects.filter(unique_id=row.form_main_unique_id, is_delete=0).first()
            consignee = PurchaseOrderConsignee.objects.filter(unique_id=row.consignee_unique_id, is_delete=0).first()
            staff = UserCreation.objects.filter(staff_id=row.approved_by, is_delete=0).first() if row.approved_by else None
            po_date = row.invoice_date.strftime("%d-%m-%Y") if row.invoice_date else ""
            dc_date = row.dc_date.strftime("%d-%m-%Y") if row.dc_date else ""
            invoice_date = row.invoice_date.strftime("%d-%m-%Y") if row.invoice_date else ""
            ws.append(
                [
                    i,
                    po.po_num if po else "",
                    po.po_date.strftime("%d-%m-%Y") if po and po.po_date else po_date,
                    consignee.con_contact_name if consignee else "",
                    consignee.con_address if consignee else "",
                    row.dc_number or "",
                    dc_date,
                    row.invoice_no or "",
                    invoice_date,
                    _status_label(row.invoice_doc_status),
                    staff.staff_name if staff else (row.approved_by or ""),
                    _indian_money(row.invoice_value),
                ]
            )

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="Docs Approval AC.xlsx"'
        return response


def _update_invoice_verification_by_dcnumber(dc_number: str) -> None:
    from django.db import connection

    if not dc_number:
        return

    with connection.cursor() as cursor:
        cursor.callproc("update_invoice_verification_by_dcnumber", [dc_number])
        while cursor.nextset():
            pass
